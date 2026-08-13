"""安全 Excel Writer（规范 §43/§44）。

只有 QA PASS 且非 dry-run 才允许写正式 Master。写入流程：
    正式 Master -> 复制到 temp -> 更新 temp -> 完整验证 -> 关闭 -> 替换正式 Master
任一步失败则正式 Master 保持原样。
"""
from __future__ import annotations

import csv
import datetime as dt
import logging
import os
import shutil
from pathlib import Path
from typing import Any

import openpyxl

from ..services.review import REVIEW_HEADERS
from .reader import ES_MAP, ZH_MAP

log = logging.getLogger(__name__)

RUN_LOG_HEADERS = ["Run ID", "运行日期", "开始时间", "结束时间", "Git Commit", "Sitemap SKU数", "Listing SKU数",
                   "ACTIVE", "NEW", "REAPPEARED", "MISSING_FIRST", "MISSING_CONTINUED", "OFFLINE",
                   "PRICE_UP", "PRICE_DOWN", "PROMO_START", "PROMO_END", "NEW_BADGE_ON", "NEW_BADGE_OFF",
                   "CONTENT_CHANGE", "异常数量", "QA状态", "运行状态", "备注"]

# 03_PRICE_HISTORY / 04_EVENT_HISTORY 表头（写行时用常量，行 dict 的 key 必须与之一致）
PRICE_HISTORY_HEADERS = ["Canonical_ID", "SKU", "日期", "旧售价 (€)", "新售价 (€)", "原价 (€)",
                         "变化类型", "变化金额 (€)", "变化幅度 (%)", "促销状态", "来源文件", "来源Sheet"]
EVENT_HISTORY_HEADERS = ["Canonical_ID", "SKU", "日期", "事件类型", "旧值", "新值", "来源文件", "备注"]


APRIL_ARCHIVE_HEADERS = [
    "四月归档ID", "正式SKU", "观测日期", "一级类目（中文）", "一级类目（西语）",
    "中文品名", "西班牙语品名", "历史售价 (€)", "规格（中文）", "规格（西语）",
    "单价（中文）", "单价（西语）", "促销（中文）", "促销（西语）",
    "字段有效范围", "匹配状态", "西语来源文件", "中文来源文件", "来源Sheet", "备注",
]

LONG_TERM_MASTER_HEADERS = [
    "实体ID", "正式SKU", "四月归档ID", "身份类型", "匹配状态", "匹配置信度", "当前状态",
    "中文品名", "西班牙语品名", "一级类目（中文）", "一级类目（西语）", "规格（中文）", "规格（西语）",
    "当前售价 (€)", "历史最低价 (€)", "历史最高价 (€)", "首次观察日期", "最后观察日期",
    "四月原始记录数", "四月归档ID集合", "来源数", "来源工作表", "商品链接", "核对备注",
]
APRIL_MATCH_AUDIT_HEADERS = [
    "四月归档ID", "正式SKU", "匹配状态", "匹配方法", "置信度", "候选SKU", "证据说明",
    "观察日期", "一级类目（中文）", "一级类目（西语）", "中文品名", "西班牙语品名",
    "历史售价 (€)", "规格（中文）", "规格（西语）", "单价（中文）", "单价（西语）",
    "促销（中文）", "促销（西语）", "来源行号", "西语来源文件", "中文来源文件",
]
SOURCE_SCHEMA_HEADERS = [
    "日期", "文件名", "Sheet", "Raw 行数", "Raw 列数", "真实 Raw Schema", "来源作用", "数据状态", "备注",
]


def _backup(master: Path, backups_dir: Path) -> Path:
    backups_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    target = backups_dir / f"Action_Master_{stamp}.xlsx"
    shutil.copy2(master, target)
    log.info("已备份 Master -> %s", target)
    return target


def _cell(value: Any) -> Any:
    """把内部值转成 Excel 单元格值：日期真日期、价格真数值、布尔转是/否。"""
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    if not s:
        return None
    # 尝试把 "YYYY-MM-DD" 转真日期（规范 §53）
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return s


def _ensure_sheet(wb, title: str, headers: list[str]) -> None:
    if title in wb.sheetnames:
        return
    ws = wb.create_sheet(title=title)
    ws.append(headers)


def _append_rows(wb, title: str, rows: list[dict], headers: list[str]) -> None:
    if not rows:
        return
    _ensure_sheet(wb, title, headers)  # 已有表则不重写表头（修复重复表头 bug）
    ws = wb[title]
    for r in rows:
        ws.append([_cell(r.get(h)) for h in headers])


def _update_or_append_current(wb, sheet: str, colmap: dict, key_to_records: dict[str, dict], internal_keys: set[str]) -> int:
    """按 header 映射更新既有行，新 SKU 追加。返回更新的行数。"""
    if sheet not in wb.sheetnames:
        return 0
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    idx = {}
    for i, h in enumerate(header):
        k = colmap.get(h)
        if k:
            idx[k] = i
    # CURRENT is not a history table. Rows absent from the authoritative
    # Presence dataset are removed; lifecycle evidence remains in state CSVs.
    wanted_skus = {str(key) for key in key_to_records}
    wanted_cids = {
        str(rec.get("canonical_id")) for rec in key_to_records.values()
        if rec.get("canonical_id")
    }
    for row_no in range(ws.max_row, 1, -1):
        sku = str(ws.cell(row=row_no, column=2).value or "")
        cid = str(ws.cell(row=row_no, column=1).value or "")
        if sku not in wanted_skus and cid not in wanted_cids:
            ws.delete_rows(row_no, 1)
    # 更新既有行
    updated = 0
    for row in ws.iter_rows(min_row=2):
        sku = row[1].value
        cid = row[0].value
        rec = key_to_records.get(str(sku)) or key_to_records.get(str(cid))
        if not rec:
            continue
        for k, col in idx.items():
            if k in internal_keys and k in rec:
                row[col].value = _cell(rec.get(k))
        updated += 1
    # 追加新 SKU
    existing = {str(row[1]) for row in ws.iter_rows(min_row=2, values_only=True) if row[1]}
    for sku, rec in key_to_records.items():
        if sku in existing:
            continue
        new_row = [_cell(rec.get(colmap[h])) for h in header]
        ws.append(new_row)
        updated += 1
    return updated


def _refresh_long_term_catalog(wb) -> None:
    """Refresh current fields without deleting historical long-term entities."""
    if "08_LONG_TERM_MASTER" not in wb.sheetnames:
        return
    catalog = wb["08_LONG_TERM_MASTER"]
    header = [cell.value for cell in catalog[7]]
    if header[:len(LONG_TERM_MASTER_HEADERS)] != LONG_TERM_MASTER_HEADERS:
        raise RuntimeError("08_LONG_TERM_MASTER header mismatch")
    idx = {name: i + 1 for i, name in enumerate(header)}

    def current_records(sheet_name: str) -> dict[str, dict[str, Any]]:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        records: dict[str, dict[str, Any]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            sku = str(rec.get("SKU") or "").strip()
            if sku:
                records[sku] = rec
        return records

    zh = current_records("01_SKU_ZH_CURRENT")
    es = current_records("02_SKU_ES_CURRENT")
    current_skus = set(zh) | set(es)
    catalog_rows: dict[str, int] = {}
    for row_no in range(8, catalog.max_row + 1):
        if str(catalog.cell(row_no, idx["身份类型"]).value or "").strip() != "OFFICIAL_SKU":
            continue
        sku = str(catalog.cell(row_no, idx["正式SKU"]).value or "").strip()
        if sku:
            catalog_rows[sku] = row_no

    def set_if_present(row_no: int, column: str, value: Any) -> None:
        if value not in (None, ""):
            catalog.cell(row_no, idx[column]).value = _cell(value)

    for sku, row_no in catalog_rows.items():
        catalog.cell(row_no, idx["当前状态"]).value = "CURRENT" if sku in current_skus else "HISTORICAL"
        if sku not in current_skus:
            catalog.cell(row_no, idx["当前售价 (€)"]).value = None

    for sku in sorted(current_skus):
        zr, er = zh.get(sku, {}), es.get(sku, {})
        row_no = catalog_rows.get(sku)
        if row_no is None:
            row_no = catalog.max_row + 1
            canonical = er.get("Canonical_ID") or zr.get("Canonical_ID")
            catalog.cell(row_no, idx["实体ID"]).value = canonical or f"SKU:{sku}"
            catalog.cell(row_no, idx["正式SKU"]).value = sku
            catalog.cell(row_no, idx["身份类型"]).value = "OFFICIAL_SKU"
            catalog.cell(row_no, idx["匹配状态"]).value = "OFFICIAL_IDENTITY"
            catalog.cell(row_no, idx["匹配置信度"]).value = 1.0
            catalog.cell(row_no, idx["四月原始记录数"]).value = 0
            catalog_rows[sku] = row_no
        catalog.cell(row_no, idx["当前状态"]).value = "CURRENT"
        set_if_present(row_no, "中文品名", zr.get("中文品名"))
        set_if_present(row_no, "西班牙语品名", er.get("西班牙语品名") or zr.get("西班牙语品名"))
        set_if_present(row_no, "一级类目（中文）", zr.get("一级类目（中文）"))
        set_if_present(row_no, "一级类目（西语）", er.get("一级类目（西语）") or zr.get("一级类目（西语）"))
        set_if_present(row_no, "规格（中文）", zr.get("规格（中文）"))
        set_if_present(row_no, "规格（西语）", er.get("规格（西语）") or zr.get("规格（西语）"))
        set_if_present(row_no, "当前售价 (€)", er.get("当前售价 (€)") or zr.get("当前售价 (€)"))
        set_if_present(row_no, "首次观察日期", er.get("首次发现日期") or zr.get("首次发现日期"))
        set_if_present(row_no, "最后观察日期", er.get("最后确认存在日期") or zr.get("最后确认存在日期"))
        set_if_present(row_no, "商品链接", er.get("商品链接") or zr.get("商品链接"))

    pending_count = sum(
        1 for row_no in range(8, catalog.max_row + 1)
        if str(catalog.cell(row_no, idx["身份类型"]).value or "").strip() == "APRIL_ARCHIVE_PENDING"
    )
    catalog["B4"] = len(catalog_rows)
    catalog["D4"] = len(current_skus)
    catalog["H4"] = pending_count
    catalog["B5"] = len(catalog_rows) + pending_count


def _migrate_legacy_sheets(wb, cfg: dict[str, Any]) -> None:
    """归档并删除旧版 05_MAPPING_REVIEW / 06_SOURCE_SUMMARY（幂等，行数不符即中止）。

    历史 Master 的这两个表不是规范 §十 的 05_RUN_LOG / 06_REVIEW_QUEUE：
      - 05_MAPPING_REVIEW 是历史人工核对审计数据（970 行）-> runtime/state/legacy_mapping_review.csv
      - 06_SOURCE_SUMMARY 是历史导入源文件汇总（17 行）-> runtime/backups/legacy_source_summary.csv
    删除前必须验证归档行数与表内一致，否则抛异常保护原文件；已归档且行数一致则直接删表。
    """
    paths_cfg = cfg.get("paths") or {}
    state_dir: Path = paths_cfg.get("state", Path("runtime/state"))
    backups_dir: Path = paths_cfg.get("backups", Path("runtime/backups"))
    for sheet, target, label in (
        ("05_MAPPING_REVIEW", state_dir / "legacy_mapping_review.csv", "MAPPING_REVIEW"),
        ("06_SOURCE_SUMMARY", backups_dir / "legacy_source_summary.csv", "SOURCE_SUMMARY"),
    ):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        header = list(all_rows[0]) if all_rows else []
        data = [r for r in all_rows[1:] if any(c is not None for c in r)]
        if target.exists():
            with target.open("r", encoding="utf-8-sig", newline="") as f:
                n_existing = sum(1 for _ in csv.reader(f)) - 1
            if n_existing == len(data):
                del wb[sheet]
                log.info("归档已存在（%s %d 行），直接删除原表", label, n_existing)
                continue
            log.warning("归档文件已存在但行数不符（%s: 已存 %d, 待存 %d），覆盖", target.name, n_existing, len(data))
        target.parent.mkdir(parents=True, exist_ok=True)
        with target.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(data)
        with target.open("r", encoding="utf-8-sig", newline="") as f:
            n_written = sum(1 for _ in csv.reader(f)) - 1
        if n_written != len(data):
            raise RuntimeError(f"{label} 归档行数不符: sheet={len(data)} csv={n_written}，中止以保护原文件")
        del wb[sheet]
        log.info("已归档 %s（%d 行）-> %s，并删除原表", label, len(data), target)


def stage_master(
    cfg: dict[str, Any],
    *,
    updated_records: dict[str, dict],
    price_events: list[dict],
    event_events: list[dict],
    run_log_row: dict | None = None,
    review_rows: list[dict] | None = None,
) -> Path:
    """暂存新的 Master 到 temp：备份 → 复制 → 更新 → 保存 → 完整验证。

    返回已就绪的 temp 路径（尚未替换正式文件），供"Master + 状态文件一起
    先生成再统一提交"使用。只有 commit_master 才会替换正式 Master。
    """
    master: Path = cfg["paths"]["master"]
    if not master.exists():
        raise FileNotFoundError(f"Master 不存在: {master}")

    _backup(master, cfg["paths"]["backups"])
    tmp: Path = cfg["paths"]["temp"] / master.name
    tmp.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(master, tmp)

    wb = openpyxl.load_workbook(tmp)
    # 迁移旧版 05/06 审计表（幂等），为新 05/06 腾出表名
    _migrate_legacy_sheets(wb, cfg)

    # 01 / 02 更新
    n_zh = _update_or_append_current(wb, "01_SKU_ZH_CURRENT", ZH_MAP, updated_records, set(ZH_MAP.values()))
    n_es = _update_or_append_current(wb, "02_SKU_ES_CURRENT", ES_MAP, updated_records, set(ES_MAP.values()))
    log.info("01 更新 %d 行, 02 更新 %d 行", n_zh, n_es)
    _refresh_long_term_catalog(wb)

    # 03 / 04 追加（表头用常量，行 dict 的 key 与之对齐）
    _append_rows(wb, "03_PRICE_HISTORY", price_events, PRICE_HISTORY_HEADERS)
    _append_rows(wb, "04_EVENT_HISTORY", event_events, EVENT_HISTORY_HEADERS)

    # 05_RUN_LOG / 06_REVIEW_QUEUE（规范 §十；不存在则创建，无行也要有表）
    _ensure_sheet(wb, "05_RUN_LOG", RUN_LOG_HEADERS)
    if run_log_row:
        _append_rows(wb, "05_RUN_LOG", [run_log_row], RUN_LOG_HEADERS)
    _ensure_sheet(wb, "06_REVIEW_QUEUE", REVIEW_HEADERS)
    if review_rows:
        _append_rows(wb, "06_REVIEW_QUEUE", review_rows, REVIEW_HEADERS)

    wb.save(tmp)
    wb.close()

    # 完整验证
    _validate(tmp)
    return tmp


def commit_master(tmp: Path, master: Path) -> Path:
    """原子替换正式 Master（tmp 必须已通过 _validate）。"""
    os.replace(tmp, master)
    log.info("正式 Master 已更新: %s", master)
    return master


def write_master(
    cfg: dict[str, Any],
    *,
    updated_records: dict[str, dict],
    price_events: list[dict],
    event_events: list[dict],
    run_log_row: dict | None = None,
    review_rows: list[dict] | None = None,
    dry_run: bool = False,
) -> Path:
    """原子更新正式 Master（stage + commit 一步完成）。dry_run 时禁止调用。"""
    if dry_run:
        raise RuntimeError("dry-run 禁止写 Master")
    tmp = stage_master(cfg, updated_records=updated_records, price_events=price_events,
                       event_events=event_events, run_log_row=run_log_row, review_rows=review_rows)
    return commit_master(tmp, cfg["paths"]["master"])


def _validate_april_archive(ws) -> None:
    """Protect the April archive from being treated as official lifecycle SKUs."""
    headers = [cell.value for cell in ws[1]]
    if headers != APRIL_ARCHIVE_HEADERS:
        raise RuntimeError("07_APRIL_ARCHIVE header mismatch")
    seen: set[str] = set()
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        archive_id = str(row[0] or "").strip()
        if not archive_id:
            continue
        if archive_id in seen:
            raise RuntimeError(f"07_APRIL_ARCHIVE duplicate archive ID at row {row_no}: {archive_id}")
        seen.add(archive_id)
        if row[1] not in (None, ""):
            raise RuntimeError(f"07_APRIL_ARCHIVE official SKU must remain blank at row {row_no}")
        if not isinstance(row[2], (dt.date, dt.datetime)):
            raise RuntimeError(f"07_APRIL_ARCHIVE observation date is not typed at row {row_no}")
        if not isinstance(row[7], (int, float)):
            raise RuntimeError(f"07_APRIL_ARCHIVE historical price is not numeric at row {row_no}")


def _validate_long_term_sheets(wb) -> None:
    """Protect the long-lived catalog and its April identity audit."""
    required = {
        "08_LONG_TERM_MASTER": LONG_TERM_MASTER_HEADERS,
        "09_APRIL_MATCH_AUDIT": APRIL_MATCH_AUDIT_HEADERS,
        "10_SOURCE_SCHEMA": SOURCE_SCHEMA_HEADERS,
    }
    present = set(required) & set(wb.sheetnames)
    if present and present != set(required):
        missing = set(required) - present
        raise RuntimeError(f"long-term workbook sheets are incomplete: {sorted(missing)}")
    if not present:
        return
    for name, expected_headers in required.items():
        ws = wb[name]
        header_row = 7 if name == "08_LONG_TERM_MASTER" else 1
        headers = [cell.value for cell in ws[header_row]][:len(expected_headers)]
        if headers != expected_headers:
            raise RuntimeError(f"{name} header mismatch")

    catalog = wb["08_LONG_TERM_MASTER"]
    entity_ids: set[str] = set()
    official_skus: set[str] = set()
    current_count = 0
    for row_no, row in enumerate(catalog.iter_rows(min_row=8, values_only=True), 8):
        entity_id = str(row[0] or "").strip()
        if not entity_id:
            continue
        if entity_id in entity_ids:
            raise RuntimeError(f"08_LONG_TERM_MASTER duplicate entity ID at row {row_no}: {entity_id}")
        entity_ids.add(entity_id)
        identity_type = str(row[3] or "").strip()
        sku = str(row[1] or "").strip()
        if identity_type == "OFFICIAL_SKU":
            if not sku:
                raise RuntimeError(f"08_LONG_TERM_MASTER official entity without SKU at row {row_no}")
            if sku in official_skus:
                raise RuntimeError(f"08_LONG_TERM_MASTER duplicate official SKU at row {row_no}: {sku}")
            official_skus.add(sku)
        elif identity_type == "APRIL_ARCHIVE_PENDING" and sku:
            raise RuntimeError(f"08_LONG_TERM_MASTER pending April entity must not have official SKU at row {row_no}")
        if str(row[6] or "").strip() == "CURRENT":
            current_count += 1

    current_sheet_skus = {
        str(row[1] or "").strip()
        for row in wb["02_SKU_ES_CURRENT"].iter_rows(min_row=2, values_only=True)
        if row[1]
    }
    if current_count != len(current_sheet_skus):
        raise RuntimeError(
            "08_LONG_TERM_MASTER CURRENT count mismatch: "
            f"catalog={current_count} current_sheet={len(current_sheet_skus)}"
        )

    audit = wb["09_APRIL_MATCH_AUDIT"]
    archive_ids: set[str] = set()
    for row_no, row in enumerate(audit.iter_rows(min_row=2, values_only=True), 2):
        archive_id = str(row[0] or "").strip()
        if not archive_id:
            continue
        if archive_id in archive_ids:
            raise RuntimeError(f"09_APRIL_MATCH_AUDIT duplicate archive ID at row {row_no}: {archive_id}")
        archive_ids.add(archive_id)
    raw_archive_ids = {
        str(row[0] or "").strip()
        for row in wb["07_APRIL_ARCHIVE"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    if archive_ids != raw_archive_ids:
        raise RuntimeError("09_APRIL_MATCH_AUDIT archive ID set does not match 07_APRIL_ARCHIVE")


def _validate(path: Path) -> None:
    """重开文件验证可读、各表存在、无空数据区错误。失败抛异常以保护原文件。"""
    wb = openpyxl.load_workbook(path, read_only=True)
    required = {"01_SKU_ZH_CURRENT", "02_SKU_ES_CURRENT", "03_PRICE_HISTORY", "04_EVENT_HISTORY",
                "05_RUN_LOG", "06_REVIEW_QUEUE"}
    missing = required - set(wb.sheetnames)
    if not missing and "07_APRIL_ARCHIVE" in wb.sheetnames:
        _validate_april_archive(wb["07_APRIL_ARCHIVE"])
    if not missing:
        _validate_long_term_sheets(wb)
    wb.close()
    if missing:
        raise RuntimeError(f"验证失败，缺少 Sheet: {missing}")
    log.info("验证通过: %s", path)
