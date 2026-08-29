"""Lossless Excel Master -> SQLite staging migration for Data Foundation V1.

The importer is deliberately independent from the daily orchestrator.  It
opens the production workbook read-only, maps only evidenced formal SKUs, and
records unmatched/duplicate source rows as auditable issues rather than
silently repairing them.
"""
from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .connection import connect, transaction
from .schema import SCHEMA_VERSION, migrate


SHEET_CONFIG = {
    "01_SKU_ZH_CURRENT": 1,
    "02_SKU_ES_CURRENT": 1,
    "03_PRICE_HISTORY": 1,
    "04_EVENT_HISTORY": 1,
    "05_RUN_LOG": 1,
    "06_REVIEW_QUEUE": 1,
    "07_APRIL_ARCHIVE": 1,
    "08_LONG_TERM_MASTER": 7,
    "09_APRIL_MATCH_AUDIT": 1,
    "10_SOURCE_SCHEMA": 1,
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _iso_value(value: Any, header: str = "") -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and any(word in header for word in ("日期", "时间", "Date", "date")):
        if 20000 <= float(value) <= 80000:
            try:
                converted = from_excel(value)
                return converted.isoformat(timespec="seconds") if isinstance(converted, datetime) else converted.isoformat()
            except Exception:
                pass
    return value


def _text(value: Any, header: str = "") -> str:
    value = _iso_value(value, header)
    if value is None:
        return ""
    return str(value).strip()


def _json_value(value: Any, header: str = "") -> Any:
    value = _iso_value(value, header)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _row_to_dict(headers: list[str], values: Iterable[Any]) -> dict[str, Any]:
    return {header: _json_value(value, header) for header, value in zip(headers, values)}


def _headers(ws, header_row: int) -> list[str]:
    values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [_text(value) for value in values]
    # openpyxl may expose a trailing blank cell when a data row is wider than
    # the header row.  Ignore only such trailing blanks; interior blanks remain
    # a hard source-schema error.
    while headers and not headers[-1]:
        headers.pop()
    if not headers or any(not header for header in headers):
        raise ValueError(f"Invalid blank header in sheet {ws.title}")
    return headers


def _records(ws, header_row: int):
    headers = _headers(ws, header_row)
    for row_no, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if all(value is None for value in values):
            continue
        yield row_no, headers, list(values)


def _map(headers: list[str], label: str) -> int:
    try:
        return headers.index(label)
    except ValueError as exc:
        raise ValueError(f"Missing required column {label!r}; got {headers!r}") from exc


def _optional_text(headers: list[str], values: list[Any], label: str, header: str = "") -> str:
    """Read an optional source column without inventing a value."""
    if label not in headers:
        return ""
    return _text(values[headers.index(label)], header or label)


def _maybe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).replace("€", "").replace(" ", "").replace("\u00a0", "")
    match = re.search(r"-?\d+(?:[.,]\d+)?", text)
    if not match:
        return None
    token = match.group(0)
    if token.count(",") == 1 and token.count(".") == 0:
        token = token.replace(",", ".")
    elif token.count(",") > 0 and token.count(".") > 0:
        token = token.replace(",", "")
    try:
        return float(token)
    except ValueError:
        return None


def _source_issue(db, migration_id: str, sheet: str, row_no: int, code: str, source_key: str, raw: dict[str, Any]):
    db.execute(
        "INSERT INTO migration_source_issues(migration_id,source_sheet,source_row_no,issue_code,source_key,raw_json) VALUES (?,?,?,?,?,?)",
        (migration_id, sheet, row_no, code, source_key, json.dumps(raw, ensure_ascii=False, default=str)),
    )


def migrate_master(master_path: Path, staging_path: Path, migration_id: str) -> dict[str, Any]:
    """Create a new staging DB from a read-only workbook."""
    if not master_path.exists():
        raise FileNotFoundError(master_path)
    source_hash = sha256_file(master_path)
    staging_path.parent.mkdir(parents=True, exist_ok=True)
    if staging_path.exists():
        staging_path.unlink()
    wb = load_workbook(master_path, read_only=True, data_only=True)
    conn = None
    try:
        missing_sheets = sorted(set(SHEET_CONFIG) - set(wb.sheetnames))
        if missing_sheets:
            raise ValueError(f"Missing required Master sheets: {', '.join(missing_sheets)}")
        conn = connect(staging_path)
        migrate(staging_path)
        counts = {name: 0 for name in ("products", "localizations", "observations", "price_history", "events", "runs", "reviews", "source_issues")}
        with transaction(conn) as db:
            db.execute("INSERT OR REPLACE INTO schema_metadata(key,value) VALUES (?,?)", ("schema_version", SCHEMA_VERSION))
            db.execute("INSERT OR REPLACE INTO schema_metadata(key,value) VALUES (?,?)", ("master_path", str(master_path)))
            db.execute("INSERT OR REPLACE INTO schema_metadata(key,value) VALUES (?,?)", ("master_sha256", source_hash))
            db.execute("INSERT OR REPLACE INTO schema_metadata(key,value) VALUES (?,?)", ("migration_id", migration_id))

            # Long-term products are the identity source.
            ws = wb["08_LONG_TERM_MASTER"]
            for row_no, headers, values in _records(ws, 7):
                raw = _row_to_dict(headers, values)
                sku = _text(values[_map(headers, "正式SKU")], "正式SKU")
                entity_id = _text(values[_map(headers, "实体ID")], "实体ID")
                if not sku:
                    _source_issue(db, migration_id, ws.title, row_no, "UNMATCHED_CANONICAL", entity_id, raw)
                    counts["source_issues"] += 1
                    continue
                status = _text(values[_map(headers, "当前状态")], "当前状态")
                db.execute(
                    """INSERT INTO products(sku,canonical_id,name_es,cat1_es,cat2_es,spec_es,product_url,current_price,historical_min_price,historical_max_price,current_status_raw,first_seen_at,last_seen_at,source_sheet,source_row_no,source_raw_json)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        sku, entity_id, _text(values[_map(headers, "西班牙语品名")]),
                        _optional_text(headers, values, "一级类目（西语）"), _optional_text(headers, values, "二级类目（西语）"),
                        _text(values[_map(headers, "规格（西语）")]), _text(values[_map(headers, "商品链接")]),
                        _maybe_float(values[_map(headers, "当前售价 (€)")]), _maybe_float(values[_map(headers, "历史最低价 (€)")]),
                        _maybe_float(values[_map(headers, "历史最高价 (€)")]), status,
                        _text(values[_map(headers, "首次观察日期")], "首次观察日期"), _text(values[_map(headers, "最后观察日期")], "最后观察日期"),
                        ws.title, row_no, json.dumps(raw, ensure_ascii=False, default=str),
                    ),
                )
                counts["products"] += 1
                zh = {
                    "name": _text(values[_map(headers, "中文品名")]),
                    "cat1": _optional_text(headers, values, "一级类目（中文）"),
                    "cat2": _optional_text(headers, values, "二级类目（中文）"),
                    "spec": _text(values[_map(headers, "规格（中文）")]),
                }
                db.execute(
                    """INSERT INTO product_localizations(sku,language,name,cat1,cat2,spec,source,review_status,source_sheet,source_row_no)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(sku,language) DO UPDATE SET name=excluded.name,cat1=excluded.cat1,cat2=excluded.cat2,spec=excluded.spec,source=excluded.source,source_sheet=excluded.source_sheet,source_row_no=excluded.source_row_no,updated_at=CURRENT_TIMESTAMP""",
                    (sku, "zh", zh["name"], zh["cat1"], zh["cat2"], zh["spec"], "MASTER_08_LONG_TERM_MASTER", "LEGACY_UNVERIFIED", ws.title, row_no),
                )
                counts["localizations"] += 1

            # Current Chinese view enriches localization descriptions and current values.
            ws = wb["01_SKU_ZH_CURRENT"]
            for row_no, headers, values in _records(ws, 1):
                raw = _row_to_dict(headers, values)
                sku = _text(values[_map(headers, "SKU")], "SKU")
                if not sku:
                    _source_issue(db, migration_id, ws.title, row_no, "EMPTY_SKU", "", raw)
                    counts["source_issues"] += 1
                    continue
                if db.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone() is None:
                    _source_issue(db, migration_id, ws.title, row_no, "ORPHAN_CURRENT_ZH", sku, raw)
                    counts["source_issues"] += 1
                    continue
                db.execute(
                    """INSERT INTO product_localizations(sku,language,name,cat1,cat2,spec,description,details,source,review_status,source_sheet,source_row_no)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(sku,language) DO UPDATE SET name=excluded.name,cat1=excluded.cat1,cat2=excluded.cat2,spec=excluded.spec,description=excluded.description,details=excluded.details,source=excluded.source,review_status=excluded.review_status,source_sheet=excluded.source_sheet,source_row_no=excluded.source_row_no,updated_at=CURRENT_TIMESTAMP""",
                    (
                        sku, "zh", _text(values[_map(headers, "中文品名")]), _text(values[_map(headers, "一级类目（中文）")]),
                        _text(values[_map(headers, "二级类目（中文）")]), _text(values[_map(headers, "规格（中文）")]),
                        _text(values[_map(headers, "中文描述")]), _text(values[_map(headers, "中文产品详情")]),
                        "MASTER_01_SKU_ZH_CURRENT", _text(values[_map(headers, "翻译状态")]) or "LEGACY_UNVERIFIED", ws.title, row_no,
                    ),
                )

            # Current Spanish details are facts; update product fields without rebuilding lifecycle state.
            ws = wb["02_SKU_ES_CURRENT"]
            for row_no, headers, values in _records(ws, 1):
                raw = _row_to_dict(headers, values)
                sku = _text(values[_map(headers, "SKU")], "SKU")
                if not sku:
                    _source_issue(db, migration_id, ws.title, row_no, "EMPTY_SKU", "", raw)
                    counts["source_issues"] += 1
                    continue
                if db.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone() is None:
                    _source_issue(db, migration_id, ws.title, row_no, "ORPHAN_CURRENT_ES", sku, raw)
                    counts["source_issues"] += 1
                    continue
                db.execute("UPDATE products SET description_es=?,details_es=?,image_url=?,current_price=?,product_url=?,updated_at=CURRENT_TIMESTAMP WHERE sku=?", (
                    _text(values[_map(headers, "描述（西语）")]), _text(values[_map(headers, "产品详情（西语）")]),
                    _text(values[_map(headers, "图片链接")]), _maybe_float(values[_map(headers, "当前售价 (€)")]),
                    _text(values[_map(headers, "商品链接")]), sku,
                ))

            # Price history, keeping unmatched canonical rows only as audit issues.
            ws = wb["03_PRICE_HISTORY"]
            for row_no, headers, values in _records(ws, 1):
                raw = _row_to_dict(headers, values)
                sku = _text(values[_map(headers, "SKU")], "SKU")
                source_key = _text(values[_map(headers, "Canonical_ID")])
                if not sku:
                    _source_issue(db, migration_id, ws.title, row_no, "UNMATCHED_CANONICAL", source_key, raw)
                    counts["source_issues"] += 1
                    continue
                if db.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone() is None:
                    _source_issue(db, migration_id, ws.title, row_no, "ORPHAN_HISTORY", sku, raw)
                    counts["source_issues"] += 1
                    continue
                db.execute("INSERT INTO price_history(sku,canonical_id,observed_at,previous_price,new_price,original_price,unit_price_raw,change_type,promotion_raw,raw_json,source_file,source_sheet) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", (
                    sku, source_key, _text(values[_map(headers, "日期")], "日期"), _maybe_float(values[_map(headers, "旧售价 (€)")]),
                    _maybe_float(values[_map(headers, "新售价 (€)")]), _maybe_float(values[_map(headers, "原价 (€)")]),
                    None, _text(values[_map(headers, "变化类型")]), _text(values[_map(headers, "促销状态")]),
                    json.dumps(raw, ensure_ascii=False, default=str), _text(values[_map(headers, "来源文件")]), _text(values[_map(headers, "来源Sheet")]),
                ))
                counts["price_history"] += 1

            # Event history.
            ws = wb["04_EVENT_HISTORY"]
            for row_no, headers, values in _records(ws, 1):
                raw = _row_to_dict(headers, values)
                sku = _text(values[_map(headers, "SKU")], "SKU")
                source_key = _text(values[_map(headers, "Canonical_ID")])
                if not sku:
                    _source_issue(db, migration_id, ws.title, row_no, "UNMATCHED_CANONICAL", source_key, raw)
                    counts["source_issues"] += 1
                    continue
                if db.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone() is None:
                    _source_issue(db, migration_id, ws.title, row_no, "ORPHAN_EVENT", sku, raw)
                    counts["source_issues"] += 1
                    continue
                db.execute("INSERT INTO events(sku,canonical_id,occurred_at,event_type,old_value,new_value,evidence,source_file,source_sheet) VALUES(?,?,?,?,?,?,?,?,?)", (
                    sku, source_key, _text(values[_map(headers, "日期")], "日期"), _text(values[_map(headers, "事件类型")]),
                    _text(values[_map(headers, "旧值")]), _text(values[_map(headers, "新值")]), _text(values[_map(headers, "备注")]),
                    _text(values[_map(headers, "来源文件")]), "04_EVENT_HISTORY",
                ))
                counts["events"] += 1

            # Run log.
            ws = wb["05_RUN_LOG"]
            for row_no, headers, values in _records(ws, 1):
                raw = _row_to_dict(headers, values)
                run_id = _text(values[_map(headers, "Run ID")])
                db.execute("INSERT INTO runs(run_id,run_date,started_at,finished_at,git_commit,sitemap_count,listing_count,current_count,new_count,reappeared_count,missing_first_count,missing_continued_count,offline_count,price_up_count,price_down_count,promo_start_count,promo_end_count,access_state,observation_complete,qa_status,commit_status,snapshot_path,source_row_no,source_raw_json) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (
                    run_id, _text(values[_map(headers, "运行日期")], "运行日期"), _text(values[_map(headers, "开始时间")], "开始时间"), _text(values[_map(headers, "结束时间")], "结束时间"),
                    _text(values[_map(headers, "Git Commit")]), *[values[_map(headers, label)] for label in ("Sitemap SKU数", "Listing SKU数", "ACTIVE", "NEW", "REAPPEARED", "MISSING_FIRST", "MISSING_CONTINUED", "OFFLINE", "PRICE_UP", "PRICE_DOWN", "PROMO_START", "PROMO_END")],
                    None, None, _text(values[_map(headers, "QA状态")]), _text(values[_map(headers, "运行状态")]), None, row_no, json.dumps(raw, ensure_ascii=False, default=str),
                ))
                counts["runs"] += 1

            # Master review queue is kept as its own audit semantic.
            ws = wb["06_REVIEW_QUEUE"]
            # Keep the inventory's declared candidate key: SKU + issue type +
            # evidence.  Date is deliberately not part of this source duplicate
            # detector; duplicate rows must remain in reviews regardless.
            review_keys: set[tuple[str, str, str]] = set()
            for row_no, headers, values in _records(ws, 1):
                raw = _row_to_dict(headers, values)
                sku = _text(values[_map(headers, "SKU")], "SKU") or None
                review_key = tuple(_text(values[_map(headers, label)]) for label in ("SKU", "问题类型", "证据"))
                if review_key in review_keys:
                    _source_issue(db, migration_id, ws.title, row_no, "SOURCE_DUPLICATE", "¦".join(review_key), raw)
                    counts["source_issues"] += 1
                review_keys.add(review_key)
                if sku and db.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone() is None:
                    _source_issue(db, migration_id, ws.title, row_no, "ORPHAN_REVIEW", sku, raw)
                    counts["source_issues"] += 1
                    sku = None
                review_id = f"MASTER06:{row_no}"
                db.execute("INSERT INTO reviews(review_id,sku,review_date,issue_type,evidence,candidate_value,confidence,suggested_action,manual_note,source_row_no,source_sheet) VALUES(?,?,?,?,?,?,?,?,?,?,?)", (
                    review_id, sku, _text(values[_map(headers, "日期")], "日期"), _text(values[_map(headers, "问题类型")]), _text(values[_map(headers, "证据")]),
                    _text(values[_map(headers, "候选值")]), _maybe_float(values[_map(headers, "置信度")]), _text(values[_map(headers, "建议动作")]),
                    _text(values[_map(headers, "人工备注")]), row_no, ws.title,
                ))
                counts["reviews"] += 1

            # Preserve the non-business source sheets as auditable issues/metadata.
            for sheet_name in ("07_APRIL_ARCHIVE", "09_APRIL_MATCH_AUDIT", "10_SOURCE_SCHEMA"):
                ws = wb[sheet_name]
                source_keys: set[str] = set()
                for row_no, headers, values in _records(ws, SHEET_CONFIG[sheet_name]):
                    raw = _row_to_dict(headers, values)
                    if sheet_name in ("07_APRIL_ARCHIVE", "09_APRIL_MATCH_AUDIT"):
                        _source_issue(db, migration_id, sheet_name, row_no, "AUDIT_ONLY_SOURCE", "", raw)
                    else:
                        source_key = f"{_text(values[1])}¦{_text(values[2])}"
                        if source_key in source_keys:
                            _source_issue(db, migration_id, sheet_name, row_no, "SOURCE_DUPLICATE", source_key, raw)
                            counts["source_issues"] += 1
                        source_keys.add(source_key)
                        _source_issue(db, migration_id, sheet_name, row_no, "SOURCE_SCHEMA_METADATA", source_key, raw)
                    counts["source_issues"] += 1

            db.execute("INSERT INTO migration_runs(migration_id,source_master_path,source_master_sha256,started_at,status,products_count,localizations_count,observations_count,price_history_count,events_count,runs_count,reviews_count,validation_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", (
                migration_id, str(master_path), source_hash, datetime.now().isoformat(timespec="seconds"), "MIGRATED", counts["products"], counts["localizations"], counts["observations"], counts["price_history"], counts["events"], counts["runs"], counts["reviews"], "PENDING",
            ))
        return {"migration_id": migration_id, "master_sha256": source_hash, "counts": counts}
    finally:
        if conn is not None:
            conn.close()
        wb.close()
