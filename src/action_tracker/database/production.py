"""SQLite V2 production transaction primitives.

The writer accepts a pre-computed CommitBundle. Collection and lifecycle code
remain outside this module; this module only persists an already QA-approved
bundle atomically.
"""
from __future__ import annotations

import csv
import hashlib
import json
import os
import sqlite3
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping

from .connection import connect
from .schema import migrate_v2


class ProductionDatabaseError(RuntimeError):
    """Production DB identity, baseline or transaction validation failure."""


def supersede_older_export_sync(db: sqlite3.Connection, new_commit_id: str) -> None:
    """Move older retryable projections behind the new SQLite head.

    Successful historical projections remain valid evidence; only PENDING and
    FAILED rows can be superseded.  Callers must invoke this inside the same
    transaction that creates the new commit.
    """
    db.execute(
        "UPDATE export_sync SET status='SUPERSEDED', error=? "
        "WHERE commit_id<>? AND status IN ('PENDING','FAILED')",
        (f"SUPERSEDED_BY:{new_commit_id}", new_commit_id),
    )


@dataclass(frozen=True)
class CommitBundle:
    run_id: str
    observation_date: str
    qa_state: str
    current_products: tuple[dict[str, Any], ...] = ()
    localization_updates: tuple[dict[str, Any], ...] = ()
    lifecycle_updates: tuple[dict[str, Any], ...] = ()
    observations: tuple[dict[str, Any], ...] = ()
    price_events: tuple[dict[str, Any], ...] = ()
    event_events: tuple[dict[str, Any], ...] = ()
    review_rows: tuple[dict[str, Any], ...] = ()
    run_record: dict[str, Any] = field(default_factory=dict)
    snapshot_path: str | None = None
    snapshot_hash: str | None = None
    base_commit_id: str | None = None
    bundle_hash: str | None = None

    def resolved_hash(self) -> str:
        if self.bundle_hash:
            return self.bundle_hash
        payload = {
            "run_id": self.run_id,
            "observation_date": self.observation_date,
            "qa_state": self.qa_state,
            "current_products": self.current_products,
            "localization_updates": self.localization_updates,
            "lifecycle_updates": self.lifecycle_updates,
            "observations": self.observations,
            "price_events": self.price_events,
            "event_events": self.event_events,
            "review_rows": self.review_rows,
            "run_record": self.run_record,
            "snapshot_path": self.snapshot_path,
            "snapshot_hash": self.snapshot_hash,
            "base_commit_id": self.base_commit_id,
        }
        encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":")).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()


class ProductionWriter:
    """Single-writer SQLite V2 commit implementation.

    The database role is checked on every commit. Shadow mode is the only
    enabled production-safe role for this first implementation; PRIMARY can be
    enabled explicitly once cutover acceptance is complete.
    """

    def __init__(self, path: Path, *, role: str = "SHADOW", localization_drop_threshold: float = 0.5) -> None:
        self.path = Path(path)
        self.role = role
        self.localization_drop_threshold = float(localization_drop_threshold)
        if self.path.exists():
            try:
                with sqlite3.connect(self.path) as raw:
                    row = raw.execute("SELECT value FROM schema_metadata WHERE key='database_role'").fetchone()
            except sqlite3.OperationalError:
                row = None
            if row and str(row[0]) != role:
                raise ProductionDatabaseError("DB_ROLE_MISMATCH_REQUIRES_EXPLICIT_CUTOVER")
        migrate_v2(self.path, role=role)

    def commit(self, bundle: CommitBundle) -> str:
        if bundle.qa_state not in {"PASS", "PASS_PRESENCE_ONLY"}:
            raise ProductionDatabaseError("DB_COMMIT_QA_NOT_PASS")
        if not bundle.run_id or not bundle.observation_date:
            raise ProductionDatabaseError("DB_COMMIT_RUN_ID_OR_DATE_MISSING")
        now = datetime.now(timezone.utc).isoformat()
        commit_id = f"{bundle.observation_date}_{bundle.run_id}_{bundle.resolved_hash()[:12]}"
        with connect(self.path) as db:
            self._check_identity(db)
            existing = db.execute("SELECT commit_id FROM commit_batches WHERE run_id=?", (bundle.run_id,)).fetchone()
            if existing:
                if existing[0] != commit_id:
                    raise ProductionDatabaseError("DB_COMMIT_RUN_ALREADY_EXISTS_WITH_DIFFERENT_BUNDLE")
                return str(existing[0])
            latest = db.execute("SELECT commit_id FROM commit_batches WHERE status='COMMITTED' ORDER BY committed_at DESC LIMIT 1").fetchone()
            latest_id = str(latest[0]) if latest else None
            if bundle.base_commit_id is not None and bundle.base_commit_id != latest_id:
                raise ProductionDatabaseError("BASELINE_CHANGED_BEFORE_COMMIT")
            try:
                db.execute("BEGIN IMMEDIATE")
                self._validate_localization_coverage(db, bundle)
                self._insert_run(db, bundle, now)
                self._upsert_products(db, bundle.current_products, commit_id, now)
                self._upsert_localizations(db, bundle.localization_updates, commit_id, now)
                self._insert_observations(db, bundle.observations)
                self._upsert_lifecycle(db, bundle.lifecycle_updates, now)
                self._insert_prices(db, bundle.price_events, bundle.run_id)
                self._insert_events(db, bundle.event_events, bundle.run_id)
                self._insert_reviews(db, bundle.review_rows, bundle.run_id)
                db.execute(
                    "INSERT INTO run_evidence(run_id,snapshot_path,snapshot_hash,evidence_json) VALUES(?,?,?,?)",
                    (bundle.run_id, bundle.snapshot_path, bundle.snapshot_hash, json.dumps(bundle.run_record, ensure_ascii=False, sort_keys=True, default=str)),
                )
                db.execute(
                    "INSERT INTO commit_batches(commit_id,run_id,base_commit_id,bundle_hash,schema_version,started_at,committed_at,product_count,observation_count,price_event_count,event_count,snapshot_path,snapshot_hash,status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (commit_id, bundle.run_id, bundle.base_commit_id, bundle.resolved_hash(), "2.0.0", now, now,
                     len(bundle.current_products), len(bundle.observations), len(bundle.price_events), len(bundle.event_events),
                     bundle.snapshot_path, bundle.snapshot_hash, "COMMITTED"),
                )
                db.execute("INSERT INTO export_sync(commit_id,status) VALUES(?, 'PENDING')", (commit_id,))
                # Compatibility projections represent only the current DB
                # head.  Once a newer formal commit exists, an older pending
                # projection must never be rebuilt over the newer facts.
                supersede_older_export_sync(db, commit_id)
                self._validate_transaction(db, bundle, commit_id)
                db.commit()
            except Exception:
                db.rollback()
                raise
        return commit_id

    def _validate_localization_coverage(self, db: sqlite3.Connection, bundle: CommitBundle) -> None:
        """Reject catastrophic localization coverage loss before any write.

        A compatibility projection bug can otherwise turn a complete baseline
        into thousands of empty fields while Presence QA still passes.  This
        compares the existing PRIMARY rows with the incoming bundle and only
        blocks a material drop; ordinary single-SKU changes remain allowed.
        """
        current_skus = {
            str(row.get("sku") or row.get("official_sku") or "").strip()
            for row in bundle.current_products
            if not row.get("_historical_minimal") and str(row.get("sku") or row.get("official_sku") or "").strip()
        }
        if not current_skus:
            return
        incoming: dict[tuple[str, str], dict[str, Any]] = {}
        for row in bundle.localization_updates:
            sku = str(row.get("sku") or row.get("official_sku") or "").strip()
            language = str(row.get("language") or "zh").strip()
            if sku in current_skus and language in {"es", "zh"}:
                incoming[(sku, language)] = row
        if not incoming:
            return
        existing_rows = db.execute(
            "SELECT official_sku,language,cat1,cat2,spec,description,details "
            "FROM product_localizations WHERE language IN ('es','zh')"
        ).fetchall()
        existing: dict[tuple[str, str], sqlite3.Row] = {
            (str(row[0]), str(row[1])): row for row in existing_rows if str(row[0]) in current_skus
        }
        field_pairs = (("cat1", "cat1"), ("cat2", "cat2"), ("spec", "spec"),
                       ("description", "description"), ("details", "details"))
        for language in ("es", "zh"):
            for db_field, incoming_field in field_pairs:
                old_count = sum(1 for (sku, lang), row in existing.items()
                                if lang == language and row[2 + field_pairs.index((db_field, incoming_field))] not in (None, ""))
                new_count = sum(1 for (sku, lang), row in incoming.items()
                                if lang == language and row.get(incoming_field) not in (None, ""))
                if old_count and (old_count - new_count) / old_count > self.localization_drop_threshold:
                    raise ProductionDatabaseError(
                        f"DB_LOCALIZATION_COVERAGE_REGRESSION:{language}.{db_field}:{old_count}->{new_count}"
                    )

    def _check_identity(self, db: sqlite3.Connection) -> None:
        values = {row[0]: row[1] for row in db.execute("SELECT key,value FROM schema_metadata")}
        if values.get("schema_family") != "ACTION_SQLITE_DATA" or values.get("schema_version") != "2.0.0":
            raise ProductionDatabaseError("DB_SCHEMA_IDENTITY_MISMATCH")
        if values.get("database_role") not in {"SHADOW", "PRIMARY"}:
            raise ProductionDatabaseError("DB_ROLE_INVALID")
        if values.get("database_role") == "PRIMARY" and self.role != "PRIMARY":
            raise ProductionDatabaseError("DB_PRIMARY_REQUIRES_PRIMARY_WRITER")

    @staticmethod
    def _insert_run(db: sqlite3.Connection, bundle: CommitBundle, now: str) -> None:
        r = bundle.run_record
        db.execute(
            "INSERT INTO runs(run_id,run_date,status,qa_state,dry_run,started_at,ended_at,schema_version) VALUES(?,?,?,?,?,?,?,?)",
            (bundle.run_id, bundle.observation_date, "COMMITTED", bundle.qa_state, int(bool(r.get("dry_run", False))),
             r.get("started_at") or now, r.get("finished_at") or now, "2.0.0"),
        )

    @staticmethod
    def _upsert_products(db: sqlite3.Connection, rows: Iterable[dict[str, Any]], commit_id: str, now: str) -> None:
        for r in rows:
            sku = str(r.get("official_sku") or r.get("sku") or "").strip()
            if not sku:
                raise ProductionDatabaseError("DB_PRODUCT_SKU_MISSING")
            cid = str(r.get("canonical_id") or f"ACT{sku.zfill(7)}")
            if r.get("_historical_minimal"):
                # Existing historical identities already hold their official
                # facts.  A minimal lifecycle-only row must never null out
                # those facts or reset badges/prices during an absence run.
                if db.execute("SELECT 1 FROM products WHERE official_sku=?", (sku,)).fetchone():
                    # Keep the product projection's status in lockstep with
                    # lifecycle_state while preserving all official facts.
                    # Without this narrow update, a missing/offline SKU could
                    # remain visible to CURRENT exports indefinitely.
                    db.execute(
                        "UPDATE products SET status=?, consecutive_missing=?, last_checked_at=?, updated_at=? WHERE official_sku=?",
                        (r.get("status") or "HISTORICAL", int(r.get("consecutive_missing", 0) or 0), r.get("last_checked_at", now), now, sku),
                    )
                    continue
            db.execute(
                """INSERT INTO products(canonical_id,official_sku,name_es,name_zh,current_price,original_price,unit_price_raw,raw_badges,
                action_new_badge,promotion_active,sustainable_badge,status,consecutive_missing,product_url,image_url,first_seen_at,
                last_seen_at,last_checked_at,source_hash,updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(official_sku) DO UPDATE SET canonical_id=excluded.canonical_id,name_es=excluded.name_es,name_zh=excluded.name_zh,
                current_price=excluded.current_price,original_price=excluded.original_price,unit_price_raw=excluded.unit_price_raw,
                raw_badges=excluded.raw_badges,action_new_badge=excluded.action_new_badge,promotion_active=excluded.promotion_active,
                sustainable_badge=excluded.sustainable_badge,status=excluded.status,consecutive_missing=excluded.consecutive_missing,
                product_url=excluded.product_url,image_url=excluded.image_url,first_seen_at=excluded.first_seen_at,last_seen_at=excluded.last_seen_at,
                last_checked_at=excluded.last_checked_at,source_hash=excluded.source_hash,updated_at=excluded.updated_at""",
                (cid, sku, r.get("name_es"), r.get("name_zh"), r.get("current_price"), r.get("original_price"), r.get("unit_price_raw", r.get("unit_price")),
                 r.get("raw_badges", r.get("raw_tags")), int(_to_bool(r.get("action_new_badge", r.get("is_new_badge", False)))),
                 int(_to_bool(r.get("promotion_active", r.get("promotion", False)))), int(_to_bool(r.get("sustainable_badge", r.get("sustainable", False)))),
                 r.get("status", "ACTIVE"), int(r.get("consecutive_missing", 0) or 0), r.get("product_url"), r.get("image_url"),
                 r.get("first_seen_at", r.get("first_seen")), r.get("last_seen_at", r.get("last_seen")), r.get("last_checked_at", now), r.get("source_hash"), now),
            )

    @staticmethod
    def _upsert_localizations(db: sqlite3.Connection, rows: Iterable[dict[str, Any]], commit_id: str, now: str) -> None:
        for r in rows:
            sku = str(r.get("official_sku") or r.get("sku") or "").strip()
            language = str(r.get("language") or "zh")
            incoming = dict(r)
            if language == "zh":
                existing_row = db.execute(
                    "SELECT name,cat1,cat2,spec,unit_price,description,details,source,review_status,source_hash,resolution_status,name_source,cat1_source,cat2_source,spec_source,unit_price_source,description_source,details_source,freshness_status,approved_by,approved_at,applied_commit_id,last_commit_id,updated_at FROM product_localizations WHERE official_sku=? AND language='zh'",
                    (sku,),
                ).fetchone()
                if existing_row is not None:
                    existing = dict(existing_row)
                    incoming_hash = str(incoming.get("source_hash") or "")
                    existing_hash = str(existing.get("source_hash") or "")
                    # Daily observation is not a localization Apply.  When
                    # official Spanish facts changed, retain the last known
                    # Chinese text and provenance and mark it STALE.  When
                    # facts did not change, preserve approval/LOCK metadata.
                    if existing_hash and incoming_hash and existing_hash != incoming_hash:
                        for key in ("name", "cat1", "cat2", "spec", "description", "details", "source_hash", "resolution_status", "name_source", "cat1_source", "cat2_source", "spec_source", "description_source", "details_source", "approved_by", "approved_at", "applied_commit_id", "last_commit_id", "updated_at"):
                            if key in existing:
                                incoming[key] = existing[key]
                        incoming["freshness_status"] = "STALE"
                        incoming["review_status"] = existing.get("review_status") or "STALE"
                    else:
                        for key in ("name", "cat1", "cat2", "spec", "description", "details", "source_hash", "resolution_status", "review_status", "name_source", "cat1_source", "cat2_source", "spec_source", "description_source", "details_source", "approved_by", "approved_at", "applied_commit_id", "last_commit_id", "updated_at"):
                            if existing.get(key) is not None:
                                incoming[key] = existing[key]
                        if str(existing.get("freshness_status") or "").upper() == "STALE":
                            incoming["freshness_status"] = "STALE"
            db.execute(
                """INSERT INTO product_localizations(official_sku,language,name,cat1,cat2,spec,unit_price,description,details,source,review_status,updated_at,last_commit_id,
                 source_hash,resolution_status,name_source,cat1_source,cat2_source,spec_source,unit_price_source,description_source,details_source,freshness_status,approved_by,approved_at,applied_commit_id)
                 VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                 ON CONFLICT(official_sku,language) DO UPDATE SET name=excluded.name,cat1=excluded.cat1,cat2=excluded.cat2,spec=excluded.spec,
                 unit_price=excluded.unit_price,description=excluded.description,details=excluded.details,source=excluded.source,review_status=excluded.review_status,
                 updated_at=excluded.updated_at,last_commit_id=excluded.last_commit_id,source_hash=excluded.source_hash,
                 resolution_status=excluded.resolution_status,name_source=excluded.name_source,cat1_source=excluded.cat1_source,
                 cat2_source=excluded.cat2_source,spec_source=excluded.spec_source,unit_price_source=excluded.unit_price_source,description_source=excluded.description_source,
                 details_source=excluded.details_source,freshness_status=excluded.freshness_status,approved_by=excluded.approved_by,
                 approved_at=excluded.approved_at,applied_commit_id=excluded.applied_commit_id""",
                 (sku, language, incoming.get("name"), incoming.get("cat1"), incoming.get("cat2"), incoming.get("spec"), incoming.get("unit_price"), incoming.get("description"), incoming.get("details"),
                  incoming.get("source"), incoming.get("review_status"), incoming.get("updated_at") or now, incoming.get("last_commit_id") or commit_id, incoming.get("source_hash"),
                  incoming.get("resolution_status"), incoming.get("name_source"), incoming.get("cat1_source"), incoming.get("cat2_source"),
                  incoming.get("spec_source"), incoming.get("unit_price_source"), incoming.get("description_source"), incoming.get("details_source"), incoming.get("freshness_status"),
                  incoming.get("approved_by"), incoming.get("approved_at"), incoming.get("applied_commit_id") or commit_id),
            )

    @staticmethod
    def _insert_observations(db: sqlite3.Connection, rows: Iterable[dict[str, Any]]) -> None:
        for r in rows:
            state = str(r.get("presence_state") or "UNKNOWN")
            if state not in {"PRESENT", "ABSENT", "UNKNOWN"}:
                raise ProductionDatabaseError("DB_OBSERVATION_STATE_INVALID")
            db.execute(
                """INSERT INTO observations(run_id,official_sku,observation_date,presence_state,sitemap_present,listing_present,nuevo_present,
                promotion_present,observation_complete,absence_capable,current_price,source_flag) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (r.get("run_id"), r.get("official_sku") or r.get("sku"), r.get("observation_date"), state, r.get("sitemap_present"),
                 r.get("listing_present"), r.get("nuevo_present"), r.get("promotion_present"), int(bool(r.get("observation_complete", False))),
                 int(bool(r.get("absence_capable", False))), r.get("current_price"), r.get("source_flag")),
            )

    @staticmethod
    def _upsert_lifecycle(db: sqlite3.Connection, rows: Iterable[dict[str, Any]], now: str) -> None:
        for r in rows:
            sku = str(r.get("official_sku") or r.get("sku") or "").strip()
            cid = str(r.get("canonical_id") or f"ACT{sku.zfill(7)}")
            db.execute(
                """INSERT INTO lifecycle_state(official_sku,canonical_id,first_seen_date,last_seen_date,current_status,missing_count,last_missing_date,
                offline_date,last_state_observation_date,ever_offline,last_run_id,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(official_sku) DO UPDATE SET canonical_id=excluded.canonical_id,first_seen_date=excluded.first_seen_date,
                last_seen_date=excluded.last_seen_date,current_status=excluded.current_status,missing_count=excluded.missing_count,
                last_missing_date=excluded.last_missing_date,offline_date=excluded.offline_date,last_state_observation_date=excluded.last_state_observation_date,
                ever_offline=excluded.ever_offline,last_run_id=excluded.last_run_id,updated_at=excluded.updated_at""",
                (sku, cid, r.get("first_seen_date", r.get("first_seen")), r.get("last_seen_date", r.get("last_seen")), r.get("current_status", r.get("last_status", "ACTIVE")),
                 int(r.get("missing_count", r.get("consecutive_missing", 0)) or 0), r.get("last_missing_date"), r.get("offline_date"),
                 r.get("last_state_observation_date"), int(_to_bool(r.get("ever_offline", False))), r.get("last_run_id"), now),
            )

    @staticmethod
    def _insert_prices(db: sqlite3.Connection, rows: Iterable[dict[str, Any]], run_id: str) -> None:
        for r in rows:
            key = str(r.get("event_key") or _event_key(run_id, r, "price"))
            db.execute(
                "INSERT OR IGNORE INTO price_history(canonical_id,official_sku,observed_at,old_price,new_price,change_type,run_id,event_key) VALUES(?,?,?,?,?,?,?,?)",
                (r.get("canonical_id"), r.get("official_sku") or r.get("sku"), r.get("observed_at") or r.get("date"), r.get("old_price"),
                 r.get("new_price"), r.get("change_type", r.get("direction", "CHANGE")), r.get("run_id", run_id), key),
            )

    @staticmethod
    def _insert_events(db: sqlite3.Connection, rows: Iterable[dict[str, Any]], run_id: str) -> None:
        for r in rows:
            key = str(r.get("event_key") or _event_key(run_id, r, "event"))
            db.execute(
                "INSERT OR IGNORE INTO event_history(canonical_id,official_sku,occurred_at,event_type,old_value,new_value,run_id,evidence,event_key) VALUES(?,?,?,?,?,?,?,?,?)",
                (r.get("canonical_id"), r.get("official_sku") or r.get("sku"), r.get("occurred_at") or r.get("date"), r.get("event_type"),
                 r.get("old_value"), r.get("new_value"), r.get("run_id", run_id), r.get("evidence"), key),
            )

    @staticmethod
    def _insert_reviews(db: sqlite3.Connection, rows: Iterable[dict[str, Any]], run_id: str) -> None:
        # Keep the dedicated V2 review table as the source of truth, while also
        # retaining the legacy sync_queue copy for older operational tooling.
        for r in rows:
            payload = {**r, "run_id": run_id}
            review_id = str(r.get("review_id") or _event_key(run_id, r, "review"))
            entity_id = str(r.get("sku") or r.get("entity_id") or review_id)
            issue_type = str(r.get("问题类型") or r.get("issue_type") or "DATA_INCONSISTENCY")
            db.execute(
                """INSERT OR IGNORE INTO reviews(review_id,run_id,entity_id,issue_type,evidence,suggested_action,status,created_at,updated_at)
                   VALUES(?,?,?,?,?,?, 'PENDING',CURRENT_TIMESTAMP,CURRENT_TIMESTAMP)""",
                (review_id, run_id, entity_id, issue_type,
                 r.get("证据") or r.get("evidence"), r.get("建议动作") or r.get("suggested_action")),
            )
            db.execute("INSERT INTO sync_queue(entity_type,entity_id,action,status) VALUES(?,?,?, 'PENDING')", ("review", review_id, json.dumps(payload, ensure_ascii=False, default=str)))

    @staticmethod
    def _validate_transaction(db: sqlite3.Connection, bundle: CommitBundle, commit_id: str) -> None:
        if db.execute("SELECT COUNT(*) FROM commit_batches WHERE commit_id=?", (commit_id,)).fetchone()[0] != 1:
            raise ProductionDatabaseError("DB_COMMIT_BATCH_MISSING")
        if db.execute("SELECT COUNT(*) FROM runs WHERE run_id=?", (bundle.run_id,)).fetchone()[0] != 1:
            raise ProductionDatabaseError("DB_RUN_MISSING")
        orphan = db.execute("SELECT COUNT(*) FROM observations o LEFT JOIN products p ON p.official_sku=o.official_sku WHERE o.run_id=? AND p.official_sku IS NULL", (bundle.run_id,)).fetchone()[0]
        if orphan:
            raise ProductionDatabaseError("DB_OBSERVATION_ORPHAN")


def _event_key(run_id: str, row: dict[str, Any], kind: str) -> str:
    payload = json.dumps({"kind": kind, "run_id": run_id, **row}, ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().casefold() in {"1", "true", "yes", "y", "on", "是", "有"}


def database_status(path: Path) -> dict[str, Any]:
    """Return a compact, read-only production database status."""
    if not Path(path).exists():
        return {"exists": False, "path": str(path)}
    with connect(Path(path)) as db:
        metadata = {row[0]: row[1] for row in db.execute("SELECT key,value FROM schema_metadata")}
        try:
            latest = db.execute("SELECT commit_id,run_id,committed_at,status FROM commit_batches ORDER BY committed_at DESC LIMIT 1").fetchone()
            pending = db.execute("SELECT COUNT(*) FROM export_sync WHERE status IN ('PENDING','FAILED')").fetchone()[0]
            products = db.execute("SELECT COUNT(*) FROM products").fetchone()[0]
            lifecycle = db.execute("SELECT COUNT(*) FROM lifecycle_state").fetchone()[0]
        except sqlite3.OperationalError:
            return {"exists": True, "path": str(path), "metadata": metadata, "legacy_schema": True,
                    "latest_commit": None, "pending_export_sync": None,
                    "products": db.execute("SELECT COUNT(*) FROM products").fetchone()[0], "lifecycle": None}
        return {"exists": True, "path": str(path), "metadata": metadata, "latest_commit": dict(latest) if latest else None,
                "pending_export_sync": pending, "products": products, "lifecycle": lifecycle}


def backup_database(source: Path, destination: Path) -> dict[str, Any]:
    """Create a WAL-safe SQLite backup using the SQLite Backup API.

    A raw ``copy`` of a live WAL database can miss frames that have not yet
    been checkpointed.  The Backup API takes a consistent snapshot while the
    source remains open and is therefore the only supported production backup
    primitive.
    """
    source = Path(source)
    destination = Path(destination)
    if not source.exists():
        raise ProductionDatabaseError("DB_MISSING")
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.unlink(missing_ok=True)
    src = sqlite3.connect(source)
    dst = sqlite3.connect(destination)
    try:
        src.backup(dst)
        dst.commit()
    finally:
        dst.close()
        src.close()
    check = sqlite3.connect(destination)
    try:
        integrity = check.execute("PRAGMA integrity_check").fetchone()[0]
    finally:
        check.close()
    if integrity != "ok":
        raise ProductionDatabaseError("DB_BACKUP_INTEGRITY_FAILED")
    return {"source": str(source), "destination": str(destination), "integrity": "PASS",
            "sha256": hashlib.sha256(destination.read_bytes()).hexdigest()}


_LOCALIZATION_RESTORE_FIELDS = {
    "es": {
        "cat1": "cat1_es", "cat2": "cat2_es", "spec": "spec_es",
        "description": "desc_es", "details": "details_es",
    },
    "zh": {
        "cat1": "cat1_zh", "cat2": "cat2_zh", "spec": "spec_zh",
        "description": "desc_zh", "details": "details_zh",
    },
}


def repair_primary_localization_regression(
    path: Path,
    *,
    trusted_snapshot: Path,
    run_id: str,
) -> dict[str, Any]:
    """Repair a confirmed PRIMARY localization-loss incident atomically.

    Only empty localization fields are restored from a trusted prior formal
    snapshot.  Current listing facts and any detail values successfully fetched
    in the affected run are left untouched.  The affected run's derived
    ``CONTENT_CHANGE`` events are then rebuilt from the repaired records.
    """
    path = Path(path)
    trusted_snapshot = Path(trusted_snapshot)
    if not path.exists():
        raise ProductionDatabaseError("DB_MISSING")
    if not trusted_snapshot.exists():
        raise ProductionDatabaseError("TRUSTED_SNAPSHOT_MISSING")
    if trusted_snapshot.name != "products_normalized.csv":
        raise ProductionDatabaseError("TRUSTED_SNAPSHOT_NOT_FORMAL")

    # A filename alone is not evidence that the CSV came from a completed
    # formal run. Require the sibling run/QA reports and bind their identity to
    # the snapshot directory before allowing any PRIMARY mutation.
    report_path = trusted_snapshot.parent / "run_report.json"
    qa_path = trusted_snapshot.parent / "qa_report.json"
    try:
        report = json.loads(report_path.read_text(encoding="utf-8"))
        qa = json.loads(qa_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ProductionDatabaseError("TRUSTED_SNAPSHOT_EVIDENCE_MISSING") from exc
    if not isinstance(report, dict) or not isinstance(qa, dict):
        raise ProductionDatabaseError("TRUSTED_SNAPSHOT_EVIDENCE_INVALID")
    try:
        report_snapshot_matches = (
            Path(str(report.get("snapshot") or "")).resolve()
            == trusted_snapshot.parent.resolve()
        )
    except (OSError, RuntimeError):
        report_snapshot_matches = False
    report_date_matches = (
        str(report.get("run_date") or "") == trusted_snapshot.parent.parent.name
    )
    if (str(report.get("run_id") or "") != trusted_snapshot.parent.name
            or not report_snapshot_matches
            or not report_date_matches
            or str(report.get("commit_status") or "") != "FULL_COMMIT"
            or bool(report.get("dry_run"))
            or qa.get("passed") is not True
            or qa.get("state") not in {"PASS", "PASS_PRESENCE_ONLY"}):
        raise ProductionDatabaseError("TRUSTED_SNAPSHOT_NOT_FORMAL")

    with trusted_snapshot.open("r", encoding="utf-8-sig", newline="") as handle:
        snapshot_rows = {
            str(row.get("sku") or "").strip(): row
            for row in csv.DictReader(handle)
            if str(row.get("sku") or "").strip()
        }
    if not snapshot_rows:
        raise ProductionDatabaseError("TRUSTED_SNAPSHOT_EMPTY")

    from ..services.hashing import content_hash, localization_source_hash

    def present(value: Any) -> bool:
        return value is not None and str(value).strip() != ""

    restored: dict[str, int] = {"es": 0, "zh": 0}
    restored_by_field: dict[str, int] = {
        f"{language}.{field}": 0
        for language, fields in _LOCALIZATION_RESTORE_FIELDS.items()
        for field in fields
    }
    with connect(path) as db:
        metadata = {row[0]: row[1] for row in db.execute("SELECT key,value FROM schema_metadata")}
        if (metadata.get("schema_family") != "ACTION_SQLITE_DATA"
                or metadata.get("schema_version") != "2.0.0"
                or metadata.get("database_role") != "PRIMARY"):
            raise ProductionDatabaseError("PRIMARY_V2_DATABASE_REQUIRED")
        affected = [str(row[0]) for row in db.execute(
            "SELECT official_sku FROM products WHERE status='CURRENT' ORDER BY official_sku"
        )]
        if not affected:
            raise ProductionDatabaseError("PRIMARY_CURRENT_EMPTY")
        missing_snapshot = sorted(set(affected) - set(snapshot_rows))
        if missing_snapshot:
            raise ProductionDatabaseError("TRUSTED_SNAPSHOT_SKU_MISMATCH")
        commit_row = db.execute(
            "SELECT commit_id,event_count FROM commit_batches WHERE run_id=?", (run_id,)
        ).fetchone()
        if commit_row is None:
            raise ProductionDatabaseError("AFFECTED_RUN_COMMIT_MISSING")
        event_count_before = int(commit_row[1] or 0)

        try:
            db.execute("BEGIN IMMEDIATE")
            for sku in affected:
                source = snapshot_rows[sku]
                for language, fields in _LOCALIZATION_RESTORE_FIELDS.items():
                    row = db.execute(
                        "SELECT name,cat1,cat2,spec,description,details FROM product_localizations "
                        "WHERE official_sku=? AND language=?", (sku, language)
                    ).fetchone()
                    if row is None:
                        raise ProductionDatabaseError("PRIMARY_LOCALIZATION_ROW_MISSING")
                    values = dict(row)
                    changes: dict[str, Any] = {}
                    for column, source_field in fields.items():
                        candidate = source.get(source_field)
                        if not present(values.get(column)) and present(candidate):
                            changes[column] = candidate
                            restored[language] += 1
                            restored_by_field[f"{language}.{column}"] += 1
                    if changes:
                        assignments = ", ".join(f"{column}=?" for column in changes)
                        db.execute(
                            f"UPDATE product_localizations SET {assignments}, updated_at=CURRENT_TIMESTAMP "
                            "WHERE official_sku=? AND language=?",
                            (*changes.values(), sku, language),
                        )

            # The source hash is a contract over the ES fact fields, so it must
            # be recomputed after a field-level recovery.  Both language rows
            # share that ES fact hash.
            es_rows = db.execute(
                "SELECT p.official_sku,es.name,es.cat1,es.cat2,es.spec,es.description,es.details "
                "FROM products p JOIN product_localizations es "
                "ON es.official_sku=p.official_sku AND es.language='es' WHERE p.status='CURRENT'"
            ).fetchall()
            for row in es_rows:
                fact = {
                    "name_es": row[1], "cat1_es": row[2], "cat2_es": row[3], "spec_es": row[4],
                    "desc_es": row[5], "details_es": row[6],
                }
                db.execute(
                    "UPDATE product_localizations SET source_hash=? WHERE official_sku=? AND language IN ('es','zh')",
                    (localization_source_hash(fact), row[0]),
                )

            # Rebuild only the derived content events for the affected run;
            # price and badge events were computed from independent Listing
            # evidence and are deliberately preserved.
            deleted_events = db.execute(
                "DELETE FROM event_history WHERE run_id=? AND event_type='CONTENT_CHANGE'", (run_id,)
            ).rowcount
            current_rows = db.execute(
                """SELECT p.canonical_id,p.official_sku,p.name_es,p.product_url,p.image_url,
                   es.name,es.cat1,es.cat2,es.spec,es.description,es.details
                   FROM products p JOIN product_localizations es
                   ON es.official_sku=p.official_sku AND es.language='es'
                   WHERE p.status='CURRENT' ORDER BY p.official_sku"""
            ).fetchall()
            rebuilt_events = 0
            for row in current_rows:
                sku = str(row[1])
                before = _snapshot_fact_record(snapshot_rows[sku])
                after = {
                    "name_es": row[5] or row[2], "cat1_es": row[6], "cat2_es": row[7],
                    "spec_es": row[8], "desc_es": row[9], "details_es": row[10],
                    "product_url": row[3], "image_url": row[4],
                }
                if content_hash(before) == content_hash(after):
                    continue
                event = {
                    "canonical_id": row[0], "sku": sku, "date": _run_date(run_id),
                    "event_type": "CONTENT_CHANGE", "old_value": content_hash(before)[:12],
                    "new_value": content_hash(after)[:12], "evidence": run_id, "run_id": run_id,
                }
                key = _event_key(run_id, event, "event")
                db.execute(
                    "INSERT INTO event_history(canonical_id,official_sku,occurred_at,event_type,old_value,new_value,run_id,evidence,event_key) "
                    "VALUES(?,?,?,?,?,?,?,?,?)",
                    (event["canonical_id"], sku, event["date"], event["event_type"], event["old_value"],
                     event["new_value"], run_id, event["evidence"], key),
                )
                rebuilt_events += 1

            event_count_after = db.execute(
                "SELECT COUNT(*) FROM event_history WHERE run_id=?", (run_id,)
            ).fetchone()[0]
            db.execute(
                "UPDATE commit_batches SET event_count=? WHERE commit_id=?",
                (event_count_after, commit_row[0]),
            )

            evidence = json.dumps({
                "run_id": run_id, "trusted_snapshot": str(trusted_snapshot), "current_sku_count": len(affected),
                "restored": restored, "restored_by_field": restored_by_field,
                "content_events_deleted": deleted_events, "content_events_rebuilt": rebuilt_events,
                "commit_event_count_before": event_count_before, "commit_event_count_after": event_count_after,
            }, ensure_ascii=False, sort_keys=True)
            db.execute(
                "INSERT INTO migration_source_issues(source_name,issue_type,entity_id,details) VALUES(?,?,?,?)",
                ("sqlite_primary_recovery", "LOCALIZATION_REGRESSION_REPAIRED", run_id, evidence),
            )
            db.commit()
        except Exception:
            db.rollback()
            raise
    return {
        "status": "REPAIRED", "run_id": run_id, "trusted_snapshot": str(trusted_snapshot),
        "current_sku_count": len(affected), "restored": restored, "restored_by_field": restored_by_field,
        "content_events_deleted": deleted_events, "content_events_rebuilt": rebuilt_events,
        "commit_event_count_before": event_count_before, "commit_event_count_after": event_count_after,
    }


def _snapshot_fact_record(row: Mapping[str, Any]) -> dict[str, Any]:
    """Map a normalized snapshot CSV row to the content-hash fact contract."""
    return {
        key: (row.get(key) or None)
        for key in ("name_es", "cat1_es", "cat2_es", "spec_es", "desc_es", "details_es", "product_url", "image_url")
    }


def _run_date(run_id: str) -> str:
    value = str(run_id).split("_", 1)[0]
    try:
        datetime.fromisoformat(value)
    except ValueError as exc:
        raise ProductionDatabaseError("RUN_ID_DATE_INVALID") from exc
    return value


def promote_database_role(path: Path, *, target_role: str = "PRIMARY") -> dict[str, str]:
    """Explicitly promote a validated V2 database; never called implicitly."""
    if target_role != "PRIMARY":
        raise ProductionDatabaseError("DB_ROLE_TARGET_UNSUPPORTED")
    if not Path(path).exists():
        raise ProductionDatabaseError("DB_MISSING")
    with connect(Path(path)) as db:
        metadata = {row[0]: row[1] for row in db.execute("SELECT key,value FROM schema_metadata")}
        if metadata.get("schema_family") != "ACTION_SQLITE_DATA" or metadata.get("schema_version") != "2.0.0":
            raise ProductionDatabaseError("DB_V2_SCHEMA_REQUIRED")
        current = metadata.get("database_role")
        if current == "PRIMARY":
            return {"previous_role": "PRIMARY", "role": "PRIMARY", "status": "ALREADY_PRIMARY"}
        if current != "SHADOW":
            raise ProductionDatabaseError("DB_ROLE_INVALID")
        if db.execute("PRAGMA integrity_check").fetchone()[0] != "ok" or db.execute("PRAGMA foreign_key_check").fetchall():
            raise ProductionDatabaseError("DB_NOT_SAFE_TO_PROMOTE")
        db.execute("UPDATE schema_metadata SET value='PRIMARY' WHERE key='database_role'")
    return {"previous_role": "SHADOW", "role": "PRIMARY", "status": "PROMOTED"}


_DETAIL_CORRECTION_FIELDS = {
    "name_es": ("products", "name_es"),
    "cat1_es": ("product_localizations", "cat1"),
    "cat2_es": ("product_localizations", "cat2"),
    "spec_es": ("product_localizations", "spec"),
    "desc_es": ("product_localizations", "description"),
    "details_es": ("product_localizations", "details"),
    "product_url": ("products", "product_url"),
    "image_url": ("products", "image_url"),
}


def apply_detail_corrections(
    path: Path,
    *,
    parent_run_id: str,
    details_by_sku: Mapping[str, Mapping[str, Any]],
    mode: str,
    source_run_date: str | None = None,
) -> dict[str, Any]:
    """Apply validated Detail facts directly to a SQLite PRIMARY database.

    Detail is an enrichment source only.  This function intentionally has no
    SQL path for prices, presence, status or lifecycle columns, and records
    every field change before rebuilding the derived content-change evidence.
    """
    if mode not in {"APPLY", "BACKFILL"}:
        raise ProductionDatabaseError("DETAIL_CORRECTION_MODE_INVALID")
    path = Path(path)
    migrate_v2(path, role="PRIMARY")
    from ..services.hashing import content_hash, localization_source_hash

    changed_skus: set[str] = set()
    changed_fields = 0
    content_events = 0
    with connect(path) as db:
        metadata = {row[0]: row[1] for row in db.execute("SELECT key,value FROM schema_metadata")}
        if metadata.get("database_role") != "PRIMARY":
            raise ProductionDatabaseError("DETAIL_CORRECTION_REQUIRES_SQLITE_PRIMARY")
        parent = db.execute(
            "SELECT c.commit_id,r.run_date,r.qa_state,r.dry_run FROM commit_batches c JOIN runs r ON r.run_id=c.run_id WHERE c.run_id=? AND c.status='COMMITTED'",
            (parent_run_id,),
        ).fetchone()
        if mode == "APPLY" and (parent is None or str(parent[2]) not in {"PASS", "PASS_PRESENCE_ONLY"} or bool(parent[3])):
            raise ProductionDatabaseError("DETAIL_CORRECTION_PARENT_NOT_FORMAL")
        event_date = str(parent[1]) if parent is not None else str(source_run_date or "")
        if not event_date:
            raise ProductionDatabaseError("DETAIL_CORRECTION_SOURCE_DATE_MISSING")
        head = db.execute("SELECT commit_id FROM commit_batches WHERE status='COMMITTED' ORDER BY committed_at DESC,commit_id DESC LIMIT 1").fetchone()
        if head is None:
            raise ProductionDatabaseError("DETAIL_CORRECTION_HEAD_MISSING")
        baseline_head_id = str(head[0])
        if mode == "APPLY" and str(parent[0]) != baseline_head_id:
            raise ProductionDatabaseError("DETAIL_CORRECTION_PARENT_NOT_CURRENT_HEAD")
        try:
            db.execute("BEGIN IMMEDIATE")
            locked_head = db.execute(
                "SELECT commit_id FROM commit_batches WHERE status='COMMITTED' ORDER BY committed_at DESC,commit_id DESC LIMIT 1"
            ).fetchone()
            if locked_head is None or str(locked_head[0]) != baseline_head_id:
                raise ProductionDatabaseError("BASELINE_CHANGED_BEFORE_DETAIL_CORRECTION")
            if mode == "APPLY":
                locked_parent = db.execute(
                    "SELECT commit_id FROM commit_batches WHERE run_id=? AND status='COMMITTED'", (parent_run_id,)
                ).fetchone()
                if locked_parent is None or str(locked_parent[0]) != baseline_head_id:
                    raise ProductionDatabaseError("DETAIL_CORRECTION_PARENT_NOT_CURRENT_HEAD")

            pending: list[dict[str, Any]] = []
            for sku in sorted(details_by_sku):
                sku = str(sku).strip()
                detail = details_by_sku[sku]
                row = db.execute(
                    """SELECT p.canonical_id,p.status,p.name_es,p.product_url,p.image_url,
                              es.name,es.cat1,es.cat2,es.spec,es.description,es.details
                       FROM products p JOIN product_localizations es
                       ON es.official_sku=p.official_sku AND es.language='es'
                       WHERE p.official_sku=?""",
                    (sku,),
                ).fetchone()
                if row is None:
                    raise ProductionDatabaseError(f"DETAIL_CORRECTION_SKU_MISSING:{sku}")
                if str(row[1]) != "CURRENT":
                    raise ProductionDatabaseError(f"DETAIL_CORRECTION_SKU_NOT_CURRENT:{sku}")
                before = {
                    "name_es": row[5] or row[2], "cat1_es": row[6], "cat2_es": row[7],
                    "spec_es": row[8], "desc_es": row[9], "details_es": row[10],
                    "product_url": row[3], "image_url": row[4],
                }
                changes: list[tuple[str, str, Any, Any]] = []
                for field, (table, column) in _DETAIL_CORRECTION_FIELDS.items():
                    value = detail.get(field)
                    if value is None or str(value).strip() == "":
                        continue
                    old = before.get(field)
                    if mode == "BACKFILL" and old is not None and str(old).strip() != "":
                        continue
                    if str(old or "") == str(value):
                        continue
                    changes.append((field, table, old, value))
                if not changes:
                    continue
                after = dict(before)
                after.update({field: value for field, _table, _old, value in changes})
                pending.append({"sku": sku, "row": row, "before": before, "after": after, "changes": changes})

            # A correction that changes facts is itself an immutable formal
            # version.  The parent remains untouched; all derived records and
            # CONTENT_CHANGE events are owned by this correction run.
            if not pending:
                db.rollback()
                return {"status": "NOOP", "mode": mode, "parent_run_id": parent_run_id,
                        "base_commit_id": baseline_head_id, "commit_id": baseline_head_id,
                        "correction_run_id": None, "applied_skus": 0, "applied_fields": 0,
                        "content_change_events": 0}
            correction_run_id = (
                f"detail_{mode.lower()}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S%fZ')}_"
                f"{uuid.uuid4().hex[:8]}"
            )
            started_at = datetime.now(timezone.utc).isoformat()
            payload = {
                "operation": "DETAIL_CORRECTION", "mode": mode,
                "parent_run_id": parent_run_id, "parent_commit_id": str(parent[0]) if parent else None,
                "base_commit_id": baseline_head_id,
                "affected_skus": len(pending),
                "affected_fields": sum(len(item["changes"]) for item in pending),
            }
            bundle_hash = hashlib.sha256(
                json.dumps({**payload, "details": details_by_sku}, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
            ).hexdigest()
            correction_commit_id = f"{event_date}_{correction_run_id}_{bundle_hash[:12]}"
            db.execute(
                "INSERT INTO runs(run_id,run_date,status,qa_state,dry_run,started_at,ended_at,schema_version) VALUES(?,?,?,?,?,?,?,?)",
                (correction_run_id, event_date, "COMMITTED", "PASS", 0, started_at, started_at, "2.0.0"),
            )
            db.execute(
                "INSERT INTO run_evidence(run_id,snapshot_path,snapshot_hash,evidence_json) VALUES(?,?,?,?)",
                (correction_run_id, None, None, json.dumps(payload, ensure_ascii=False, sort_keys=True)),
            )
            db.execute(
                "INSERT INTO commit_batches(commit_id,run_id,base_commit_id,bundle_hash,schema_version,started_at,committed_at,product_count,observation_count,price_event_count,event_count,status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (correction_commit_id, correction_run_id, baseline_head_id, bundle_hash, "2.0.0", started_at, started_at,
                 len(pending), 0, 0, 0, "COMMITTED"),
            )
            db.execute("INSERT INTO export_sync(commit_id,status) VALUES(?, 'PENDING')", (correction_commit_id,))
            supersede_older_export_sync(db, correction_commit_id)

            for item in pending:
                sku = item["sku"]
                row = item["row"]
                before = item["before"]
                after = item["after"]
                changes = item["changes"]
                for field, table, old, value in changes:
                    if table == "products":
                        db.execute(f"UPDATE products SET {_DETAIL_CORRECTION_FIELDS[field][1]}=?,updated_at=CURRENT_TIMESTAMP WHERE official_sku=?", (value, sku))
                        if field == "name_es":
                            # The product convenience column and the ES
                            # localization are one official fact and must not
                            # diverge after a Detail correction.
                            db.execute("UPDATE product_localizations SET name=?,updated_at=CURRENT_TIMESTAMP,last_commit_id=?,applied_commit_id=? WHERE official_sku=? AND language='es'", (value, correction_commit_id, correction_commit_id, sku))
                    else:
                        db.execute(f"UPDATE product_localizations SET {_DETAIL_CORRECTION_FIELDS[field][1]}=?,updated_at=CURRENT_TIMESTAMP,last_commit_id=?,applied_commit_id=? WHERE official_sku=? AND language='es'", (value, correction_commit_id, correction_commit_id, sku))
                fact_hash = localization_source_hash(after)
                db.execute("UPDATE products SET source_hash=?,updated_at=CURRENT_TIMESTAMP WHERE official_sku=?", (fact_hash, sku))
                db.execute("UPDATE product_localizations SET source_hash=?,last_commit_id=?,applied_commit_id=?,freshness_status='CURRENT',updated_at=CURRENT_TIMESTAMP WHERE official_sku=? AND language='es'", (fact_hash, correction_commit_id, correction_commit_id, sku))
                text_changed = any(field in {"name_es", "cat1_es", "cat2_es", "spec_es", "desc_es", "details_es"} for field, _table, _old, _value in changes)
                if text_changed:
                    # Chinese content was not regenerated by this operation;
                    # retain its old provenance and explicitly mark it stale.
                    db.execute("UPDATE product_localizations SET freshness_status='STALE',updated_at=CURRENT_TIMESTAMP WHERE official_sku=? AND language='zh'", (sku,))
                for field, _table, old, value in changes:
                    db.execute(
                        "INSERT INTO detail_corrections(correction_id,parent_run_id,official_sku,field_name,old_value,new_value,mode,source_hash) VALUES(?,?,?,?,?,?,?,?)",
                        (uuid.uuid4().hex, parent_run_id, sku, field, None if old is None else str(old), str(value), mode, fact_hash),
                    )
                if content_hash(before) != content_hash(after):
                    event = {
                        "canonical_id": row[0], "sku": sku, "date": event_date, "event_type": "CONTENT_CHANGE",
                        "old_value": content_hash(before)[:12], "new_value": content_hash(after)[:12], "evidence": f"DETAIL_{mode}", "run_id": correction_run_id,
                    }
                    key = _event_key(correction_run_id, event, "detail-correction")
                    db.execute("DELETE FROM event_history WHERE event_key=?", (key,))
                    db.execute(
                        "INSERT INTO event_history(canonical_id,official_sku,occurred_at,event_type,old_value,new_value,run_id,evidence,event_key) VALUES(?,?,?,?,?,?,?,?,?)",
                        (event["canonical_id"], sku, event["date"], event["event_type"], event["old_value"], event["new_value"], correction_run_id, event["evidence"], key),
                    )
                    content_events += 1
                changed_skus.add(sku)
                changed_fields += len(changes)
            db.execute("UPDATE commit_batches SET event_count=? WHERE commit_id=?", (content_events, correction_commit_id))
            db.commit()
        except Exception:
            db.rollback()
            raise
    return {"status": "SUCCESS", "parent_run_id": parent_run_id,
            "source_parent_run_id": parent_run_id,
            "source_parent_commit_id": str(parent[0]) if parent else None,
            "base_commit_id": baseline_head_id, "correction_run_id": correction_run_id,
            "commit_id": correction_commit_id,
            "applied_skus": len(changed_skus), "applied_fields": changed_fields,
            "content_change_events": content_events, "mode": mode}


def apply_localization_correction(
    path: Path,
    *,
    run_id: str,
    localizations_by_sku: Mapping[str, Mapping[str, Any]],
    source_hashes: Mapping[str, str],
    apply_date: str | None = None,
) -> dict[str, Any]:
    """Create an immutable zh-only correction commit on SQLite PRIMARY.

    This deliberately has no SQL path for products, observations, prices,
    lifecycle or ES rows.  The current committed head is locked before any
    mutation and becomes ``base_commit_id`` for the new correction version.
    """
    path = Path(path); migrate_v2(path, role="PRIMARY")
    fields = {"name", "cat1", "cat2", "spec", "unit_price", "description", "details"}
    with connect(path) as db:
        role = db.execute("SELECT value FROM schema_metadata WHERE key='database_role'").fetchone()
        if not role or str(role[0]) != "PRIMARY":
            raise ProductionDatabaseError("LOCALIZATION_CORRECTION_REQUIRES_SQLITE_PRIMARY")
        head = db.execute("SELECT commit_id FROM commit_batches WHERE status='COMMITTED' ORDER BY committed_at DESC,commit_id DESC LIMIT 1").fetchone()
        if not head: raise ProductionDatabaseError("LOCALIZATION_CORRECTION_HEAD_MISSING")
        base = str(head[0]); now = datetime.now(timezone.utc).isoformat()
        db.execute("BEGIN IMMEDIATE")
        locked = db.execute("SELECT commit_id FROM commit_batches WHERE status='COMMITTED' ORDER BY committed_at DESC,commit_id DESC LIMIT 1").fetchone()
        if not locked or str(locked[0]) != base: raise ProductionDatabaseError("BASELINE_CHANGED_BEFORE_LOCALIZATION_CORRECTION")
        payload = {"operation": "LOCALIZATION_CORRECTION", "run_id": run_id, "base_commit_id": base, "skus": sorted(localizations_by_sku)}
        bundle_hash = hashlib.sha256(json.dumps({"payload": payload, "values": localizations_by_sku, "hashes": source_hashes}, ensure_ascii=False, sort_keys=True, default=str).encode()).hexdigest()
        correction_run = f"localization_{run_id}_{uuid.uuid4().hex[:8]}"
        commit_id = f"{run_id}_{correction_run}_{bundle_hash[:12]}"
        business_date = str(apply_date or datetime.now().astimezone().date().isoformat())
        db.execute("INSERT INTO runs(run_id,run_date,status,qa_state,dry_run,started_at,ended_at,schema_version) VALUES(?,?,?,?,?,?,?,?)", (correction_run, business_date, "COMMITTED", "PASS", 0, now, now, "2.0.0"))
        db.execute("INSERT INTO run_evidence(run_id,snapshot_path,snapshot_hash,evidence_json) VALUES(?,?,?,?)", (correction_run, None, None, json.dumps(payload, ensure_ascii=False, sort_keys=True)))
        db.execute("INSERT INTO commit_batches(commit_id,run_id,base_commit_id,bundle_hash,schema_version,started_at,committed_at,product_count,observation_count,price_event_count,event_count,status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", (commit_id, correction_run, base, bundle_hash, "2.0.0", now, now, len(localizations_by_sku), 0, 0, 0, "COMMITTED"))
        db.execute("INSERT INTO export_sync(commit_id,status) VALUES(?, 'PENDING')", (commit_id,)); supersede_older_export_sync(db, commit_id)
        changed = 0
        for sku, values in sorted(localizations_by_sku.items()):
            sku = str(sku).strip(); source_hash_value = str(source_hashes.get(sku) or "")
            if not source_hash_value: raise ProductionDatabaseError(f"LOCALIZATION_SOURCE_HASH_MISSING:{sku}")
            if db.execute("SELECT 1 FROM products WHERE official_sku=? AND status='CURRENT'", (sku,)).fetchone() is None: raise ProductionDatabaseError(f"LOCALIZATION_SKU_NOT_CURRENT:{sku}")
            metadata_keys = {f"{f}_source" for f in fields} | {"provenance", "sources"}
            if set(values) - fields - metadata_keys:
                raise ProductionDatabaseError(f"LOCALIZATION_FIELD_NOT_ALLOWED:{sku}")
            from ..services.hashing import localization_source_hash
            es = db.execute("SELECT name,cat1,cat2,spec,description,details FROM product_localizations WHERE official_sku=? AND language='es'", (sku,)).fetchone()
            if es is not None:
                expected_hash = localization_source_hash({"name_es": es[0], "cat1_es": es[1], "cat2_es": es[2], "spec_es": es[3], "desc_es": es[4], "details_es": es[5]})
                if source_hash_value != expected_hash:
                    raise ProductionDatabaseError(f"LOCALIZATION_SOURCE_HASH_MISMATCH:{sku}")
            # Older migrated fixtures can legitimately lack an ES projection;
            # the caller-provided hash is then retained as the explicit
            # evidence hash.  New production databases always contain ES and
            # take the strict branch above.
            current = db.execute("SELECT * FROM product_localizations WHERE official_sku=? AND language='zh'", (sku,)).fetchone()
            current_dict = dict(current) if current else {}
            vals = {f: current_dict.get(f) for f in fields}; src = {f: current_dict.get(f"{f}_source") for f in fields}
            provenance = values.get("provenance") or values.get("sources") or {}
            for f, v in values.items():
                if f in fields and v is not None:
                    vals[f] = str(v)
                    src[f] = str(values.get(f"{f}_source") or (provenance.get(f) if isinstance(provenance, Mapping) else "") or "localization_engine")
            db.execute("""INSERT INTO product_localizations(official_sku,language,name,cat1,cat2,spec,unit_price,description,details,source,review_status,updated_at,last_commit_id,source_hash,resolution_status,name_source,cat1_source,cat2_source,spec_source,unit_price_source,description_source,details_source,freshness_status,approved_by,approved_at,applied_commit_id)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(official_sku,language) DO UPDATE SET name=excluded.name,cat1=excluded.cat1,cat2=excluded.cat2,spec=excluded.spec,unit_price=excluded.unit_price,description=excluded.description,details=excluded.details,source=excluded.source,review_status=excluded.review_status,updated_at=excluded.updated_at,last_commit_id=excluded.last_commit_id,source_hash=excluded.source_hash,resolution_status=excluded.resolution_status,name_source=excluded.name_source,cat1_source=excluded.cat1_source,cat2_source=excluded.cat2_source,spec_source=excluded.spec_source,unit_price_source=excluded.unit_price_source,description_source=excluded.description_source,details_source=excluded.details_source,freshness_status=excluded.freshness_status,approved_by=excluded.approved_by,approved_at=excluded.approved_at,applied_commit_id=excluded.applied_commit_id""", (sku, "zh", vals["name"], vals["cat1"], vals["cat2"], vals["spec"], vals["unit_price"], vals["description"], vals["details"], "LOCALIZATION", "APPROVED", now, commit_id, source_hash_value, "APPLIED", src["name"], src["cat1"], src["cat2"], src["spec"], src["unit_price"], src["description"], src["details"], "CURRENT", "LOCALIZATION", now, commit_id))
            changed += 1
        db.commit()
    return {"status": "SUCCESS", "base_commit_id": base, "commit_id": commit_id, "correction_run_id": correction_run, "applied_skus": changed}


def mark_export_sync(path: Path, commit_id: str, *, master: Path, known: Path, offline: Path) -> dict[str, Any]:
    """Mark compatibility projections as synchronized after verifying files.

    This never creates or edits the projections.  Generation remains the
    responsibility of the existing Excel/CSV writer; this function records an
    auditable, content-addressed acknowledgement in SQLite.
    """
    files = {"master": Path(master), "known": Path(known), "offline": Path(offline)}
    hashes: dict[str, str | None] = {}
    missing: list[str] = []
    for name, file_path in files.items():
        if not file_path.exists():
            missing.append(name)
            hashes[name] = None
        else:
            hashes[name] = hashlib.sha256(file_path.read_bytes()).hexdigest()
    status = "SUCCESS" if not missing else "PENDING"
    error = None if not missing else "MISSING_PROJECTION:" + ",".join(missing)
    from .connection import connect

    with connect(Path(path)) as db:
        current = db.execute("SELECT status FROM export_sync WHERE commit_id=?", (commit_id,)).fetchone()
        if current is None:
            raise ProductionDatabaseError("EXPORT_SYNC_COMMIT_MISSING")
        if str(current[0]) == "SUPERSEDED":
            raise ProductionDatabaseError("EXPORT_SYNC_COMMIT_SUPERSEDED")
        db.execute(
            "UPDATE export_sync SET master_status=?,known_status=?,offline_status=?,master_sha256=?,known_sha256=?,offline_sha256=?,last_attempt_at=CURRENT_TIMESTAMP,error=?,status=? WHERE commit_id=?",
            ("SUCCESS" if "master" not in missing else "PENDING",
             "SUCCESS" if "known" not in missing else "PENDING",
             "SUCCESS" if "offline" not in missing else "PENDING",
             hashes["master"], hashes["known"], hashes["offline"], error, status, commit_id),
        )
    return {"commit_id": commit_id, "status": status, "missing": missing, "hashes": hashes}


def sync_pending_exports(path: Path, *, master: Path, known: Path, offline: Path,
                         commit_id: str | None = None) -> list[dict[str, Any]]:
    """Retry export-sync acknowledgements for pending SQLite commits."""
    if not Path(path).exists():
        raise ProductionDatabaseError("DB_MISSING")
    with connect(Path(path)) as db:
        tables = {row[0] for row in db.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        if "export_sync" not in tables:
            raise ProductionDatabaseError("DB_V2_SCHEMA_REQUIRED")
        query = "SELECT commit_id FROM export_sync WHERE status IN ('PENDING','FAILED')"
        args: tuple[Any, ...] = ()
        if commit_id:
            query += " AND commit_id=?"
            args = (commit_id,)
        ids = [str(row[0]) for row in db.execute(query, args).fetchall()]
    return [mark_export_sync(Path(path), value, master=master, known=known, offline=offline) for value in ids]


def persist_image_manifest(path: Path, manifest_path: Path) -> dict[str, int]:
    """Mirror image manifest metadata into V2 without storing image bytes."""
    from ..images.assets import ImageManifest

    manifest = ImageManifest(Path(manifest_path))
    inserted = 0
    skipped = 0
    with connect(Path(path)) as db:
        try:
            db.execute("SELECT 1 FROM image_assets LIMIT 1")
        except sqlite3.OperationalError as exc:
            raise ProductionDatabaseError("DB_V2_IMAGE_TABLE_MISSING") from exc
        for record in manifest.records.values():
            product = db.execute("SELECT 1 FROM products WHERE official_sku=?", (record.sku,)).fetchone()
            if not product:
                skipped += 1
                continue
            db.execute(
                """INSERT INTO image_assets(official_sku,canonical_id,source_image_url,master_image_path,source_hash,master_hash,width,height,status,first_downloaded_at,last_checked_at,updated_at,error_type)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(official_sku) DO UPDATE SET canonical_id=excluded.canonical_id,source_image_url=excluded.source_image_url,
                   master_image_path=excluded.master_image_path,source_hash=excluded.source_hash,master_hash=excluded.master_hash,
                   width=excluded.width,height=excluded.height,status=excluded.status,first_downloaded_at=COALESCE(image_assets.first_downloaded_at, excluded.first_downloaded_at),
                   last_checked_at=excluded.last_checked_at,updated_at=excluded.updated_at,error_type=excluded.error_type""",
                (record.sku, record.canonical_id, record.source_image_url, record.master_image_path,
                 record.source_hash, record.master_hash, record.master_width or record.source_width,
                 record.master_height or record.source_height,
                 "AVAILABLE" if record.available else (
                     "DERIVATIVE_FAILED" if record.derivative_status == "FAILED" else record.download_status
                 ),
                 record.first_downloaded_at or record.last_downloaded_at, record.last_checked_at,
                 datetime.now(timezone.utc).isoformat(), record.error_type or None),
            )
            inserted += 1
    return {"manifest_records": len(manifest.records), "upserted": inserted, "skipped_unknown_product": skipped}


def validate_production_database(path: Path) -> dict[str, Any]:
    """Run invariant checks without mutating the database."""
    if not Path(path).exists():
        raise ProductionDatabaseError("DB_MISSING")
    with connect(Path(path)) as db:
        metadata = {row[0]: row[1] for row in db.execute("SELECT key,value FROM schema_metadata")}
        if metadata.get("schema_family") != "ACTION_SQLITE_DATA" or metadata.get("schema_version") != "2.0.0":
            raise ProductionDatabaseError("DB_V2_SCHEMA_REQUIRED")
        integrity = db.execute("PRAGMA integrity_check").fetchone()[0]
        foreign_keys = db.execute("PRAGMA foreign_key_check").fetchall()
        if integrity != "ok":
            raise ProductionDatabaseError("DB_INTEGRITY_FAILED")
        if foreign_keys:
            raise ProductionDatabaseError("DB_FOREIGN_KEY_FAILED")
        bad_presence = db.execute("SELECT COUNT(*) FROM observations WHERE presence_state NOT IN ('PRESENT','ABSENT','UNKNOWN')").fetchone()[0]
        if bad_presence:
            raise ProductionDatabaseError("DB_PRESENCE_STATE_FAILED")
        return {"integrity": "PASS", "foreign_keys": "PASS", "presence_states": "PASS",
                "products": db.execute("SELECT COUNT(*) FROM products").fetchone()[0],
                "observations": db.execute("SELECT COUNT(*) FROM observations").fetchone()[0],
                "commits": db.execute("SELECT COUNT(*) FROM commit_batches").fetchone()[0]}


def cutover_preflight(path: Path, *, master: Path, known: Path, offline: Path) -> dict[str, Any]:
    """Read-only gate for a Shadow database before an explicit cutover.

    This deliberately does not promote, migrate, or rewrite any file.  The
    caller must still choose the cutover window and perform the explicit
    promotion after this check passes.
    """
    validation = validate_production_database(Path(path))
    with connect(Path(path)) as db:
        role = db.execute("SELECT value FROM schema_metadata WHERE key='database_role'").fetchone()
        pending = db.execute("SELECT COUNT(*) FROM export_sync WHERE status IN ('PENDING','FAILED')").fetchone()[0]
    if not role or role[0] != "SHADOW":
        raise ProductionDatabaseError("CUTOVER_REQUIRES_SHADOW_DATABASE")
    if pending:
        raise ProductionDatabaseError("CUTOVER_EXPORT_SYNC_PENDING")
    for label, candidate in (("master", master), ("known", known), ("offline", offline)):
        if not Path(candidate).exists():
            raise ProductionDatabaseError(f"CUTOVER_PROJECTION_MISSING:{label}")
    from .parity import compare_with_legacy_files
    parity = compare_with_legacy_files(
        {"paths": {"master": Path(master), "state": Path(known).parent}},
        db_path=Path(path),
    )
    if parity["status"] != "PASS":
        raise ProductionDatabaseError("CUTOVER_PARITY_FAILED")
    return {
        "status": "PASS", "database_role": "SHADOW", "pending_export_sync": pending,
        "validation": validation, "parity": parity,
    }


def import_legacy_baseline_v2(db_path: Path, *, master_path: Path, state_dir: Path, observed_at: str) -> str:
    """Build a V2 baseline from the legacy read-only Master/State files.

    This is an explicit migration command, not part of daily collection. The
    source files are never modified and the resulting bundle is committed once
    with an auditable BASELINE run id.
    """
    from ..excel.reader import load_current, load_long_term_official, read_price_history, read_event_history
    from ..state import load_known_skus

    current = load_current(Path(master_path))
    known = load_known_skus(Path(state_dir))
    product_by_sku = {str(r.get("sku") or "").strip(): dict(r) for r in current.values() if str(r.get("sku") or "").strip()}
    for sku, record in known.items():
        product_by_sku.setdefault(sku, {"sku": sku, "canonical_id": record.get("canonical_id"), "status": record.get("last_status", "OFFLINE"), "last_seen": record.get("last_seen_date")})
    # Preserve every official long-term identity so historical price/event
    # rows never become orphaned merely because the SKU is not in today's
    # known_skus projection.  These identities do not receive observations or
    # lifecycle rows unless they are present in known_skus.
    try:
        long_term = load_long_term_official(Path(master_path))
    except (KeyError, ValueError):
        long_term = {}
    for sku, record in long_term.items():
        product_by_sku.setdefault(sku, {
            "sku": sku, "canonical_id": record.get("canonical_id"),
            "name_es": record.get("name_es"), "name_zh": record.get("name_zh"),
            "product_url": record.get("product_url"), "first_seen": record.get("first_seen"),
            "last_seen": record.get("last_seen"), "status": record.get("status") or "HISTORICAL",
        })
    products = tuple(product_by_sku.values())
    lifecycle = []
    for sku, record in known.items():
        lifecycle.append({
            "sku": sku,
            "canonical_id": record.get("canonical_id"),
            "first_seen": record.get("first_seen_date"),
            "last_seen": record.get("last_seen_date"),
            "current_status": record.get("last_status", "ACTIVE"),
            "missing_count": record.get("missing_count", 0),
            "last_missing_date": record.get("last_missing_date"),
            "offline_date": record.get("offline_date"),
            "last_state_observation_date": record.get("last_state_observation_date"),
            "ever_offline": record.get("ever_offline", False),
            "last_run_id": record.get("last_run_id"),
        })
    baseline_run_id = f"BASELINE_{observed_at}"
    try:
        price_rows = read_price_history(Path(master_path))
    except KeyError:
        price_rows = []
    try:
        event_rows = read_event_history(Path(master_path))
    except KeyError:
        event_rows = []
    prices = tuple(_legacy_price_event(row, baseline_run_id) for row in price_rows
                   if str(row.get("SKU") or "").strip())
    events = tuple(_legacy_event(row, baseline_run_id) for row in event_rows
                   if str(row.get("SKU") or "").strip())
    reviews = tuple(_review_row(row, baseline_run_id, index) for index, row in enumerate(_read_master_sheet(Path(master_path), "06_REVIEW_QUEUE"), start=2))
    current_sku_set = set(current)
    bundle = CommitBundle(
        run_id=baseline_run_id, observation_date=observed_at, qa_state="PASS",
        current_products=tuple(products), lifecycle_updates=tuple(lifecycle),
        observations=tuple({"run_id": baseline_run_id, "sku": str(r.get("sku")), "observation_date": observed_at,
                            "presence_state": "PRESENT", "observation_complete": True, "absence_capable": True}
                           for r in products if str(r.get("sku") or "") in current_sku_set),
        price_events=prices, event_events=events, review_rows=reviews,
        run_record={"dry_run": False, "source": "legacy_baseline"}, snapshot_path=str(master_path),
    )
    target = Path(db_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    # The shipped V1 mirror uses incompatible ``products``/``runs`` schemas;
    # SQLite cannot upgrade those tables with CREATE IF NOT EXISTS.  Build a
    # new V2 database beside the target and atomically replace it only after a
    # complete commit.  The caller is responsible for keeping the old target
    # backup (the CLI cutover workflow does this first).
    staged: Path | None = None
    if target.exists() and not _is_v2_database(target):
        staged = target.with_name(f".{target.name}.{uuid.uuid4().hex}.v2")
    commit_path = staged or target
    try:
        commit_id = ProductionWriter(commit_path).commit(bundle)
        _import_master_run_log(commit_path, master_path, baseline_run_id)
        _record_unmatched_event_rows(commit_path, event_rows)
        if staged is not None:
            os.replace(staged, target)
        return commit_id
    finally:
        if staged is not None:
            staged.unlink(missing_ok=True)


def _is_v2_database(path: Path) -> bool:
    """Return whether a database carries the V2 identity metadata."""
    try:
        with connect(Path(path)) as db:
            row = db.execute(
                "SELECT value FROM schema_metadata WHERE key='schema_family'"
            ).fetchone()
    except sqlite3.Error:
        return False
    return bool(row and row[0] == "ACTION_SQLITE_DATA")


def _read_master_sheet(path: Path, sheet: str) -> list[dict[str, Any]]:
    """Read a legacy Master sheet as dictionaries without mutating the file."""
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        header = [cell.value for cell in ws[1]]
        return [dict(zip(header, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    finally:
        wb.close()


def _review_row(row: Mapping[str, Any], run_id: str, source_row: int = 0) -> dict[str, Any]:
    payload = {**dict(row), "_source_row": source_row}
    return {"review_id": _event_key(run_id, payload, "review"),
            "sku": str(row.get("SKU") or "").strip(),
            "issue_type": row.get("问题类型") or "DATA_INCONSISTENCY",
            "evidence": row.get("证据"), "suggested_action": row.get("建议动作")}


def _legacy_price_event(row: Mapping[str, Any], run_id: str) -> dict[str, Any]:
    return {"canonical_id": row.get("Canonical_ID"), "sku": row.get("SKU"),
            "date": row.get("日期"), "old_price": row.get("旧售价 (€)"),
            "new_price": row.get("新售价 (€)"), "change_type": row.get("变化类型"),
            "run_id": run_id}


def _legacy_event(row: Mapping[str, Any], run_id: str) -> dict[str, Any]:
    return {"canonical_id": row.get("Canonical_ID"), "sku": row.get("SKU"),
            "date": row.get("日期"), "event_type": row.get("事件类型"),
            "old_value": row.get("旧值"), "new_value": row.get("新值"),
            "evidence": row.get("备注"), "run_id": run_id}


def _import_master_run_log(path: Path, master_path: Path, baseline_run_id: str) -> None:
    """Import historical run metadata and preserve source statistics in evidence."""
    rows = _read_master_sheet(Path(master_path), "05_RUN_LOG")
    if not rows:
        return
    with connect(Path(path)) as db:
        for row in rows:
            run_id = str(row.get("Run ID") or "").strip()
            if not run_id:
                continue
            run_date = _date_text(row.get("运行日期"))
            started = _join_datetime(run_date, row.get("开始时间"))
            ended = _join_datetime(run_date, row.get("结束时间"))
            db.execute(
                "INSERT OR IGNORE INTO runs(run_id,run_date,status,qa_state,dry_run,started_at,ended_at,schema_version) VALUES(?,?,?,?,?,?,?,?)",
                (run_id, run_date, "COMMITTED", row.get("QA状态") or "PASS", 0, started, ended, "2.0.0"),
            )
            evidence = json.dumps({"source": "legacy_master", "source_sheet": "05_RUN_LOG", "row": row},
                                  ensure_ascii=False, sort_keys=True, default=str)
            db.execute("INSERT OR IGNORE INTO run_evidence(run_id,snapshot_path,snapshot_hash,evidence_json) VALUES(?,?,?,?)",
                       (run_id, str(master_path), None, evidence))


def _record_unmatched_event_rows(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    """Retain audit evidence for archive events without an official SKU."""
    with connect(Path(path)) as db:
        for row in rows:
            if str(row.get("SKU") or "").strip():
                continue
            details = json.dumps(dict(row), ensure_ascii=False, sort_keys=True, default=str)
            db.execute(
                "INSERT INTO migration_source_issues(source_name,issue_type,entity_id,details) VALUES(?,?,?,?)",
                ("04_EVENT_HISTORY", "OFFICIAL_SKU_UNMATCHED", row.get("Canonical_ID"), details),
            )


def _date_text(value: Any) -> str:
    if hasattr(value, "date"):
        value = value.date()
    return str(value or "")[:10]


def _join_datetime(run_date: str, value: Any) -> str:
    text = str(value or "")
    return f"{run_date}T{text}" if run_date and text and "T" not in text else (text or run_date)
