"""字典构建使用的只读历史证据与人工标准化种子。"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

import openpyxl


_CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")
# 详情缺少规格时，页面有时把收藏按钮或品牌导航文案错写进规格列。
# 这些值不是商品事实，必须隔离，不能当作可翻译的西语规格。
_UI_SPEC_RE = re.compile(
    r"^\s*(?:añadir\s+a\s+tus\s+favoritos|todo\s+de\s+.+?)\s*$", re.IGNORECASE,
)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _headers(ws) -> dict[str, int]:
    values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return {_text(value).replace("\ufeff", ""): index for index, value in enumerate(values)}


def _cell(row: tuple[Any, ...], columns: dict[str, int], name: str) -> str:
    index = columns.get(name)
    return _text(row[index]) if index is not None and index < len(row) else ""


def load_standardized_seed(path: Path) -> dict[str, dict[str, str]]:
    """读取用户已生成的中文标准化表，不改动该 Excel。"""
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "全量中文清单" not in wb.sheetnames:
            raise ValueError(f"STANDARDIZED_SEED_SHEET_MISSING: {path}")
        ws = wb["全量中文清单"]
        columns = _headers(ws)
        required = {"SKU", "中文标准品名", "中文规格", "品牌"}
        missing = sorted(required - set(columns))
        if missing:
            raise ValueError(f"STANDARDIZED_SEED_SCHEMA_MISSING: {missing}")
        result: dict[str, dict[str, str]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = _cell(row, columns, "SKU")
            if not sku:
                continue
            if sku in result:
                raise ValueError(f"STANDARDIZED_SEED_DUPLICATE_SKU: {sku}")
            result[sku] = {
                "name_zh": _cell(row, columns, "中文标准品名"),
                "brand_id": _cell(row, columns, "品牌"),
                "cat1_zh": _cell(row, columns, "一级类目（中文）"),
                "cat2_zh": _cell(row, columns, "二级类目（中文）"),
                "spec_zh": _cell(row, columns, "中文规格"),
                "seed_status": _cell(row, columns, "中文状态"),
            }
        return result
    finally:
        wb.close()


def load_brand_reference(path: Path) -> list[dict[str, str]]:
    """读取用户确认过的品牌清单，作为允许保留的拉丁品牌集合。"""
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "品牌清单" not in wb.sheetnames:
            raise ValueError(f"BRAND_REFERENCE_SHEET_MISSING: {path}")
        ws = wb["品牌清单"]
        columns = _headers(ws)
        if "品牌" not in columns:
            raise ValueError("BRAND_REFERENCE_SCHEMA_MISSING: 品牌")
        result: list[dict[str, str]] = []
        seen: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            brand = _cell(row, columns, "品牌")
            if not brand:
                continue
            key = brand.casefold()
            if key in seen:
                continue
            seen.add(key)
            result.append({
                "brand_id": brand,
                "canonical_name": brand,
                "aliases_es": brand,
                "keep_original": "1",
                "is_action_brand": "1" if "Action官网明确自有品牌" in _cell(row, columns, "识别类型") else "0",
                "confidence": "REFERENCE",
                "review_status": "HUMAN_REVIEWED",
                "notes": "2026-08-24 品牌识别清单",
            })
        return result
    finally:
        wb.close()


def load_clean_historical_spanish_reference(path: Path) -> dict[str, dict[str, str]]:
    """按字段选取最近的无中文污染历史证据。

    旧版按整行选择记录：同一行的品名可能是干净西语，但规格或类目已经
    被中文污染，结果会把本来可以恢复的字段一起丢弃。现在每个西语字段
    独立比较日期，既不拼造来源，也不使用中文反向翻译。
    """
    if not path.exists():
        return {}
    result: dict[str, dict[str, str]] = {}
    field_dates: dict[tuple[str, str], str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            sku = _text(row.get("sku"))
            if not sku:
                continue
            reference_date = _text(row.get("date"))
            for field, source_field in (
                ("name_es", "name_es"), ("spec_es", "spec_es"), ("cat1_es", "category_es"),
            ):
                value = _text(row.get(source_field))
                if not value or _CJK_RE.search(value):
                    continue
                key = (sku, field)
                if reference_date < field_dates.get(key, ""):
                    continue
                field_dates[key] = reference_date
                candidate = result.setdefault(sku, {})
                candidate[field] = value
                candidate["reference_date"] = max(reference_date, candidate.get("reference_date", ""))
                candidate["reference_source"] = _text(row.get("source_file"))
    return result


def has_cjk(value: object) -> bool:
    return bool(_CJK_RE.search(_text(value)))


def is_polluted_source_field(field: str, value: object) -> bool:
    """判断已知网页 UI 文案是否误落入西语商品事实字段。"""
    return field == "spec_es" and bool(_UI_SPEC_RE.match(_text(value)))


def restore_spanish_facts(records: dict[str, dict], reference: dict[str, dict[str, str]]) -> int:
    """只在 Master 西语字段含中文时用历史西语证据恢复；不覆盖正常官网字段。"""
    repaired = 0
    for sku, record in records.items():
        evidence = reference.get(sku)
        if not evidence:
            continue
        for field in ("name_es", "spec_es", "cat1_es"):
            if has_cjk(record.get(field)) and evidence.get(field) and not has_cjk(evidence[field]):
                record[field] = evidence[field]
                repaired += 1
    return repaired
