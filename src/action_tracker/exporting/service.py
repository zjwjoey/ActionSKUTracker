"""正式商品清单导出服务：只读来源、严格验证、生成 Excel 与 manifest。"""
from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import math
import os
import re
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import openpyxl

from ..excel.reader import load_current
from ..services.normalization import parse_bool_zh, parse_price
from .dictionary_join import (
    DictionaryJoinError,
    build_zh_rows,
    load_dictionary_context,
    unresolved_brand_ids_for_records,
)
from .excel_writer import write_catalog_xlsx
from .profiles import ExportProfile, ExportProfileError, load_profile


class ExportValidationError(ValueError):
    """正式导出来源或记录不满足冻结契约。"""


@dataclass(frozen=True)
class ExportSource:
    export_date: str
    run_id: str
    kind: str
    records: tuple[dict[str, Any], ...]
    source_master_file_hash: str | None
    directory: Path | None = None


_SOURCE_HASH_FIELDS = (
    "sku", "canonical_id", "name_es", "cat1_es", "cat2_es", "spec_es", "current_price",
    "original_price", "unit_price", "desc_es", "details_es", "product_url", "image_url",
    "is_new_badge", "promotion", "sustainable", "discount", "raw_tags", "status",
    "first_seen", "last_seen",
)
_CJK_RE = re.compile(r"[\u3400-\u9fff]")


def export_catalog(
    cfg: dict[str, Any],
    *,
    language: str,
    export_date: str,
    no_images: bool,
    run_id: str | None = None,
) -> dict[str, Any]:
    """导出一个正式全量清单；整个过程只读取来源并写入 exports 目录。"""
    _validate_date(export_date)
    try:
        profile = load_profile(cfg, language=language, no_images=no_images)
    except ExportProfileError as exc:
        raise ExportValidationError(str(exc)) from exc
    if profile.language != language:
        raise ExportValidationError(f"EXPORT_PROFILE_LANGUAGE_MISMATCH: {profile.profile_id}")
    source = resolve_formal_source(cfg, export_date=export_date, requested_run_id=run_id, profile=profile)
    validate_source_records(source.records, export_date=export_date)
    dictionary_hash = None
    fallback_counts: dict[str, int] = {}
    unresolved_brand_ids: list[str] = []
    if language == "es":
        validate_spanish_source_fields(source.records)
        rows = build_es_rows(source.records)
    elif language == "zh":
        try:
            dictionary = load_dictionary_context(cfg)
            rows, fallback_counts = build_zh_rows(source.records, dictionary)
            dictionary_hash = dictionary.content_hash
            unresolved_brand_ids = unresolved_brand_ids_for_records(source.records, dictionary)
        except DictionaryJoinError as exc:
            raise ExportValidationError(str(exc)) from exc
        validate_zh_rows_against_source(rows, source.records)
    else:
        raise ExportValidationError(f"EXPORT_LANGUAGE_UNSUPPORTED: {language}")
    validate_output_rows(rows)

    date_compact = export_date.replace("-", "")
    output_path = Path(cfg["paths"]["exports"]) / profile.filename_for(date_compact)
    headers = [str(column["header"]) for column in profile.columns]
    expected_skus = {str(r["编号"]) for r in rows}
    # 先写入并验证旁路临时文件；工作簿和 manifest 通过校验后成对发布，
    # 避免验证失败或 manifest 写入失败时留下半套导出物。
    preview_path = output_path.with_name(f".{output_path.stem}.preview.xlsx")
    try:
        image_stats = write_catalog_xlsx(
            preview_path, headers=headers, rows=rows, workbook_format=profile.workbook_format,
            image_root=(Path(cfg["paths"]["images"]) / "derivatives" / "excel_250") if not no_images else None,
            embed_images=not no_images,
        )
        _verify_written_workbook(preview_path, headers=headers, expected_skus=expected_skus)
        source_hash = canonical_source_hash(source.records)
        manifest = {
            "run_id": source.run_id,
            "export_date": export_date,
            "sku_count": len(rows),
            "sku_set_hash": _sku_set_hash(expected_skus),
            "source_master_hash": source_hash,
            "source_master_file_hash": source.source_master_file_hash,
            "source_kind": source.kind,
            "profile_id": profile.profile_id,
            "profile_version": profile.version,
            "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "output_file": output_path.name,
            "validation_results": {
                "source_records": "PASS",
                "output_rows": "PASS",
                "workbook": "PASS",
            },
            "detail_retry_ids": _detail_retry_ids(source),
            "image_profile": "excel_250_white_v1" if not no_images else None,
            "image_embedded_count": image_stats["embedded_count"],
            "image_missing_count": image_stats["missing_count"],
        }
        if language == "zh":
            manifest["dictionary_hash"] = dictionary_hash
            manifest["dictionary_fallback_counts"] = fallback_counts
            manifest["dictionary_unresolved_brand_ids"] = unresolved_brand_ids
        manifest_path = output_path.with_suffix(".manifest.json")
        _publish_export_pair(preview_path, output_path, manifest_path, manifest)
    finally:
        if preview_path.exists():
            preview_path.unlink()
    return {
        "output": str(output_path),
        "manifest": str(manifest_path),
        "run_id": source.run_id,
        "sku_count": len(rows),
        "source_kind": source.kind,
        "profile": profile.profile_id,
        "image_embedded_count": image_stats["embedded_count"],
        "image_missing_count": image_stats["missing_count"],
    }


def resolve_formal_source(
    cfg: dict[str, Any],
    *,
    export_date: str,
    requested_run_id: str | None,
    profile: ExportProfile,
) -> ExportSource:
    """解析唯一正式 run；最新正式 run 才允许直接读取 Master CURRENT。"""
    candidates = _formal_snapshot_candidates(Path(cfg["paths"]["snapshots"]), export_date, profile.source_policy)
    if requested_run_id:
        candidates = [candidate for candidate in candidates if candidate["run_id"] == requested_run_id]
        if not candidates:
            raise ExportValidationError(f"FORMAL_RUN_NOT_FOUND: {requested_run_id} on {export_date}")
    if not candidates:
        raise ExportValidationError(f"NO_FORMAL_EXPORT_SOURCE: {export_date}")
    if len(candidates) != 1:
        raise ExportValidationError(f"AMBIGUOUS_FORMAL_EXPORT_SOURCE: {export_date}")
    selected = candidates[0]
    master = Path(cfg["paths"]["master"])
    latest_master_run = _latest_successful_master_run(master, profile.source_policy)
    if latest_master_run == selected["run_id"]:
        records = tuple(load_current(master).values())
        return ExportSource(
            export_date=export_date,
            run_id=selected["run_id"],
            kind="MASTER_CURRENT",
            records=records,
            source_master_file_hash=_sha256_file(master),
            directory=selected["directory"],
        )
    snapshot_records = _read_csv_records(selected["directory"] / "products_normalized.csv")
    records = tuple(_merge_detail_retries(selected["directory"], snapshot_records))
    if not records:
        raise ExportValidationError(f"FORMAL_SNAPSHOT_PRODUCTS_MISSING: {selected['directory']}")
    return ExportSource(
        export_date=export_date,
        run_id=selected["run_id"],
        kind="FORMAL_SNAPSHOT",
        records=records,
        source_master_file_hash=None,
        directory=selected["directory"],
    )


def _formal_snapshot_candidates(snapshots: Path, export_date: str, policy: dict[str, Any]) -> list[dict[str, Any]]:
    date_dir = snapshots / export_date
    if not date_dir.exists():
        return []
    allowed_qa = {str(value) for value in policy.get("allowed_qa_states") or ()}
    required_commit = str(policy.get("required_commit_status") or "FULL_COMMIT")
    candidates: list[dict[str, Any]] = []
    for directory in date_dir.iterdir():
        if not directory.is_dir():
            continue
        report_path, qa_path = directory / "run_report.json", directory / "qa_report.json"
        if not report_path.exists() or not qa_path.exists():
            continue
        try:
            report = json.loads(report_path.read_text(encoding="utf-8"))
            qa = json.loads(qa_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise ExportValidationError(f"FORMAL_SNAPSHOT_JSON_INVALID: {directory}") from exc
        if not isinstance(report, dict) or not isinstance(qa, dict):
            raise ExportValidationError(f"FORMAL_SNAPSHOT_JSON_INVALID: {directory}")
        run_id = str(report.get("run_id") or "")
        if not run_id or str(report.get("run_date") or report.get("observation_date") or "") != export_date:
            continue
        if str(report.get("commit_status") or "") != required_commit:
            continue
        if policy.get("reject_dry_run", True) and _is_truthy(report.get("dry_run")):
            continue
        if not bool(qa.get("passed")) or str(qa.get("state") or "") not in allowed_qa:
            continue
        candidates.append({"run_id": run_id, "directory": directory, "report": report, "qa": qa})
    return candidates


def _latest_successful_master_run(master: Path, policy: dict[str, Any]) -> str | None:
    if not master.exists():
        raise ExportValidationError(f"MASTER_MISSING: {master}")
    allowed_qa = {str(value) for value in policy.get("allowed_qa_states") or ()}
    workbook = openpyxl.load_workbook(master, read_only=True, data_only=True)
    try:
        if "05_RUN_LOG" not in workbook.sheetnames:
            raise ExportValidationError("MASTER_RUN_LOG_MISSING")
        ws = workbook["05_RUN_LOG"]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows, ()) or ())
        index = {str(value): pos for pos, value in enumerate(headers)}
        required = {"Run ID", "QA状态", "运行状态"}
        if not required.issubset(index):
            raise ExportValidationError(f"MASTER_RUN_LOG_SCHEMA_MISSING: {sorted(required - set(index))}")
        successful: list[str] = []
        for row in rows:
            run_id = str(row[index["Run ID"]] or "").strip()
            qa_state = str(row[index["QA状态"]] or "").strip()
            status = str(row[index["运行状态"]] or "").strip()
            if run_id and qa_state in allowed_qa and status == "SUCCESS":
                successful.append(run_id)
        return successful[-1] if successful else None
    finally:
        workbook.close()


def validate_source_records(records: Iterable[dict[str, Any]], *, export_date: str) -> None:
    """验证 SKU、状态、价格和 URL；不在此处修补来源数据。"""
    rows = list(records)
    if not rows:
        raise ExportValidationError("EXPORT_SOURCE_EMPTY")
    skus: set[str] = set()
    for row in rows:
        sku = str(row.get("sku") or "").strip()
        if not sku:
            raise ExportValidationError("EXPORT_SOURCE_EMPTY_SKU")
        if sku in skus:
            raise ExportValidationError(f"EXPORT_SOURCE_DUPLICATE_SKU: {sku}")
        skus.add(sku)
        if str(row.get("status") or "").strip() != "CURRENT":
            raise ExportValidationError(f"EXPORT_SOURCE_NON_CURRENT_SKU: {sku}")
        current = _required_price(row.get("current_price"), sku=sku, field="current_price")
        original = _optional_price(row.get("original_price"), sku=sku, field="original_price")
        if original is not None and not math.isfinite(original):
            raise ExportValidationError(f"EXPORT_SOURCE_BAD_PRICE: {sku}/original_price")
        product_url = str(row.get("product_url") or "").strip()
        if not _is_http_url(product_url):
            raise ExportValidationError(f"EXPORT_SOURCE_BAD_PRODUCT_URL: {sku}")
        image_url = str(row.get("image_url") or "").strip()
        if image_url and not _is_http_url(image_url):
            raise ExportValidationError(f"EXPORT_SOURCE_BAD_IMAGE_URL: {sku}")
        last_seen = str(row.get("last_seen") or "").strip()
        if not last_seen:
            raise ExportValidationError(f"EXPORT_SOURCE_COLLECTION_DATE_MISSING: {sku}")
        if last_seen[:10] != export_date:
            raise ExportValidationError(f"EXPORT_SOURCE_COLLECTION_DATE_MISMATCH: {sku}/{last_seen}")
        if current <= 0:
            raise ExportValidationError(f"EXPORT_SOURCE_BAD_PRICE: {sku}/current_price")
        if original is not None and original <= 0:
            raise ExportValidationError(f"EXPORT_SOURCE_BAD_PRICE: {sku}/original_price")


def validate_spanish_source_fields(records: Iterable[dict[str, Any]]) -> None:
    """西语事实列含中文说明发生源污染，必须阻断正式 ES 导出。"""
    for row in records:
        sku = str(row.get("sku") or "").strip()
        for field in ("name_es", "cat1_es", "cat2_es", "spec_es", "unit_price", "desc_es", "details_es", "raw_tags"):
            value = _text(row.get(field))
            if value and _CJK_RE.search(value):
                raise ExportValidationError(f"EXPORT_SOURCE_NON_SPANISH_FACT: {sku}/{field}")


def build_es_rows(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """构建 Profile v1 西班牙语行；不使用任何中文字典。"""
    ordered = sorted(records, key=_sku_sort_key)
    return [
        {
            "图片": None,
            "编号": str(record.get("sku") or "").strip(),
            "标题": _text(record.get("name_es")),
            "分类1": _text(record.get("cat1_es")),
            "分类2": _text(record.get("cat2_es")),
            "规格": _text(record.get("spec_es")),
            "折后价": _required_price(record.get("current_price"), sku=str(record.get("sku")), field="current_price"),
            "原价": _display_original_price(record),
            "单价": _text(record.get("unit_price")),
            "描述": _text_or_none(record.get("desc_es")),
            "产品详情": _text_or_none(record.get("details_es")),
            "图片链接": _text_or_none(record.get("image_url")),
            "商品链接": _text(record.get("product_url")),
            "备注": _es_remarks(record),
        }
        for record in ordered
    ]


def validate_output_rows(rows: Iterable[dict[str, Any]]) -> None:
    expected = {"图片", "编号", "标题", "分类1", "分类2", "规格", "折后价", "原价", "单价",
                "描述", "产品详情", "图片链接", "商品链接", "备注"}
    seen: set[str] = set()
    for row in rows:
        if set(row) != expected:
            raise ExportValidationError("EXPORT_OUTPUT_SCHEMA_MISMATCH")
        sku = str(row["编号"] or "").strip()
        if not sku or sku in seen:
            raise ExportValidationError(f"EXPORT_OUTPUT_DUPLICATE_OR_EMPTY_SKU: {sku}")
        seen.add(sku)
        if not isinstance(row["折后价"], (float, int)):
            raise ExportValidationError(f"EXPORT_OUTPUT_PRICE_NOT_NUMERIC: {sku}")
        original = row["原价"]
        if original is not None and (not isinstance(original, (float, int)) or original <= row["折后价"]):
            raise ExportValidationError(f"EXPORT_OUTPUT_INVALID_ORIGINAL_PRICE: {sku}")
        if not _is_http_url(str(row["商品链接"] or "")):
            raise ExportValidationError(f"EXPORT_OUTPUT_BAD_PRODUCT_URL: {sku}")


def validate_zh_rows_against_source(rows: Iterable[dict[str, Any]], records: Iterable[dict[str, Any]]) -> None:
    """中文导出只能改变派生字段，SKU、价格与 URL 必须和正式西语事实完全一致。"""
    expected = {str(record.get("sku") or "").strip(): record for record in records}
    actual = {str(row.get("编号") or "").strip(): row for row in rows}
    if set(actual) != set(expected):
        raise ExportValidationError("ZH_EXPORT_SKU_SET_MISMATCH")
    for sku, record in expected.items():
        row = actual[sku]
        if row["折后价"] != _required_price(record.get("current_price"), sku=sku, field="current_price"):
            raise ExportValidationError(f"ZH_EXPORT_PRICE_MISMATCH: {sku}")
        if row["原价"] != _display_original_price(record):
            raise ExportValidationError(f"ZH_EXPORT_ORIGINAL_PRICE_MISMATCH: {sku}")
        if _text_or_none(row["图片链接"]) != _text_or_none(record.get("image_url")):
            raise ExportValidationError(f"ZH_EXPORT_IMAGE_URL_MISMATCH: {sku}")
        if _text(row["商品链接"]) != _text(record.get("product_url")):
            raise ExportValidationError(f"ZH_EXPORT_PRODUCT_URL_MISMATCH: {sku}")


def canonical_source_hash(records: Iterable[dict[str, Any]]) -> str:
    normalized = []
    for record in sorted(records, key=_sku_sort_key):
        normalized.append({field: _canonical_value(record.get(field)) for field in _SOURCE_HASH_FIELDS})
    payload = json.dumps(normalized, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _verify_written_workbook(path: Path, *, headers: list[str], expected_skus: set[str]) -> None:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        if workbook.sheetnames != ["商品全量"]:
            raise ExportValidationError("EXPORT_XLSX_SHEET_MISMATCH")
        ws = workbook["商品全量"]
        actual_headers = [cell.value for cell in ws[1]]
        if actual_headers != headers or ws.freeze_panes != "A2":
            raise ExportValidationError("EXPORT_XLSX_HEADER_OR_FREEZE_MISMATCH")
        if ws.auto_filter.ref != f"A1:N{len(expected_skus) + 1}":
            raise ExportValidationError("EXPORT_XLSX_FILTER_MISMATCH")
        sku_column = headers.index("编号") + 1
        actual_skus = {str(ws.cell(row=row, column=sku_column).value or "").strip() for row in range(2, ws.max_row + 1)}
        if actual_skus != expected_skus or len(actual_skus) != len(expected_skus):
            raise ExportValidationError("EXPORT_XLSX_SKU_SET_MISMATCH")
    finally:
        workbook.close()


def _read_csv_records(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


_DETAIL_EXPORT_FIELDS = ("name_es", "cat1_es", "cat2_es", "spec_es", "desc_es", "details_es", "product_url", "image_url")


def _detail_retry_ids(source: ExportSource) -> list[str]:
    """返回来源快照中实际存在的详情补抓目录，供 manifest 追溯。"""
    if source.directory is None:
        return []
    return [path.name for path in _iter_valid_detail_retry_dirs(source.directory)]


def _iter_valid_detail_retry_dirs(directory: Path) -> list[Path]:
    """只接受有完整成功报告的详情补抓目录。

    目录里可能同时保留失败/中断的历史尝试；这些尝试不能覆盖正式快照，
    否则会把不完整详情误当成已验证事实。
    """
    retry_root = directory / "detail_retries"
    if not retry_root.exists():
        return []
    valid: list[Path] = []
    for retry_dir in sorted(path for path in retry_root.iterdir() if path.is_dir()):
        report_path = retry_dir / "detail_retry_report.json"
        checkpoint = retry_dir / "detail_fetch.jsonl"
        if not report_path.exists() or not checkpoint.exists():
            continue
        try:
            report = json.loads(report_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(report, dict):
            continue
        planned = report.get("planned")
        completed = report.get("completed")
        pending_after = report.get("pending_after")
        if not bool(report.get("detail_retry_pass")):
            continue
        if planned is None or completed is None or pending_after is None:
            continue
        try:
            complete = int(planned) == int(completed) and int(pending_after) == 0
        except (TypeError, ValueError):
            complete = False
        if complete:
            valid.append(retry_dir)
    return valid


def _merge_detail_retries(directory: Path, records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """将已落盘的 detail-retry 证据合并到历史快照的商品事实。"""
    retry_dirs = _iter_valid_detail_retry_dirs(directory)
    if not retry_dirs:
        return records
    by_sku = {str(row.get("sku") or ""): row for row in records}
    for retry_dir in retry_dirs:
        checkpoint = retry_dir / "detail_fetch.jsonl"
        if not checkpoint.exists():
            continue
        try:
            lines = checkpoint.read_text(encoding="utf-8").splitlines()
        except OSError:
            continue
        for line in lines:
            try:
                payload = json.loads(line)
                sku = str(payload.get("sku") or "")
                detail = payload.get("detail") or {}
            except (json.JSONDecodeError, TypeError):
                continue
            if sku not in by_sku or not isinstance(detail, dict):
                continue
            for field in _DETAIL_EXPORT_FIELDS:
                value = detail.get(field)
                if value not in (None, ""):
                    by_sku[sku][field] = value
    return records


def _display_original_price(record: dict[str, Any]) -> float | None:
    current = _required_price(record.get("current_price"), sku=str(record.get("sku")), field="current_price")
    original = _optional_price(record.get("original_price"), sku=str(record.get("sku")), field="original_price")
    return original if original is not None and original > current else None


def _required_price(value: Any, *, sku: str, field: str) -> float:
    parsed = _optional_price(value, sku=sku, field=field)
    if parsed is None or not math.isfinite(parsed):
        raise ExportValidationError(f"EXPORT_SOURCE_BAD_PRICE: {sku}/{field}")
    return parsed


def _optional_price(value: Any, *, sku: str, field: str) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        parsed = parse_price(value)
    except (TypeError, ValueError) as exc:
        raise ExportValidationError(f"EXPORT_SOURCE_BAD_PRICE: {sku}/{field}") from exc
    return None if parsed is None else float(parsed)


def _es_remarks(record: dict[str, Any]) -> str:
    values = ["Estado: CURRENT"]
    if parse_bool_zh(record.get("is_new_badge")):
        values.append("Nuevo")
    if parse_bool_zh(record.get("promotion")):
        values.append("Promoción")
    if parse_bool_zh(record.get("sustainable")):
        values.append("Sostenible")
    discount = _optional_price(record.get("discount"), sku=str(record.get("sku")), field="discount")
    if discount is not None:
        values.append(f"Descuento: {discount:g}")
    raw_tags = _text_or_none(record.get("raw_tags"))
    if raw_tags:
        values.append(f"Etiquetas oficiales: {raw_tags}")
    return "；".join(values)


def _sku_sort_key(record: dict[str, Any]) -> tuple[int, int, str]:
    sku = str(record.get("sku") or "").strip()
    if sku.isdigit():
        return (0, int(sku), sku)
    return (1, 0, sku)


def _is_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def _validate_date(value: str) -> None:
    try:
        dt.date.fromisoformat(value)
    except ValueError as exc:
        raise ExportValidationError(f"EXPORT_DATE_INVALID: {value}") from exc


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile_mkstemp(path)
    os.close(fd)
    temp_path = Path(temporary)
    try:
        temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        temp_path.replace(path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _publish_export_pair(
    preview_path: Path,
    output_path: Path,
    manifest_path: Path,
    manifest: dict[str, Any],
) -> None:
    """Publish workbook and manifest as a recoverable pair.

    A failed manifest replacement restores the previous workbook/manifest
    pair, so a new workbook is never left beside an old provenance record.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_tmp: Path | None = None
    old_output_tmp: Path | None = None
    old_manifest_tmp: Path | None = None
    try:
        fd, manifest_name = tempfile.mkstemp(prefix=f".{manifest_path.stem}.", suffix=".tmp", dir=output_path.parent)
        os.close(fd)
        manifest_tmp = Path(manifest_name)
        manifest_tmp.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        if output_path.exists():
            fd, name = tempfile.mkstemp(prefix=f".{output_path.stem}.old.", suffix=".xlsx", dir=output_path.parent)
            os.close(fd)
            backup = Path(name)
            try:
                shutil.copy2(output_path, backup)
            except BaseException:
                if backup.exists():
                    backup.unlink()
                raise
            old_output_tmp = backup
        if manifest_path.exists():
            fd, name = tempfile.mkstemp(prefix=f".{manifest_path.stem}.old.", suffix=".json", dir=output_path.parent)
            os.close(fd)
            backup = Path(name)
            try:
                shutil.copy2(manifest_path, backup)
            except BaseException:
                if backup.exists():
                    backup.unlink()
                raise
            old_manifest_tmp = backup
        preview_path.replace(output_path)
        manifest_tmp.replace(manifest_path)
    except BaseException:
        if old_output_tmp and old_output_tmp.exists():
            old_output_tmp.replace(output_path)
        elif output_path.exists() and not old_output_tmp:
            output_path.unlink()
        if old_manifest_tmp and old_manifest_tmp.exists():
            old_manifest_tmp.replace(manifest_path)
        elif manifest_path.exists() and not old_manifest_tmp:
            manifest_path.unlink()
        raise
    finally:
        for path in (manifest_tmp, old_output_tmp, old_manifest_tmp):
            if path and path.exists():
                path.unlink()


def _sku_set_hash(skus: Iterable[str]) -> str:
    payload = "\n".join(sorted(str(sku) for sku in skus)).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _is_truthy(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def tempfile_mkstemp(path: Path) -> tuple[int, str]:
    import tempfile
    return tempfile.mkstemp(prefix=f".{path.stem}.", suffix=".tmp", dir=path.parent)


def _canonical_value(value: Any) -> str | float | int | bool | None:
    if value is None:
        return None
    if isinstance(value, (bool, int, float)):
        return value
    return str(value).strip()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _text_or_none(value: Any) -> str | None:
    text = _text(value)
    return text or None
