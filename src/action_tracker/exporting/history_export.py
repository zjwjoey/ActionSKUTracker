"""Standalone historical Presence export; it never derives today's status."""
from __future__ import annotations

import datetime as dt
import hashlib
import json
from pathlib import Path
from typing import Any

from .history import HistoryExportError, build_history_only_rows, load_presence_history
from .template1 import HISTORY_HEADERS, write_history_xlsx


def export_history(cfg: dict[str, Any], *, export_date: str) -> dict[str, Any]:
    """Export the configured historical Presence union as a single workbook."""
    _validate_date(export_date)
    try:
        history = load_presence_history(cfg)
        rows = build_history_only_rows(history)
    except HistoryExportError:
        raise
    output = Path(cfg["paths"]["exports"]) / f"{export_date.replace('-', '')}Action商品上下架明细.xlsx"
    temporary = output.with_name(f".{output.stem}.preview.xlsx")
    write_history_xlsx(temporary, history_rows=rows, history_dates=history.dates)
    try:
        verification = verify_history_xlsx(temporary, history_dates=history.dates, expected_skus={row["编号"] for row in rows})
        temporary.replace(output)
    finally:
        if temporary.exists():
            temporary.unlink()

    manifest_path = output.with_suffix(".manifest.json")
    manifest = {
        "template_id": "action_history_presence",
        "template_version": 1,
        "export_date": export_date,
        "history_union_sku_count": len(rows),
        "history_dates": list(history.dates),
        "seed_path": history.seed_path,
        "seed_row_count": history.seed_row_count,
        "seed_sha256": _file_hash(Path(history.seed_path)) if history.seed_path else None,
        "source_stats": [
            {**stat.__dict__, "sha256": _file_hash(Path(stat.path))}
            for stat in history.source_stats
        ],
        "validation_results": {"workbook": "PASS", "presence_values": "PASS"},
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return {"output": str(output), "manifest": str(manifest_path), "sku_count": len(rows), "date_count": len(history.dates), "verification": verification}


def verify_history_xlsx(path: Path, *, history_dates: tuple[str, ...], expected_skus: set[str]) -> dict[str, int]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        if workbook.sheetnames != ["商品上下架明细"]:
            raise HistoryExportError("HISTORY_EXPORT_SHEET_MISMATCH")
        ws = workbook["商品上下架明细"]
        headers = [cell.value for cell in ws[1]]
        expected_headers = list(HISTORY_HEADERS) + [_compact_date(date) for date in history_dates]
        if headers != expected_headers:
            raise HistoryExportError("HISTORY_EXPORT_HEADERS_MISMATCH")
        if ws.freeze_panes != "A2" or not ws.auto_filter.ref:
            raise HistoryExportError("HISTORY_EXPORT_FORMAT_MISMATCH")
        sku_col = headers.index("编号") + 1
        skus = [str(ws.cell(row=row, column=sku_col).value or "").strip() for row in range(2, ws.max_row + 1)]
        if set(skus) != expected_skus or len(skus) != len(set(skus)) or "" in skus:
            raise HistoryExportError("HISTORY_EXPORT_SKU_SET_MISMATCH")
        for date in history_dates:
            col = headers.index(_compact_date(date)) + 1
            values = [ws.cell(row=row, column=col).value for row in range(2, ws.max_row + 1)]
            if any(value not in (0, 1) for value in values):
                raise HistoryExportError(f"HISTORY_EXPORT_BAD_PRESENCE: {date}")
        return {"sku_count": len(skus), "date_count": len(history_dates)}
    finally:
        workbook.close()


def _file_hash(path: Path) -> str | None:
    if not path.exists():
        raise HistoryExportError(f"HISTORY_SOURCE_MISSING: {path}")
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _validate_date(value: str) -> None:
    try:
        dt.date.fromisoformat(value)
    except ValueError as exc:
        raise HistoryExportError(f"HISTORY_EXPORT_DATE_INVALID: {value}") from exc


def _compact_date(value: str) -> str:
    return f"{value[2:4]}.{value[5:7]}.{value[8:10]}"
