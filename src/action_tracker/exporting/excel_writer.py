"""导出专用 Excel 写入器，不复用 Master writer，也绝不修改 Master。"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_catalog_xlsx(
    path: Path,
    *,
    headers: list[str],
    rows: Iterable[dict[str, Any]],
    workbook_format: dict[str, Any],
) -> None:
    """原子写入单工作表商品清单。调用方负责所有业务校验。"""
    materialized = list(rows)
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = str(workbook_format.get("sheet_name") or "商品全量")
    ws.append(headers)
    for row in materialized:
        ws.append([row.get(header) for header in headers])

    header_cfg = workbook_format.get("header") or {}
    fill = PatternFill("solid", fgColor=str(header_cfg.get("fill") or "1F4E78"))
    font = Font(bold=bool(header_cfg.get("bold", True)), color=str(header_cfg.get("font_color") or "FFFFFF"))
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = str(workbook_format.get("freeze_panes") or "A2")
    if workbook_format.get("auto_filter", True):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(materialized) + 1}"

    index = {header: pos + 1 for pos, header in enumerate(headers)}
    text_wrap = set((workbook_format.get("body") or {}).get("wrap_text_columns") or [])
    max_row_height = float((workbook_format.get("body") or {}).get("max_row_height") or 405)
    hyperlink_labels = dict(workbook_format.get("hyperlink_display_text") or {})
    widths = {
        "图片": 12, "编号": 14, "标题": 28, "分类1": 16, "分类2": 18, "规格": 26,
        "折后价": 13, "原价": 13, "单价": 16, "描述": 48, "产品详情": 56,
        "图片链接": 42, "商品链接": 58, "备注": 34,
    }
    for header, col in index.items():
        ws.column_dimensions[get_column_letter(col)].width = widths.get(header, 18)

    price_format = str((workbook_format.get("price") or {}).get("number_format") or "€#,##0.00")
    for row_no in range(2, len(materialized) + 2):
        for header, col in index.items():
            cell = ws.cell(row=row_no, column=col)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=header in text_wrap,
            )
            if header == "编号" and cell.value is not None:
                cell.number_format = "@"
            elif header in {"折后价", "原价"} and cell.value is not None:
                cell.number_format = price_format
            elif header in {"图片链接", "商品链接"} and cell.value:
                target = str(cell.value)
                cell.value = str(hyperlink_labels.get(header) or target)
                cell.hyperlink = target
                cell.style = "Hyperlink"
        description_lines = max(
            (_wrapped_line_count(ws.cell(row=row_no, column=index[header]).value, widths[header])
             for header in ("描述", "产品详情") if header in index),
            default=1,
        )
        ws.row_dimensions[row_no].height = min(max_row_height, max(20, 15 * description_lines + 4))

    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(prefix=f".{path.stem}.", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    temp_path = Path(temporary)
    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    finally:
        workbook.close()
        if temp_path.exists():
            temp_path.unlink()


def _wrapped_line_count(value: Any, width: int) -> int:
    """按列宽估算行高，避免描述/详情在常见 Excel 缩放下被截断。"""
    if value is None:
        return 1
    text = str(value)
    if not text:
        return 1
    return sum(max(1, (len(line) + width - 1) // width) for line in text.splitlines() or [""])
