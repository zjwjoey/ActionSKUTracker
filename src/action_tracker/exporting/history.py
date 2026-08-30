"""只读历史 Presence 构建：供 Template 1 和独立历史导出复用。"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import yaml


class HistoryExportError(ValueError):
    """历史来源缺失、结构不一致或 Presence 值非法。"""


PresenceValue = int | str
PRESENCE_UNKNOWN = "UNKNOWN"


@dataclass(frozen=True)
class HistorySourceStat:
    date: str
    path: str
    raw_rows: int
    unique_skus: int
    duplicate_rows: int
    presence_capability: bool
    absence_capability: bool
    observation_complete: bool
    evidence_level: str
    status: str


@dataclass(frozen=True)
class PresenceHistory:
    dates: tuple[str, ...]
    presence_by_sku: dict[str, dict[str, PresenceValue]]
    latest_by_sku: dict[str, dict[str, Any]]
    seed_by_sku: dict[str, dict[str, Any]]
    source_stats: tuple[HistorySourceStat, ...]
    seed_path: str | None
    seed_row_count: int


def build_history_only_rows(history: PresenceHistory) -> list[dict[str, Any]]:
    """Build the historical SKU union without inventing a current observation."""
    return build_presence_rows(history)


_DATE_HEADER_RE = re.compile(r"^(\d{2})\.(\d{2})\.(\d{2})$")


def load_presence_history(cfg: dict[str, Any]) -> PresenceHistory:
    config_path = Path(cfg.get("history_sources_path") or Path(cfg["project_root"]) / "config" / "history_sources.yaml")
    if not config_path.exists():
        raise HistoryExportError(f"HISTORY_CONFIG_MISSING: {config_path}")
    raw = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    seed = raw.get("seed") or {}
    presence: dict[str, dict[str, PresenceValue]] = {}
    latest: dict[str, dict[str, Any]] = {}
    seed_rows: dict[str, dict[str, Any]] = {}
    dates: set[str] = set()
    stats: list[HistorySourceStat] = []
    seed_path: str | None = None
    seed_row_count = 0
    source_by_date: dict[str, HistorySourceStat] = {}

    if seed.get("path"):
        path = _resolve_path(cfg, seed["path"])
        seed_path = str(path)
        seed_records, seed_dates = _read_seed(path, seed)
        dates.update(seed_dates)
        seed_row_count = len(seed_records)
        seed_cfg = _capabilities(seed, source_name="seed")
        seed_skus = {record["sku"] for record in seed_records}
        for date in seed_dates:
            source_by_date[date] = HistorySourceStat(
                date, str(path), seed_row_count, len(seed_skus), seed_row_count - len(seed_skus), *seed_cfg,
            )
        for record in seed_records:
            sku = record["sku"]
            seed_rows[sku] = record
            presence.setdefault(sku, {}).update(record["presence"])
            latest.setdefault(sku, {}).update(record.get("fields") or {})

    seen_dates = set(dates)
    for source in raw.get("sources") or []:
        date = _parse_iso_date(source.get("date"))
        if date in seen_dates:
            # The seed is an explicitly confirmed snapshot. Preserve its date
            # instead of silently letting a differently formatted source win.
            continue
        path = _resolve_path(cfg, source.get("path"))
        capabilities = _capabilities(source, source_name=date)
        records = _read_source(path, source)
        sku_values = [row["sku"] for row in records]
        unique = set(sku_values)
        for record in records:
            sku = record["sku"]
            presence.setdefault(sku, {})[date] = 1
            fields = record.get("fields") or {}
            if fields:
                latest.setdefault(sku, {}).update(fields)
        dates.add(date)
        seen_dates.add(date)
        stat = HistorySourceStat(
            date, str(path), len(records), len(unique), len(records) - len(unique), *capabilities,
        )
        stats.append(stat)
        source_by_date[date] = stat

    # Fill absence only when the source explicitly proves completeness and
    # absence capability. Partial sources remain UNKNOWN; not observed is not 0.
    all_skus = set(presence)
    for date, stat in source_by_date.items():
        default: PresenceValue = 0 if stat.absence_capability and stat.observation_complete else PRESENCE_UNKNOWN
        for sku in all_skus:
            presence.setdefault(sku, {}).setdefault(date, default)

    return PresenceHistory(
        dates=tuple(sorted(dates)),
        presence_by_sku=presence,
        latest_by_sku=latest,
        seed_by_sku=seed_rows,
        source_stats=tuple(sorted(source_by_date.values(), key=lambda item: item.date)),
        seed_path=seed_path,
        seed_row_count=seed_row_count,
    )


def build_presence_rows(
    history: PresenceHistory,
    *,
    export_date: str | None = None,
    current_records: list[dict[str, Any]] | None = None,
    zh_rows: list[dict[str, Any]] | None = None,
    dictionary: Any | None = None,
) -> list[dict[str, Any]]:
    """合并历史、当前 CURRENT 和字典展示字段；过去日期永不由生命周期反推。"""
    if not history.dates:
        raise HistoryExportError("HISTORY_DATES_EMPTY")
    current_records = current_records or []
    zh_rows = zh_rows or []
    current_by_sku = {str(row.get("sku") or "").strip(): row for row in current_records}
    zh_by_sku = {str(row.get("编号") or "").strip(): row for row in zh_rows}
    all_skus = set(history.presence_by_sku) | set(current_by_sku)
    dates = list(history.dates)
    if export_date is not None and export_date not in dates:
        dates.append(export_date)
        dates.sort()
    rows: list[dict[str, Any]] = []
    for sku in sorted(all_skus, key=_sku_key):
        current = current_by_sku.get(sku, {})
        old = history.latest_by_sku.get(sku, {})
        seed = history.seed_by_sku.get(sku, {})
        zh = zh_by_sku.get(sku, {})
        # Current facts win; then latest historical fact; then confirmed seed.
        name_es = _first(current.get("name_es"), old.get("name_es"), seed.get("name_es"))
        product_url = _first(current.get("product_url"), old.get("product_url"), seed.get("product_url"))
        image_url = _first(current.get("image_url"), old.get("image_url"), seed.get("image_url"))
        dictionary_product = dictionary.product_by_sku.get(sku, {}) if dictionary is not None else {}
        brand_id = _first(current.get("brand_id"), dictionary_product.get("brand_id"), old.get("brand_id"), seed.get("brand_id"))
        brand = ""
        if brand_id:
            brand = _first((dictionary.brand_by_id.get(str(brand_id)) or {}).get("canonical_name"), brand_id)
        row: dict[str, Any] = {
            "编号": sku,
            "中文品名": _first(zh.get("标题"), seed.get("name_zh")),
            "品牌": brand,
            "一级类目（中文）": _first(zh.get("分类1"), seed.get("cat1_zh")),
            "二级类目（中文）": _first(zh.get("分类2"), seed.get("cat2_zh")),
            "西班牙语品名": name_es,
            "规格（西语）": _first(current.get("spec_es"), old.get("spec_es"), seed.get("spec_es")),
            "商品链接": product_url,
        }
        row["图片链接"] = image_url
        row["presence"] = dict(history.presence_by_sku.get(sku, {}))
        if export_date is not None:
            row["presence"][export_date] = 1 if sku in current_by_sku else 0
        for date in dates:
            value = row["presence"].get(date, PRESENCE_UNKNOWN)
            if value not in (0, 1, PRESENCE_UNKNOWN):
                raise HistoryExportError(f"HISTORY_BAD_PRESENCE: {sku}/{date}")
            row[date] = value
        row["首次出现日期"] = next((date for date in dates if row[date] == 1), "")
        row["最近出现日期"] = next((date for date in reversed(dates) if row[date] == 1), "")
        row["当前状态"] = ("在售" if row[export_date] == 1 else "不在售") if export_date is not None else ""
        rows.append(row)
    return rows


def _capabilities(config: dict[str, Any], *, source_name: str) -> tuple[bool, bool, bool, str, str]:
    required = ("presence_capability", "absence_capability", "observation_complete", "evidence_level")
    if any(key not in config for key in required):
        raise HistoryExportError(f"HISTORY_CAPABILITY_CONFIG_MISSING: {source_name}")
    values = tuple(config[key] for key in required)
    if not all(isinstance(value, bool) for value in values[:3]):
        raise HistoryExportError(f"HISTORY_CAPABILITY_CONFIG_INVALID: {source_name}")
    evidence = str(values[3]).strip().upper()
    if evidence not in {"A", "B", "C", "D"} or not values[0]:
        raise HistoryExportError(f"HISTORY_CAPABILITY_CONFIG_INVALID: {source_name}")
    status = "COMPLETE" if values[1] and values[2] else "PARTIAL"
    return values[0], values[1], values[2], evidence, status


def _read_seed(path: Path, seed_cfg: dict[str, Any]) -> tuple[list[dict[str, Any]], set[str]]:
    if not path.exists():
        raise HistoryExportError(f"HISTORY_SEED_MISSING: {path}")
    records = _read_workbook(path, seed_cfg.get("sheet"), seed_cfg.get("sku_header") or "编号")
    headers = list(records[0].keys()) if records else []
    date_columns: dict[str, str] = {}
    for header in headers:
        match = _DATE_HEADER_RE.match(str(header).strip())
        if match:
            date_columns[header] = f"20{match.group(1)}-{match.group(2)}-{match.group(3)}"
    if not date_columns:
        raise HistoryExportError(f"HISTORY_SEED_DATE_COLUMNS_MISSING: {path}")
    field_map = seed_cfg.get("fields") or {}
    output: list[dict[str, Any]] = []
    for raw in records:
        sku = str(raw.get(seed_cfg.get("sku_header") or "编号") or "").strip()
        if not sku:
            continue
        values: dict[str, int] = {}
        for header, date in date_columns.items():
            value = raw.get(header)
            if str(value).strip() not in {"0", "1"} and value not in {0, 1}:
                raise HistoryExportError(f"HISTORY_SEED_BAD_PRESENCE: {path}/{sku}/{header}")
            values[date] = int(value)
        fields = {key: raw.get(header) for key, header in field_map.items() if header in raw}
        output.append({"sku": sku, "presence": values, **fields, "fields": fields})
    return output, set(date_columns.values())


def _read_source(path: Path, source_cfg: dict[str, Any]) -> list[dict[str, Any]]:
    if not path.exists():
        raise HistoryExportError(f"HISTORY_SOURCE_MISSING: {path}")
    sku_header = str(source_cfg.get("sku_header") or "").strip()
    if not sku_header:
        raise HistoryExportError(f"HISTORY_SKU_HEADER_MISSING: {path}")
    records = _read_workbook(path, source_cfg.get("sheet"), sku_header)
    field_map = source_cfg.get("fields") or {}
    output: list[dict[str, Any]] = []
    for raw in records:
        sku = str(raw.get(sku_header) or "").strip()
        if not sku:
            continue
        fields = {key: raw.get(header) for key, header in field_map.items() if header in raw}
        output.append({"sku": sku, "fields": fields})
    return output


def _read_workbook(path: Path, sheet: str | None, sku_header: str) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet and sheet not in workbook.sheetnames:
            raise HistoryExportError(f"HISTORY_SHEET_MISSING: {path}/{sheet}")
        worksheet = workbook[sheet] if sheet else workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(iterator, ())]
        if sku_header not in headers:
            raise HistoryExportError(f"HISTORY_SKU_HEADER_NOT_FOUND: {path}/{sku_header}")
        return [dict(zip(headers, row)) for row in iterator]
    finally:
        workbook.close()


def _resolve_path(cfg: dict[str, Any], raw_path: Any) -> Path:
    if not raw_path:
        raise HistoryExportError("HISTORY_SOURCE_PATH_MISSING")
    path = Path(str(raw_path))
    return path if path.is_absolute() else Path(cfg["project_root"]) / path


def _parse_iso_date(value: Any) -> str:
    try:
        return dt.date.fromisoformat(str(value)).isoformat()
    except ValueError as exc:
        raise HistoryExportError(f"HISTORY_DATE_INVALID: {value}") from exc


def _first(*values: Any) -> Any:
    for value in values:
        if value is not None and str(value).strip():
            return value
    return ""


def _sku_key(value: str) -> tuple[int, int, str]:
    return (0, int(value), value) if value.isdigit() else (1, 0, value)
