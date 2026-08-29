"""Template 1 三表导出（第一版：历史 Presence + 两张无图清单）。"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .history import PresenceHistory


HISTORY_HEADERS = ("序号", "编号", "中文品名", "图片链接", "商品链接")
CATALOG_HEADERS = (
    "图片", "编号", "标题", "分类1", "分类2", "规格", "折后价", "原价", "单价",
    "描述", "产品详情", "图片链接", "商品链接", "备注",
)


def write_template1_xlsx(
    path: Path,
    *,
    history_rows: list[dict[str, Any]],
    history_dates: tuple[str, ...],
    es_rows: list[dict[str, Any]],
    zh_rows: list[dict[str, Any]],
) -> None:
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "商品上下架明细"
    _write_history_sheet(first, history_rows, history_dates)
    _write_catalog_sheet(workbook.create_sheet("今日西班牙语清单"), es_rows)
    _write_catalog_sheet(workbook.create_sheet("今日中文清单"), zh_rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.tmp.xlsx")
    try:
        workbook.save(temporary)
        temporary.replace(path)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()


def write_history_xlsx(path: Path, *, history_rows: list[dict[str, Any]], history_dates: tuple[str, ...]) -> None:
    """Write the standalone historical Presence workbook using Template 1 formatting."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "商品上下架明细"
    _write_history_sheet(sheet, history_rows, history_dates)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.tmp.xlsx")
    try:
        workbook.save(temporary)
        temporary.replace(path)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()


def verify_template1_xlsx(path: Path, *, export_date: str, current_skus: set[str]) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        expected_names = ["商品上下架明细", "今日西班牙语清单", "今日中文清单"]
        if workbook.sheetnames != expected_names:
            raise ValueError("TEMPLATE1_SHEET_MISMATCH")
        es = _sheet_skus(workbook["今日西班牙语清单"])
        zh = _sheet_skus(workbook["今日中文清单"])
        if es != current_skus or zh != current_skus or es != zh:
            raise ValueError("TEMPLATE1_CURRENT_SKU_SET_MISMATCH")
        history = workbook["商品上下架明细"]
        date_header = _compact_date(export_date)
        headers = [cell.value for cell in history[1]]
        if date_header not in headers:
            raise ValueError("TEMPLATE1_DATE_COLUMN_MISSING")
        date_col = headers.index(date_header) + 1
        ones = {
            str(history.cell(row=row, column=2).value or "").strip()
            for row in range(2, history.max_row + 1)
            if history.cell(row=row, column=date_col).value == 1
        }
        if ones != current_skus:
            raise ValueError("TEMPLATE1_PRESENCE_SET_MISMATCH")
        for name in expected_names:
            sheet = workbook[name]
            if sheet.freeze_panes != "A2" or not sheet.auto_filter.ref:
                raise ValueError(f"TEMPLATE1_FORMAT_MISMATCH: {name}")
        return {"history_sku_count": history.max_row - 1, "current_sku_count": len(current_skus), "presence_one_count": len(ones)}
    finally:
        workbook.close()


def _write_history_sheet(ws: Any, rows: list[dict[str, Any]], dates: tuple[str, ...]) -> None:
    headers = list(HISTORY_HEADERS) + [_compact_date(date) for date in dates]
    ws.append(headers)
    for number, row in enumerate(rows, 1):
        values = [
            number, row.get("编号"), row.get("中文品名"), row.get("图片链接"), row.get("商品链接"),
        ] + [row.get(date, 0) for date in dates]
        ws.append(values)
    _format_sheet(ws, wrap_columns={"中文品名", "一级类目（中文）", "二级类目（中文）", "西班牙语品名", "规格（西语）"})
    for col in (4, 5):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for item in cell:
                _set_hyperlink(item)
    for col in range(len(HISTORY_HEADERS) + 1, len(headers) + 1):
        for row_no in range(2, ws.max_row + 1):
            ws.cell(row=row_no, column=col).number_format = "0"


def _write_catalog_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    ws.append(list(CATALOG_HEADERS))
    for row in rows:
        ws.append([row.get(header) for header in CATALOG_HEADERS])
    _format_sheet(ws, wrap_columns={"标题", "分类1", "分类2", "规格", "描述", "产品详情", "备注"})
    headers = list(CATALOG_HEADERS)
    for header in ("图片链接", "商品链接"):
        col = headers.index(header) + 1
        for row_no in range(2, ws.max_row + 1):
            _set_hyperlink(ws.cell(row=row_no, column=col))
    for row_no in range(2, ws.max_row + 1):
        for header in ("折后价", "原价"):
            cell = ws.cell(row=row_no, column=headers.index(header) + 1)
            if cell.value is not None:
                cell.number_format = "€#,##0.00"


def _format_sheet(ws: Any, *, wrap_columns: set[str]) -> None:
    headers = [cell.value for cell in ws[1]]
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    widths = {"序号": 9, "编号": 14, "中文品名": 28, "品牌": 18, "一级类目（中文）": 18, "二级类目（中文）": 20,
              "西班牙语品名": 30, "规格（西语）": 24, "图片链接": 42, "商品链接": 58, "首次出现日期": 14,
              "最近出现日期": 14, "当前状态": 12, "图片": 12, "标题": 28, "分类1": 16, "分类2": 18,
              "规格": 26, "折后价": 13, "原价": 13, "单价": 16, "描述": 48, "产品详情": 56, "备注": 34}
    for index, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(str(header), 14)
    for row_no in range(2, ws.max_row + 1):
        for index, header in enumerate(headers, 1):
            cell = ws.cell(row=row_no, column=index)
            cell.alignment = Alignment(vertical="top", wrap_text=str(header) in wrap_columns)
            if str(header) == "编号":
                cell.number_format = "@"
        ws.row_dimensions[row_no].height = 20


def _set_hyperlink(cell: Any) -> None:
    if cell.value and str(cell.value).startswith(("http://", "https://")):
        cell.hyperlink = str(cell.value)
        cell.style = "Hyperlink"


def _sheet_skus(ws: Any) -> set[str]:
    headers = [cell.value for cell in ws[1]]
    col = headers.index("编号") + 1
    return {str(ws.cell(row=row, column=col).value or "").strip() for row in range(2, ws.max_row + 1)}


def _compact_date(value: Any) -> str:
    text = str(value or "")[:10]
    if len(text) == 10 and text[4] == "-" and text[7] == "-":
        return f"{text[2:4]}.{text[5:7]}.{text[8:10]}"
    return text
