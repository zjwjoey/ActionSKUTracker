"""Auditable dictionary Apply preview and the disabled-by-default commit gate."""
from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import os
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

import openpyxl

from .dictionary import DICTIONARY_BASELINE_FILENAMES
from .dictionary_resolver import RecordResolution, resolve_record
from .excel.reader import ES_MAP, ZH_MAP, load_current
from .excel.writer import commit_master, restore_master_from_backup, stage_master
from .exporting.dictionary_join import load_dictionary_context
from .exporting.profiles import load_profile
from .exporting.service import ExportValidationError, resolve_formal_source
from .services.runtime import RunLock


class DictionaryApplyError(RuntimeError):
    """字典 Apply 预览或 Gate 失败。"""


@dataclass(frozen=True)
class CommitOutcome:
    master_hash: str
    backup_path: Path


# Apply is limited to derived Chinese fields.  Official Spanish facts,
# prices, URLs, lifecycle and Presence are immutable safety invariants.
ALLOWLIST = {
    "name": "name_zh", "cat1": "cat1_zh", "cat2": "cat2_zh", "spec": "spec_zh",
    "description": "desc_zh", "details": "details_zh",
}
IMMUTABLE_FIELDS = (
    "sku", "canonical_id", "name_es", "cat1_es", "cat2_es", "spec_es", "current_price",
    "original_price", "unit_price", "product_url", "image_url", "status", "first_seen", "last_seen",
    "is_new_badge", "promotion", "sustainable", "discount", "raw_tags", "presence",
)
APPROVED_FIELD_STATUSES = frozenset({
    "APPROVED", "HUMAN_APPROVED", "AUTO_APPROVED", "CONFIRMED", "LOCKED", "HUMAN_REVIEWED",
})
APPROVED_FIELD_SOURCES = frozenset({
    "manual_override", "product_dictionary", "category_dictionary", "term_dictionary",
    "model_cache", "master_zh",
})


def _field_applyable(result: Any) -> bool:
    """Fail closed for explicit field-level approval metadata.

    Existing dictionary sources predate the metadata column and are trusted by
    their resolver status.  If a resolution carries an approval status, it is
    authoritative for that field only; a pending/rejected field can never be
    applied merely because the surrounding SKU is ``AUTO_READY``.
    """
    if result is None or result.status != "READY" or result.source not in APPROVED_FIELD_SOURCES:
        return False
    approval = str(getattr(result, "approval_status", "") or "").strip().upper()
    return not approval or approval in APPROVED_FIELD_STATUSES


def dictionary_apply(cfg: dict[str, Any], *, run_id: str, dry_run: bool = True) -> dict[str, Any]:
    _recover_interrupted_apply(cfg, run_id)
    production_enabled = _production_enabled(cfg)
    if not dry_run and not production_enabled:
        raise DictionaryApplyError("PRODUCTION_DICTIONARY_APPLY_DISABLED")
    profile = load_profile(cfg, language="es", no_images=True)
    try:
        source = resolve_formal_source(cfg, export_date=_run_date(cfg, run_id), requested_run_id=run_id, profile=profile)
    except ExportValidationError as exc:
        raise DictionaryApplyError(str(exc)) from exc
    context = load_dictionary_context(cfg)
    resolutions = [resolve_record(dict(record), context) for record in source.records]
    master_path = Path(cfg["paths"]["master"])
    before_hash = _master_hash(cfg)
    master_records = _load_apply_master_records(master_path) if master_path.exists() else {}
    # Compare the staged Apply candidate with the exact Master that will be
    # replaced.  Comparing the observation snapshot with Master is invalid:
    # they may be different dates and therefore legitimately differ in
    # last_seen, prices, tags, or SKU membership.
    candidate_records = _build_allowlisted_records(master_records, resolutions)
    immutable_count = _immutable_diff_count(master_records, candidate_records)
    output_dir = Path(cfg["paths"]["dictionary"]) / "apply" / run_id
    output_dir.mkdir(parents=True, exist_ok=True)
    preview_rows, summary = _preview_rows(source.records, resolutions)
    review_rows = _review_rows(resolutions)
    diff_headers = ["sku", "field", "old_value", "new_value", "source", "resolver_status", "reason"]
    _write_csv(output_dir / "apply_preview.csv", diff_headers, preview_rows)
    _write_csv(output_dir / "field_diff.csv", diff_headers, preview_rows)
    _write_csv(output_dir / "review_required.csv", ["sku", "readiness", "field", "status", "source", "reason"], review_rows)
    auto_count = sum(item.readiness == "AUTO_READY" for item in resolutions)
    review_count = sum(item.readiness == "REVIEW_REQUIRED" for item in resolutions)
    blocked_count = sum(item.readiness == "SOURCE_BLOCKED" for item in resolutions)
    run_date = _run_date(cfg, run_id)
    manifest = {
        "run_id": run_id, "run_date": run_date, "date": run_date, "dry_run": dry_run,
        "master_hash_before": before_hash, "temporary_master_hash": before_hash if dry_run else None,
        "master_hash_after_if_committed": None, "master_hash_after": None,
        "dictionary_hash": context.content_hash, "dictionary_baseline_hash": _dictionary_baseline_hash(cfg),
        "total_current_skus": len(source.records), "auto_ready_count": auto_count,
        "review_required_count": review_count, "source_blocked_count": blocked_count,
        "preview_field_count": len(preview_rows) + summary["unchanged_field_count"],
        "actual_changed_field_count": summary["actual_changed_field_count"],
        "unchanged_field_count": summary["unchanged_field_count"], "immutable_fact_change_count": immutable_count,
        "field_change_summary": summary["fields"], "applied_sku_count": len({row["sku"] for row in preview_rows}),
        "applied_field_count": len(preview_rows), "formal_write": False,
        "production_enabled": production_enabled, "committed": False, "commit_state": "DRY_RUN",
        "backup_path": None, "rollback_status": "NOT_APPLICABLE",
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
    }
    if dry_run:
        _write_json(output_dir / "apply_manifest.json", manifest)
    else:
        errors = _gate_errors(cfg, source, resolutions, context, master_records, before_hash)
        if errors:
            raise DictionaryApplyError("DICTIONARY_APPLY_GATE_REJECTED: " + ",".join(errors))
        # Persist a recoverable intent before replacing Master.  If the process
        # stops after replace, the backup path identifies the only valid rollback.
        manifest.update({"commit_state": "PENDING", "rollback_status": "NOT_NEEDED_YET"})
        _write_json(output_dir / "apply_manifest.json", manifest)
        outcome: CommitOutcome | None = None
        try:
            outcome = _commit_allowlisted(
                cfg, source.records, resolutions, master_records, before_hash, run_id,
                on_backup_ready=lambda backup: _write_json(
                    output_dir / "apply_manifest.json",
                    {**manifest, "backup_path": str(backup), "commit_state": "READY_TO_REPLACE"},
                ),
            )
            manifest.update({
                "temporary_master_hash": outcome.master_hash,
                "master_hash_after_if_committed": outcome.master_hash,
                "master_hash_after": outcome.master_hash,
                "formal_write": True, "committed": True, "commit_state": "COMMITTED",
                "backup_path": str(outcome.backup_path), "rollback_status": "NOT_NEEDED",
            })
            _write_json(output_dir / "apply_manifest.json", manifest)
        except Exception as exc:
            if outcome is not None:
                _restore_after_manifest_failure(cfg, outcome, before_hash, manifest, output_dir)
            raise DictionaryApplyError(f"DICTIONARY_APPLY_COMMIT_FAILED: {exc}") from exc
    return {"run_id": run_id, "dry_run": dry_run, "output_dir": str(output_dir), **{
        key: manifest[key] for key in ("auto_ready_count", "applied_sku_count", "applied_field_count", "review_required_count", "source_blocked_count", "actual_changed_field_count")
    }}


def _preview_rows(records: Iterable[dict[str, Any]], resolutions: list[RecordResolution]) -> tuple[list[dict[str, str]], dict[str, Any]]:
    by_sku = {str(record.get("sku") or ""): record for record in records}
    rows: list[dict[str, str]] = []
    fields = {field: {"ACTUAL_CHANGE": 0, "UNCHANGED": 0, "BLANK_TO_VALUE": 0, "VALUE_TO_NEW_VALUE": 0} for field in ALLOWLIST}
    for item in resolutions:
        if item.readiness != "AUTO_READY":
            continue
        record = by_sku[item.sku]
        for field, target in ALLOWLIST.items():
            result = item.fields.get(field)
            if not _field_applyable(result):
                continue
            old, new = str(record.get(target) or "").strip(), str(result.value or "").strip()
            if old == new:
                fields[field]["UNCHANGED"] += 1
                continue
            fields[field]["ACTUAL_CHANGE"] += 1
            kind = "BLANK_TO_VALUE" if not old else "VALUE_TO_NEW_VALUE"
            fields[field][kind] += 1
            rows.append({"sku": item.sku, "field": target, "old_value": old, "new_value": new,
                         "source": result.source, "resolver_status": item.readiness, "reason": "AUTO_READY 字段预览"})
    return rows, {"fields": fields, "actual_changed_field_count": len(rows),
                  "unchanged_field_count": sum(item["UNCHANGED"] for item in fields.values())}


def _review_rows(resolutions: list[RecordResolution]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in resolutions:
        if item.readiness == "AUTO_READY":
            continue
        for reason in item.review_reasons or (item.readiness,):
            field = _field_for_reason(reason)
            result = item.fields.get(field)
            rows.append({"sku": item.sku, "readiness": item.readiness, "field": field,
                         "status": result.status if result else "BLOCKED", "source": result.source if result else "source_quality", "reason": reason})
    return rows


def _field_for_reason(reason: str) -> str:
    return {"NAME_REVIEW": "name", "SPEC_REVIEW": "spec", "CATEGORY_REVIEW": "cat1", "BRAND_CANDIDATE": "brand", "TERM_REVIEW": "unit_price"}.get(reason, "")


def _gate_errors(cfg: dict[str, Any], source: Any, resolutions: list[RecordResolution], context: Any,
                master_records: dict[str, dict[str, Any]], before_hash: str | None) -> list[str]:
    errors: list[str] = []
    report = _read_json(source.directory / "run_report.json") if source.directory else {}
    qa = _read_json(source.directory / "qa_report.json") if source.directory else {}
    if report.get("dry_run") is not False:
        errors.append("RUN_NOT_FORMAL")
    if report.get("commit_status") != "FULL_COMMIT":
        errors.append("COMMIT_STATUS_NOT_FULL_COMMIT")
    allowed = set(load_profile(cfg, language="es", no_images=True).source_policy.get("allowed_qa_states") or ("PASS", "PASS_PRESENCE_ONLY"))
    if not qa.get("passed") or qa.get("state") not in allowed:
        errors.append("QA_NOT_PASS")
    audit_files = sorted(Path(cfg["paths"]["dictionary"]).glob("audit_report_*.json"), reverse=True)
    audit = _read_json(audit_files[0]) if audit_files else {}
    if not audit_files or int((audit.get("summary") or {}).get("fail") or 0) != 0:
        errors.append("DICTIONARY_AUDIT_NOT_PASS")
    if not _dictionary_binding_is_valid(cfg, context):
        errors.append("DICTIONARY_HASH_BINDING_INVALID")
    if str(audit.get("dictionary_hash") or "") != str(getattr(context, "content_hash", "")):
        errors.append("DICTIONARY_AUDIT_HASH_MISMATCH")
    if str(audit.get("baseline_manifest_hash") or "") != str(_dictionary_baseline_hash(cfg) or ""):
        errors.append("DICTIONARY_AUDIT_MANIFEST_HASH_MISMATCH")
    run_date = str((_read_json(source.directory / "run_report.json") if source.directory else {}).get("run_date") or "")
    if run_date and str(audit.get("latest_date") or "") < run_date:
        errors.append("DICTIONARY_AUDIT_STALE")
    if any(item.readiness != "AUTO_READY" for item in resolutions):
        errors.append("RESOLVER_NOT_AUTO_READY")
    if _require_confirmed_brands_for_apply(cfg) and any(item.brand_classification == "PROVISIONAL" for item in resolutions):
        errors.append("PROVISIONAL_BRAND_NOT_ALLOWED_FOR_APPLY")
    if set(master_records) != {str(record.get("sku") or "") for record in source.records}:
        errors.append("TARGET_CURRENT_SKU_SET_MISMATCH")
    if not before_hash or _master_hash_from_path(Path(cfg["paths"]["master"])) != before_hash:
        errors.append("MASTER_CHANGED_CONCURRENTLY")
    if not _production_enabled(cfg):
        errors.append("PRODUCTION_DICTIONARY_APPLY_DISABLED")
    return errors


def _commit_allowlisted(cfg: dict[str, Any], records: Iterable[dict[str, Any]], resolutions: list[RecordResolution],
                        master_records: dict[str, dict[str, Any]], before_hash: str, run_id: str,
                        on_backup_ready: Callable[[Path], None] | None = None) -> CommitOutcome:
    updated = _build_allowlisted_records(master_records, resolutions)
    lock = RunLock(Path(cfg["paths"]["state"]), stale_minutes=int((cfg.get("run") or {}).get("lock_stale_minutes", 180)))
    lock.acquire(run_id, command="dictionary-apply")
    tmp: Path | None = None
    backup: Path | None = None
    replaced = False
    try:
        if _master_hash_from_path(Path(cfg["paths"]["master"])) != before_hash:
            raise DictionaryApplyError("MASTER_CHANGED_CONCURRENTLY")
        staged = stage_master(cfg, updated_records=updated, price_events=[], event_events=[], return_backup=True)
        tmp, backup = staged
        if on_backup_ready is not None:
            on_backup_ready(backup)
        _assert_master_safe(Path(cfg["paths"]["master"]), tmp)
        commit_master(tmp, Path(cfg["paths"]["master"]))
        tmp = None
        replaced = True
        _assert_master_safe(backup, Path(cfg["paths"]["master"]))
        master_hash = _master_hash_from_path(Path(cfg["paths"]["master"]))
        if not master_hash:
            raise DictionaryApplyError("MASTER_POST_COMMIT_HASH_MISSING")
        return CommitOutcome(master_hash=master_hash, backup_path=backup)
    except Exception as exc:
        if replaced and backup is not None:
            try:
                restore_master_from_backup(backup, Path(cfg["paths"]["master"]))
                if _master_hash_from_path(Path(cfg["paths"]["master"])) != before_hash:
                    raise DictionaryApplyError("MASTER_ROLLBACK_HASH_MISMATCH")
            except Exception as rollback_exc:
                raise DictionaryApplyError(
                    f"MASTER_STATE_UNKNOWN_AFTER_COMMIT_FAILURE: backup={backup}; rollback_error={rollback_exc}"
                ) from exc
            raise DictionaryApplyError(f"POST_COMMIT_FAILURE_ROLLED_BACK: backup={backup}; error={exc}") from exc
        raise
    finally:
        lock.release()
        if tmp and tmp.exists():
            tmp.unlink()


def _assert_master_safe(before: Path, staged: Path) -> None:
    old, new = _load_apply_master_records(before), _load_apply_master_records(staged)
    if list(old) != list(new) or set(old) != set(new):
        raise DictionaryApplyError("IMMUTABLE_SKU_ORDER_OR_SET_CHANGED")
    for sku in old:
        for field in IMMUTABLE_FIELDS:
            if _norm(old[sku].get(field)) != _norm(new[sku].get(field)):
                raise DictionaryApplyError(f"IMMUTABLE_FACT_CHANGED:{sku}:{field}")


def _build_allowlisted_records(
    master_records: dict[str, dict[str, Any]], resolutions: Iterable[RecordResolution],
) -> dict[str, dict[str, Any]]:
    """Build the exact Master candidate using only the derived-field allowlist."""
    updated = {sku: dict(record) for sku, record in master_records.items()}
    for item in resolutions:
        if item.readiness != "AUTO_READY" or item.sku not in updated:
            continue
        for field, target in ALLOWLIST.items():
            result = item.fields.get(field)
            if _field_applyable(result):
                updated[item.sku][target] = result.value
    return updated


def _immutable_diff_count(
    before_records: dict[str, dict[str, Any]], after_records: dict[str, dict[str, Any]],
) -> int:
    """Count immutable changes between the pre-Apply Master and its candidate.

    This deliberately does not compare a dated observation snapshot with the
    current Master.  Snapshot-vs-Master differences are expected between runs;
    only the proposed replacement workbook is relevant to this invariant.
    """
    count = 0
    for sku, before in before_records.items():
        after = after_records.get(sku)
        if not after:
            count += 1
            continue
        count += sum(_norm(before.get(field)) != _norm(after.get(field)) for field in IMMUTABLE_FIELDS)
    count += len(set(after_records) - set(before_records))
    return count


def _norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _production_enabled(cfg: dict[str, Any]) -> bool:
    value = (cfg.get("dictionary_apply") or {}).get("production_enabled", False)
    if isinstance(value, bool):
        return value
    raise DictionaryApplyError("PRODUCTION_DICTIONARY_APPLY_CONFIG_INVALID")


def _require_confirmed_brands_for_apply(cfg: dict[str, Any]) -> bool:
    value = (cfg.get("dictionary_apply") or {}).get("require_confirmed_brands_for_apply", True)
    if isinstance(value, bool):
        return value
    raise DictionaryApplyError("DICTIONARY_APPLY_BRAND_POLICY_CONFIG_INVALID")


def _dictionary_binding_is_valid(cfg: dict[str, Any], context: Any) -> bool:
    """Bind both the selected dictionary and Git baseline files to one manifest."""
    baseline = Path(cfg["paths"]["dictionary_baseline"])
    manifest = _read_json(baseline / "baseline_manifest.json")
    files = manifest.get("files") if isinstance(manifest, dict) else None
    if not isinstance(files, dict) or set(files) != set(DICTIONARY_BASELINE_FILENAMES):
        return False
    for filename in DICTIONARY_BASELINE_FILENAMES:
        expected = str((files.get(filename) or {}).get("sha256") or "")
        baseline_file = baseline / filename
        selected_directory = getattr(context, "directory", None)
        if selected_directory is None:
            return False
        selected_file = Path(selected_directory) / filename
        if not expected or not baseline_file.exists() or not selected_file.exists():
            return False
        if _hash(baseline_file) != expected or _hash(selected_file) != expected:
            return False
    return True


def _recover_interrupted_apply(cfg: dict[str, Any], run_id: str) -> None:
    """在下一次调用前处理上次替换后未完成 manifest 的恢复状态。"""
    manifest_path = Path(cfg["paths"]["dictionary"]) / "apply" / run_id / "apply_manifest.json"
    if not manifest_path.exists():
        return
    manifest = _read_json(manifest_path)
    if manifest.get("commit_state") not in {"PENDING", "READY_TO_REPLACE"}:
        return
    backup_value = str(manifest.get("backup_path") or "")
    before_hash = str(manifest.get("master_hash_before") or "")
    if not backup_value or not before_hash:
        raise DictionaryApplyError("MASTER_RECOVERY_REQUIRED: incomplete apply manifest")
    master = Path(cfg["paths"]["master"])
    current_hash = _master_hash_from_path(master)
    if current_hash == before_hash:
        manifest.update({"commit_state": "RECOVERED_NO_REPLACE", "committed": False, "rollback_status": "NOT_NEEDED"})
        _write_json(manifest_path, manifest)
        return
    backup = Path(backup_value)
    restore_master_from_backup(backup, master)
    if _master_hash_from_path(master) != before_hash:
        raise DictionaryApplyError("MASTER_RECOVERY_HASH_MISMATCH")
    manifest.update({"commit_state": "ROLLED_BACK", "committed": False, "rollback_status": "SUCCESS"})
    _write_json(manifest_path, manifest)


def _load_apply_master_records(path: Path) -> dict[str, dict[str, Any]]:
    """Read CURRENT only after schema, SKU uniqueness and ES/ZH set checks pass."""
    if not path.exists():
        raise DictionaryApplyError(f"MASTER_MISSING: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        records_by_sheet: dict[str, list[str]] = {}
        for sheet, mapping in (("01_SKU_ZH_CURRENT", ZH_MAP), ("02_SKU_ES_CURRENT", ES_MAP)):
            if sheet not in wb.sheetnames:
                raise DictionaryApplyError(f"MASTER_SCHEMA_MISSING_SHEET:{sheet}")
            ws = wb[sheet]
            headers = [str(cell.value or "").strip() for cell in ws[1]]
            missing = [header for header in mapping if header not in headers]
            if missing:
                raise DictionaryApplyError(f"MASTER_SCHEMA_MISSING_HEADERS:{sheet}:{','.join(missing)}")
            sku_index = headers.index("SKU")
            seen: set[str] = set()
            order: list[str] = []
            for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                sku = str(row[sku_index] if sku_index < len(row) and row[sku_index] is not None else "").strip()
                if not sku:
                    if any(value not in (None, "") for value in row):
                        raise DictionaryApplyError(f"MASTER_EMPTY_SKU:{sheet}:row={row_no}")
                    continue
                if sku in seen:
                    raise DictionaryApplyError(f"MASTER_DUPLICATE_SKU:{sheet}:{sku}:row={row_no}")
                seen.add(sku)
                order.append(sku)
            records_by_sheet[sheet] = order
        if set(records_by_sheet["01_SKU_ZH_CURRENT"]) != set(records_by_sheet["02_SKU_ES_CURRENT"]):
            raise DictionaryApplyError("MASTER_ES_ZH_SKU_SET_MISMATCH")
        if records_by_sheet["01_SKU_ZH_CURRENT"] != records_by_sheet["02_SKU_ES_CURRENT"]:
            raise DictionaryApplyError("MASTER_ES_ZH_SKU_ORDER_MISMATCH")
    finally:
        wb.close()
    return load_current(path)


def _restore_after_manifest_failure(
    cfg: dict[str, Any], outcome: CommitOutcome, before_hash: str, manifest: dict[str, Any], output_dir: Path,
) -> None:
    try:
        restore_master_from_backup(outcome.backup_path, Path(cfg["paths"]["master"]))
        if _master_hash(cfg) != before_hash:
            raise DictionaryApplyError("MASTER_ROLLBACK_HASH_MISMATCH")
        manifest.update({"committed": False, "commit_state": "ROLLED_BACK", "rollback_status": "SUCCESS"})
        _write_json(output_dir / "apply_manifest.json", manifest)
    except Exception as rollback_exc:
        raise DictionaryApplyError(
            f"MASTER_STATE_UNKNOWN_AFTER_MANIFEST_FAILURE: backup={outcome.backup_path}; rollback_error={rollback_exc}"
        ) from rollback_exc


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        with tmp.open("w", encoding="utf-8", newline="\n") as handle:
            handle.write(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(tmp, path)
    finally:
        tmp.unlink(missing_ok=True)


def _read_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return value if isinstance(value, dict) else {}


def _write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _run_date(cfg: dict[str, Any], run_id: str) -> str:
    matches = list(Path(cfg["paths"]["snapshots"]).glob(f"*/{run_id}/run_report.json"))
    if len(matches) != 1:
        raise DictionaryApplyError(f"FORMAL_RUN_NOT_FOUND: {run_id}")
    report = _read_json(matches[0])
    return str(report.get("run_date") or report.get("observation_date") or matches[0].parent.parent.name)


def _master_hash(cfg: dict[str, Any]) -> str | None:
    path = Path(cfg["paths"]["master"])
    return _master_hash_from_path(path) if path.exists() else None


def _master_hash_from_path(path: Path) -> str | None:
    if not path.exists():
        return None
    return _hash(path)


def _dictionary_baseline_hash(cfg: dict[str, Any]) -> str | None:
    path = Path(cfg["paths"]["dictionary_baseline"]) / "baseline_manifest.json"
    return _hash(path) if path.exists() else None


def _hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()
