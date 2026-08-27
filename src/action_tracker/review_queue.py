"""跨生命周期与字典的统一人工审核队列。

Master 的 06_REVIEW_QUEUE 是既有生命周期审计表，不能被字典流程重写。
本模块把它及字典侧证据只读汇集为运行时队列，并以稳定 review_id 去重。
"""
from __future__ import annotations

import csv
import hashlib
import json
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

import openpyxl

from .dictionary import (
    BRAND_DICTIONARY_HEADERS,
    DictionaryValidationError,
    OVERRIDE_HEADERS,
    PRODUCT_DICTIONARY_HEADERS,
    PRODUCT_OVERRIDE_FIELDS,
    TERM_DICTIONARY_HEADERS,
    load_dictionary_rows,
    write_dictionary_csv,
)


REVIEW_QUEUE_HEADERS = [
    "review_id", "issue_type", "sku", "field", "current_value", "suggested_value",
    "evidence", "reason", "created_at", "status", "source", "updated_at", "resolution",
]
REVIEW_QUEUE_STATUSES = frozenset({"PENDING", "APPROVED", "REJECTED", "RESOLVED"})
REVIEW_QUEUE_TYPES = frozenset({
    # 生命周期已有问题类型
    "SITEMAP_ONLY", "LISTING_ONLY", "UNKNOWN", "SKU_MATCH_CONFLICT", "IDENTITY_CONFLICT",
    "PRICE_ANOMALY", "LABEL_PARSE_FAILED", "DETAIL_FETCH_FAILED", "DETAIL_PARSE_FAILED",
    "TRANSLATION_FAILED", "TRANSLATION_LOW_CONFIDENCE", "CATEGORY_CONFLICT", "IMAGE_MISSING",
    "DATA_INCONSISTENCY",
    # 字典侧统一类型
    "BRAND_CANDIDATE", "TERM_CANDIDATE", "NAME_REVIEW", "SPEC_REVIEW", "CATEGORY_REVIEW",
    "SOURCE_HASH_CHANGED", "MODEL_LOW_CONFIDENCE", "SOURCE_DAMAGED", "SOURCE_POLLUTED",
    "DICTIONARY_CONFLICT",
})


class ReviewQueueError(RuntimeError):
    """审核队列结构、决策或写回目标不安全。"""


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _now() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _stable_review_id(row: Mapping[str, object]) -> str:
    payload = "\x1f".join(_text(row.get(key)) for key in (
        "issue_type", "sku", "field", "current_value", "suggested_value", "evidence",
    ))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]


def _queue_path(cfg: dict[str, Any]) -> Path:
    return Path(cfg["paths"]["review_queue"]) / "review_queue.csv"


def _normalize_row(raw: Mapping[str, object]) -> dict[str, str]:
    row = {header: _text(raw.get(header)) for header in REVIEW_QUEUE_HEADERS}
    if row["issue_type"] not in REVIEW_QUEUE_TYPES:
        raise ReviewQueueError(f"INVALID_ISSUE_TYPE: {row['issue_type']}")
    if row["status"] not in REVIEW_QUEUE_STATUSES:
        raise ReviewQueueError(f"INVALID_REVIEW_STATUS: {row['status']}")
    if not row["review_id"]:
        row["review_id"] = _stable_review_id(row)
    if not row["created_at"]:
        row["created_at"] = _now()
    if not row["updated_at"]:
        row["updated_at"] = row["created_at"]
    return row


def load_queue(cfg: dict[str, Any]) -> dict[str, dict[str, str]]:
    path = _queue_path(cfg)
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None or set(REVIEW_QUEUE_HEADERS) - set(reader.fieldnames):
            raise ReviewQueueError("REVIEW_QUEUE_SCHEMA_MISMATCH")
        result: dict[str, dict[str, str]] = {}
        for raw in reader:
            row = _normalize_row(raw)
            if row["review_id"] in result:
                raise ReviewQueueError(f"DUPLICATE_REVIEW_ID: {row['review_id']}")
            result[row["review_id"]] = row
        return result


def _write_queue(cfg: dict[str, Any], rows: Iterable[Mapping[str, object]]) -> bool:
    path = _queue_path(cfg)
    normalized = [_normalize_row(row) for row in rows]
    ids = [row["review_id"] for row in normalized]
    if len(ids) != len(set(ids)):
        raise ReviewQueueError("DUPLICATE_REVIEW_ID")
    path.parent.mkdir(parents=True, exist_ok=True)
    staged = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp")
    with staged.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=REVIEW_QUEUE_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(normalized)
        fh.flush()
        os.fsync(fh.fileno())
    if path.exists() and path.read_bytes() == staged.read_bytes():
        staged.unlink()
        return False
    os.replace(staged, path)
    return True


def _draft(*, issue_type: str, sku: str = "", field: str = "", current_value: str = "",
           suggested_value: str = "", evidence: str = "", reason: str = "", source: str = "") -> dict[str, str]:
    return _normalize_row({
        "review_id": "", "issue_type": issue_type, "sku": sku, "field": field,
        "current_value": current_value, "suggested_value": suggested_value, "evidence": evidence,
        "reason": reason, "created_at": _now(), "status": "PENDING", "source": source,
        "updated_at": _now(), "resolution": "",
    })


def _master_drafts(master: Path) -> list[dict[str, str]]:
    if not master.exists():
        return []
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    try:
        if "06_REVIEW_QUEUE" not in wb.sheetnames:
            return []
        ws = wb["06_REVIEW_QUEUE"]
        headers = [_text(cell.value) for cell in ws[1]]
        required = ["日期", "SKU", "问题类型", "证据", "候选值", "置信度", "建议动作", "人工备注"]
        if headers != required:
            raise ReviewQueueError("MASTER_REVIEW_QUEUE_SCHEMA_MISMATCH")
        out = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            issue_type = _text(row["问题类型"])
            if not issue_type:
                continue
            out.append(_draft(
                issue_type=issue_type, sku=_text(row["SKU"]), evidence=_text(row["证据"]),
                suggested_value=_text(row["候选值"]), reason=_text(row["人工备注"]),
                source="MASTER_06_REVIEW_QUEUE",
            ))
        return out
    finally:
        wb.close()


def _dictionary_drafts(cfg: dict[str, Any], run_id: str | None) -> list[dict[str, str]]:
    dictionary = Path(cfg["paths"]["dictionary"])
    products = {row["sku"]: row for row in load_dictionary_rows(
        dictionary / "product_dictionary.csv", headers=PRODUCT_DICTIONARY_HEADERS, key_fields=("sku",),
    )}
    formal_terms = load_dictionary_rows(
        dictionary / "term_dictionary.csv", headers=TERM_DICTIONARY_HEADERS, key_fields=("term_es", "term_type"),
    )
    known_term_words = {_text(row["term_es"]).casefold() for row in formal_terms}
    out: list[dict[str, str]] = []
    # 西语源字段损坏/污染均是字典层问题，不依赖 run。
    damage = dictionary / "source_damage_report.csv"
    if damage.exists():
        with damage.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                status = _text(row.get("status"))
                if status not in {"SOURCE_DAMAGED", "SOURCE_POLLUTED"}:
                    continue
                out.append(_draft(
                    issue_type=status, sku=_text(row.get("sku")), field=_text(row.get("damaged_fields")),
                    evidence=_text(row.get("notes")), reason="西语官网事实不可用，禁止反向翻译填充。",
                    source="SOURCE_DAMAGE_REPORT",
                ))
    # 旧品牌候选：已确认的行不再入队，待核验项保留候选品牌证据。
    brands = dictionary / "brand_review_queue.csv"
    if brands.exists():
        with brands.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                if _text(row.get("review_status")) in {"已规则审核", "HUMAN_REVIEWED"}:
                    continue
                out.append(_draft(
                    issue_type="BRAND_CANDIDATE", sku=_text(row.get("sku")), field="brand_id",
                    current_value=_text(row.get("current_name_zh")), suggested_value=_text(row.get("canonical_entities")),
                    evidence=_text(row.get("latin_candidates")), reason=_text(row.get("notes")),
                    source="LEGACY_BRAND_REVIEW_QUEUE",
                ))
    if not run_id:
        return out
    enrichment = dictionary / "enrichment" / run_id
    review_candidates = enrichment / "review_candidates.csv"
    if review_candidates.exists():
        with review_candidates.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                sku, reason = _text(row.get("sku")), _text(row.get("reason"))
                product = products.get(sku, {})
                if reason == "NAME_OR_SPEC_REVIEW":
                    if _text(product.get("translation_status")) != "NEEDS_REVIEW" and _text(product.get("review_status")) != "NEEDS_REVIEW":
                        continue
                    out.append(_draft(
                        issue_type="NAME_REVIEW", sku=sku, field="name_zh_standard",
                        current_value=_text(product.get("name_zh_standard")), evidence=_text(row.get("detail")),
                        reason="同源中文未确认。", source=f"ENRICHMENT:{run_id}",
                    ))
                    out.append(_draft(
                        issue_type="SPEC_REVIEW", sku=sku, field="spec_zh_standard",
                        current_value=_text(product.get("spec_zh_standard")), evidence=_text(row.get("detail")),
                        reason="同源中文规格未确认。", source=f"ENRICHMENT:{run_id}",
                    ))
                elif reason == "BRAND_IDENTIFICATION_PENDING":
                    if _text(product.get("brand_id")):
                        continue
                    out.append(_draft(
                        issue_type="BRAND_CANDIDATE", sku=sku, field="brand_id",
                        current_value="", evidence=_text(row.get("detail")),
                        reason="缺少可确认品牌；程序不会猜测。", source=f"ENRICHMENT:{run_id}",
                    ))
                elif reason == "CATEGORY_REVIEW":
                    if _text(product.get("cat1_zh")):
                        continue
                    out.append(_draft(
                        issue_type="CATEGORY_REVIEW", sku=sku, field="cat1_zh",
                        current_value=_text(product.get("cat1_zh")), evidence=_text(row.get("detail")),
                        reason="未命中固定一级类目。", source=f"ENRICHMENT:{run_id}",
                    ))
    # 已缓存但标记低置信度的模型结果只能产生审核项，不能被 Resolver 当作可信译文。
    model_path = dictionary / "model_translation_overrides.csv"
    if model_path.exists():
        with model_path.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                if _text(row.get("quality_status")).upper() != "NEEDS_REVIEW":
                    continue
                sku = _text(row.get("sku"))
                out.append(_draft(
                    issue_type="MODEL_LOW_CONFIDENCE", sku=sku,
                    field="name_zh_standard/spec_zh_standard",
                    current_value=_text(row.get("name_zh_standard")),
                    suggested_value=_text(row.get("spec_zh_standard")),
                    evidence=_text(row.get("source_hash")),
                    reason="历史模型缓存质量状态为 NEEDS_REVIEW，禁止自动采用。",
                    source="MODEL_TRANSLATION_OVERRIDES",
                ))
    selected = enrichment / "selected_skus.csv"
    if selected.exists():
        with selected.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                if "SOURCE_HASH_CHANGED" not in _text(row.get("reasons")):
                    continue
                sku, observed_hash = _text(row.get("sku")), _text(row.get("source_hash"))
                # 若 enrich 或之后的字典重建已经吸收该事实，旧证据不会重复入队。
                if _text(products.get(sku, {}).get("source_hash")) == observed_hash:
                    continue
                out.append(_draft(
                    issue_type="SOURCE_HASH_CHANGED", sku=sku, field="source_hash",
                    current_value=_text(products.get(sku, {}).get("source_hash")), suggested_value=observed_hash,
                    evidence="snapshot 官网事实哈希已变化。", reason="中文派生值需重新核对。",
                    source=f"ENRICHMENT:{run_id}",
                ))
    term_candidates = dictionary / "term_candidates" / run_id / "term_candidates.csv"
    if term_candidates.exists():
        with term_candidates.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                term = _text(row.get("term_es"))
                if not term or term.casefold() in known_term_words or _text(row.get("review_status")) != "PENDING":
                    continue
                term_type = _text(row.get("term_type")) or _text(row.get("term_type_suggested")) or "general"
                # 兼容旧版候选包；正式候选使用 canonical 字段名。
                category = _text(row.get("cat1_distribution")) or _text(row.get("cat1_zh_distribution"))
                contexts = _text(row.get("sample_contexts")) or _text(row.get("examples_es"))
                zh_contexts = _text(row.get("existing_zh_contexts"))
                out.append(_draft(
                    issue_type="TERM_CANDIDATE", field=term_type,
                    current_value=term, evidence=(
                        f"覆盖SKU={_text(row.get('sku_count'))}；出现={_text(row.get('occurrence_count'))}；"
                        f"类目={category}；上下文={contexts}；"
                        f"现有中文上下文={zh_contexts}"
                    ), reason="候选提取仅供人工确认；不自动写入术语字典。",
                    source=f"TERM_CANDIDATES:{run_id}",
                ))
    return out


def build_review_queue(cfg: dict[str, Any], *, run_id: str | None = None) -> dict[str, Any]:
    """汇集当前有效问题；同一稳定问题只保留一个 review_id。"""
    existing = load_queue(cfg)
    drafts = _master_drafts(Path(cfg["paths"]["master"])) + _dictionary_drafts(cfg, run_id)
    merged = dict(existing)
    active_ids = {draft["review_id"] for draft in drafts}
    # 字典来源是当前状态的投影；在人工决定后完成字典重建，若该投影已不再
    # 产生同一问题，就自动结束待办，但保留整行作为可审计历史。Master 的
    # 生命周期审计行不做这种推断，仍由其原有生命周期规则决定是否消失。
    resolved = 0
    for row in merged.values():
        if (
            row["review_id"] not in active_ids
            and row["source"].startswith(("ENRICHMENT:", "SOURCE_DAMAGE_REPORT", "LEGACY_BRAND_REVIEW_QUEUE", "TERM_CANDIDATES:"))
            and row["status"] in {"PENDING", "APPROVED"}
        ):
            row["status"] = "RESOLVED"
            row["resolution"] = row["resolution"] or "源问题已不再出现"
            row["updated_at"] = _now()
            resolved += 1
    created = 0
    for draft in drafts:
        review_id = draft["review_id"]
        if review_id in merged:
            # 绝不覆盖人工决策、created_at 或 resolution。
            continue
        merged[review_id] = draft
        created += 1
    changed = _write_queue(cfg, [merged[key] for key in sorted(merged)])
    status_counts = {status: sum(row["status"] == status for row in merged.values()) for status in sorted(REVIEW_QUEUE_STATUSES)}
    return {
        "queue": str(_queue_path(cfg)), "total": len(merged), "new": created,
        "resolved": resolved, "status_counts": status_counts, "changed": changed,
        "run_id": run_id or "", "sources_read": len(drafts),
    }


def _upsert_product_override(cfg: dict[str, Any], *, sku: str, field: str, value: str, reason: str) -> None:
    if field not in PRODUCT_OVERRIDE_FIELDS:
        raise ReviewQueueError(f"UNSUPPORTED_PRODUCT_OVERRIDE_FIELD: {field}")
    product_path = Path(cfg["paths"]["dictionary"]) / "product_dictionary.csv"
    product_rows = load_dictionary_rows(product_path, headers=PRODUCT_DICTIONARY_HEADERS, key_fields=("sku",))
    if sku not in {row["sku"] for row in product_rows}:
        raise ReviewQueueError(f"PRODUCT_NOT_IN_DICTIONARY: {sku}")
    path = Path(cfg["paths"]["dictionary"]) / "manual_overrides.csv"
    rows = load_dictionary_rows(path, headers=OVERRIDE_HEADERS, key_fields=("scope", "key", "field"))
    index = {(row["scope"], row["key"], row["field"]): row for row in rows}
    index[("product", sku, field)] = {
        "scope": "product", "key": sku, "field": field, "value": value,
        "reason": reason, "source": "UNIFIED_REVIEW_QUEUE", "locked": "0", "updated_at": _now(),
    }
    write_dictionary_csv(path, [index[key] for key in sorted(index)], OVERRIDE_HEADERS, key_fields=("scope", "key", "field"))


def _upsert_brand(cfg: dict[str, Any], brand: str) -> None:
    path = Path(cfg["paths"]["dictionary"]) / "brand_dictionary.csv"
    rows = load_dictionary_rows(path, headers=BRAND_DICTIONARY_HEADERS, key_fields=("brand_id",))
    if not any(_text(row.get("brand_id")).casefold() == brand.casefold() for row in rows):
        rows.append({
            "brand_id": brand, "canonical_name": brand, "aliases_es": brand, "keep_original": "1",
            "is_action_brand": "0", "confidence": "HUMAN_REVIEWED", "review_status": "HUMAN_REVIEWED",
            "notes": "统一审核队列人工确认。",
        })
        write_dictionary_csv(path, rows, BRAND_DICTIONARY_HEADERS, key_fields=("brand_id",))


def _upsert_term(cfg: dict[str, Any], *, term_es: str, term_type: str, term_zh: str) -> None:
    if not term_es or not term_type or not term_zh:
        raise ReviewQueueError("TERM_DECISION_REQUIRES_TERM_ES_TERM_TYPE_AND_VALUE")
    path = Path(cfg["paths"]["dictionary"]) / "term_dictionary.csv"
    rows = load_dictionary_rows(path, headers=TERM_DICTIONARY_HEADERS, key_fields=("term_es", "term_type"))
    index = {(row["term_es"], row["term_type"]): row for row in rows}
    index[(term_es, term_type)] = {
        "term_es": term_es, "term_zh": term_zh, "term_type": term_type, "forbidden_zh": "",
        "keep_original": "0", "review_status": "HUMAN_REVIEWED", "notes": "统一审核队列人工确认。",
    }
    write_dictionary_csv(path, [index[key] for key in sorted(index)], TERM_DICTIONARY_HEADERS, key_fields=("term_es", "term_type"))


def decide_review(
    cfg: dict[str, Any], *, review_id: str, decision: str, value: str = "", term_type: str = "",
) -> dict[str, Any]:
    """保存人工决定；批准的字典项写到对应字典，而非写回官网或 Master。"""
    decision = decision.upper()
    if decision not in REVIEW_QUEUE_STATUSES:
        raise ReviewQueueError(f"INVALID_DECISION: {decision}")
    queue = load_queue(cfg)
    if review_id not in queue:
        raise ReviewQueueError(f"REVIEW_NOT_FOUND: {review_id}")
    row = queue[review_id]
    route = "status_only"
    if decision == "APPROVED":
        if row["issue_type"] == "BRAND_CANDIDATE":
            if not value or not row["sku"]:
                raise ReviewQueueError("BRAND_APPROVAL_REQUIRES_SKU_AND_VALUE")
            _upsert_brand(cfg, value)
            _upsert_product_override(cfg, sku=row["sku"], field="brand_id", value=value, reason=f"review_id={review_id}")
            route = "brand_dictionary+manual_overrides"
        elif row["issue_type"] in {"NAME_REVIEW", "CATEGORY_REVIEW"}:
            if not value or not row["sku"]:
                raise ReviewQueueError("PRODUCT_APPROVAL_REQUIRES_SKU_AND_VALUE")
            _upsert_product_override(cfg, sku=row["sku"], field=row["field"], value=value, reason=f"review_id={review_id}")
            route = "manual_overrides"
        elif row["issue_type"] == "TERM_CANDIDATE":
            _upsert_term(cfg, term_es=row["current_value"], term_type=term_type or row["field"], term_zh=value)
            route = "term_dictionary"
        else:
            raise ReviewQueueError(f"APPROVAL_ROUTE_NOT_DEFINED: {row['issue_type']}")
    row["status"] = decision
    row["resolution"] = value or row["resolution"]
    row["updated_at"] = _now()
    _write_queue(cfg, [queue[key] for key in sorted(queue)])
    return {"review_id": review_id, "status": decision, "route": route, "rebuild_required": route != "status_only"}
