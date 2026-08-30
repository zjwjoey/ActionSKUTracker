from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from action_tracker.exporting.history import PRESENCE_UNKNOWN, load_presence_history
from action_tracker.exporting.history_export import HistoryExportError, export_history

from test_exporting import _cfg


def _write_seed(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ACTION商品上下架明细"
    sheet.append(["序号", "编号", "中文品名", "图片链接", "商品链接", "26.08.24", "26.08.25"])
    sheet.append([1, "9002", "历史商品2", "https://images.example/9002.jpg", "https://www.action.com/es-es/p/9002/", 0, 1])
    sheet.append([2, "9001", "历史商品1", "", "https://www.action.com/es-es/p/9001/", 1, 0])
    workbook.save(path)
    workbook.close()


def _history_cfg(tmp_path: Path, seed: Path, *, partial: bool = False, source: Path | None = None) -> dict:
    cfg = _cfg(tmp_path)
    config = tmp_path / "history_sources.yaml"
    lines = [
        "seed:", f"  path: '{seed.as_posix()}'", "  sheet: ACTION商品上下架明细", "  sku_header: 编号",
        "  presence_capability: true", "  absence_capability: true", "  observation_complete: true", "  evidence_level: A",
        "  fields:", "    name_zh: 中文品名", "    image_url: 图片链接", "    product_url: 商品链接",
    ]
    if partial:
        lines += [
            "sources:", "  - date: '2026-08-26'", f"    path: '{(source or seed).as_posix()}'",
            "    sheet: ACTION商品上下架明细", "    sku_header: 编号", "    presence_capability: true",
            "    absence_capability: false", "    observation_complete: false", "    evidence_level: B", "    fields: {}",
        ]
    else:
        lines.append("sources: []")
    config.write_text("\n".join(lines) + "\n", encoding="utf-8")
    cfg["history_sources_path"] = config
    return cfg


def test_export_history_writes_union_and_manifest(tmp_path):
    seed = tmp_path / "seed.xlsx"
    _write_seed(seed)
    cfg = _history_cfg(tmp_path, seed)
    result = export_history(cfg, export_date="2026-08-30")

    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        assert workbook.sheetnames == ["商品上下架明细", "历史来源审计"]
        sheet = workbook["商品上下架明细"]
        headers = [cell.value for cell in sheet[1]]
        assert headers[:5] == ["序号", "编号", "中文品名", "品牌", "一级类目（中文）"]
        assert headers[-2:] == ["26.08.24", "26.08.25"]
        date_start = headers.index("26.08.24") + 1
        rows = {str(sheet.cell(row=row, column=2).value): [sheet.cell(row=row, column=col).value for col in range(date_start, date_start + 2)] for row in range(2, sheet.max_row + 1)}
        assert rows == {"9001": [1, 0], "9002": [0, 1]}
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref == "A1:O3"
        assert workbook["历史来源审计"].max_row == 3
    finally:
        workbook.close()

    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["template_id"] == "action_history_presence"
    assert manifest["history_union_sku_count"] == 2
    assert manifest["history_dates"] == ["2026-08-24", "2026-08-25"]
    assert manifest["unknown_presence_count"] == 0
    assert len(manifest["seed_sha256"]) == 64


def test_partial_source_absence_is_unknown(tmp_path):
    seed = tmp_path / "seed.xlsx"
    _write_seed(seed)
    partial = tmp_path / "partial.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ACTION商品上下架明细"
    sheet.append(["序号", "编号", "中文品名", "图片链接", "商品链接"])
    sheet.append([1, "9002", "历史商品2", "", ""])
    workbook.save(partial)
    workbook.close()
    history = load_presence_history(_history_cfg(tmp_path, seed, partial=True, source=partial))
    assert history.presence_by_sku["9001"]["2026-08-26"] == PRESENCE_UNKNOWN


def test_export_history_rejects_invalid_presence(tmp_path):
    seed = tmp_path / "seed.xlsx"
    _write_seed(seed)
    workbook = openpyxl.load_workbook(seed)
    workbook["ACTION商品上下架明细"]["F2"] = 2
    workbook.save(seed)
    workbook.close()
    with pytest.raises(HistoryExportError, match="HISTORY_SEED_BAD_PRESENCE"):
        export_history(_history_cfg(tmp_path, seed), export_date="2026-08-30")


def test_missing_capability_config_is_rejected(tmp_path):
    seed = tmp_path / "seed.xlsx"
    _write_seed(seed)
    cfg = _history_cfg(tmp_path, seed)
    cfg["history_sources_path"].write_text(
        cfg["history_sources_path"].read_text(encoding="utf-8").replace("  evidence_level: A\n", ""),
        encoding="utf-8",
    )
    with pytest.raises(HistoryExportError, match="HISTORY_CAPABILITY_CONFIG_MISSING"):
        load_presence_history(cfg)
