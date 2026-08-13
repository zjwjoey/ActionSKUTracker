import datetime as dt

import openpyxl
import pytest

from action_tracker.excel import writer


def _base_workbook(path):
    wb = openpyxl.Workbook()
    zh = wb.active
    zh.title = "01_SKU_ZH_CURRENT"
    zh.append(["Canonical_ID", "SKU"])
    zh.append(["ACT0001001", "1001"])
    es = wb.create_sheet("02_SKU_ES_CURRENT")
    es.append(["Canonical_ID", "SKU"])
    es.append(["ACT0001001", "1001"])
    wb.create_sheet("03_PRICE_HISTORY").append(writer.PRICE_HISTORY_HEADERS)
    wb.create_sheet("04_EVENT_HISTORY").append(writer.EVENT_HISTORY_HEADERS)
    wb.create_sheet("05_RUN_LOG").append(writer.RUN_LOG_HEADERS)
    wb.create_sheet("06_REVIEW_QUEUE")
    archive = wb.create_sheet("07_APRIL_ARCHIVE")
    archive.append(writer.APRIL_ARCHIVE_HEADERS)
    archive.append([
        "vivienda_813161_0", None, dt.date(2026, 4, 5), "家居生活", "Vivienda",
        "花盆", "Maceta", 8.48, "40 厘米", "40 cm", "8.48/件", "8.48/ud.",
        None, None, "名称/规格/价格", "UNMATCHED_ARCHIVE_ID", "es.xlsx", "zh.xlsx",
        "Sheet1", "图片排除",
    ])
    return wb


def _add_long_term_sheets(wb, *, catalog_current="CURRENT", pending_sku=None):
    catalog = wb.create_sheet("08_LONG_TERM_MASTER")
    for _ in range(6):
        catalog.append([])
    catalog.append(writer.LONG_TERM_MASTER_HEADERS)
    catalog.append([
        "ACT0001001", "1001", None, "OFFICIAL_SKU", "OFFICIAL_IDENTITY", 1.0,
        catalog_current, "商品", "Producto", "家居", "Vivienda", "", "", 1.0,
        1.0, 1.0, dt.date(2026, 1, 9), dt.date(2026, 8, 13), 0, None, 2,
        "CURRENT|HISTORY", None, None,
    ])
    catalog.append([
        "APRIL:vivienda_813161_0", pending_sku, "vivienda_813161_0",
        "APRIL_ARCHIVE_PENDING", "UNMATCHED_ARCHIVE_ID", 0.0,
        "ARCHIVE_PENDING_MATCH", "花盆", "Maceta", "家居", "Vivienda", "40 厘米", "40 cm",
        None, 8.48, 8.48, dt.date(2026, 4, 5), dt.date(2026, 4, 5), 1,
        "vivienda_813161_0", 1, "07_APRIL_ARCHIVE", None, None,
    ])
    audit = wb.create_sheet("09_APRIL_MATCH_AUDIT")
    audit.append(writer.APRIL_MATCH_AUDIT_HEADERS)
    audit.append([
        "vivienda_813161_0", None, "UNMATCHED_ARCHIVE_ID", "NO_EXACT_REFERENCE", 0.0,
        None, None, dt.date(2026, 4, 5), "家居", "Vivienda", "花盆", "Maceta",
        8.48, "40 厘米", "40 cm", "8.48/件", "8.48/ud.", None, None, 2, "es.xlsx", "zh.xlsx",
    ])
    schema = wb.create_sheet("10_SOURCE_SCHEMA")
    schema.append(writer.SOURCE_SCHEMA_HEADERS)
    schema.append([dt.date(2026, 4, 5), "es.xlsx", "Sheet1", 1, 8, "a|b", "archive", "valid", None])


def test_validate_long_term_master(tmp_path):
    path = tmp_path / "master.xlsx"
    wb = _base_workbook(path)
    _add_long_term_sheets(wb)
    wb.save(path)
    wb.close()
    writer._validate(path)


def test_rejects_partial_long_term_sheets(tmp_path):
    path = tmp_path / "master.xlsx"
    wb = _base_workbook(path)
    ws = wb.create_sheet("08_LONG_TERM_MASTER")
    for _ in range(6):
        ws.append([])
    ws.append(writer.LONG_TERM_MASTER_HEADERS)
    wb.save(path)
    wb.close()
    with pytest.raises(RuntimeError, match="incomplete"):
        writer._validate(path)


def test_pending_april_entity_cannot_claim_official_sku(tmp_path):
    path = tmp_path / "master.xlsx"
    wb = _base_workbook(path)
    _add_long_term_sheets(wb, pending_sku="813161")
    wb.save(path)
    wb.close()
    with pytest.raises(RuntimeError, match="must not have official SKU"):
        writer._validate(path)


def test_catalog_current_count_must_match_current_sheet(tmp_path):
    path = tmp_path / "master.xlsx"
    wb = _base_workbook(path)
    _add_long_term_sheets(wb, catalog_current="HISTORICAL")
    wb.save(path)
    wb.close()
    with pytest.raises(RuntimeError, match="CURRENT count mismatch"):
        writer._validate(path)


def test_refresh_long_term_catalog_keeps_history_and_adds_new_sku(tmp_path):
    path = tmp_path / "master.xlsx"
    wb = _base_workbook(path)
    _add_long_term_sheets(wb)
    es = wb["02_SKU_ES_CURRENT"]
    es.append(["ACT0001002", "1002"])
    writer._refresh_long_term_catalog(wb)
    catalog = wb["08_LONG_TERM_MASTER"]
    rows = list(catalog.iter_rows(min_row=8, values_only=True))
    official = {str(r[1]): r for r in rows if r[3] == "OFFICIAL_SKU"}
    assert set(official) == {"1001", "1002"}
    assert official["1001"][6] == "CURRENT"
    assert official["1002"][6] == "CURRENT"
    assert catalog["B4"].value == 2
    assert catalog["D4"].value == 2
    assert catalog["H4"].value == 1
    wb.close()
