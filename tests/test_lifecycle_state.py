"""P0 生命周期状态持久化测试（known_skus / offline_skus 推进规则 + 提交门禁 + 原子写）。

对应验收 15 项 + 跨日完整生命周期验证（规范 §56"更新 known_skus"）。
"""
import datetime as dt
import os
from pathlib import Path

import openpyxl
import pytest

from action_tracker import state as st
from action_tracker.excel import writer
from action_tracker.monitor.sku_monitor import SkuStatus
from action_tracker.orchestrator import daily as daily_mod
from action_tracker.services.lifecycle import classify


def _sku(sku: str = "1001", status: str = "ACTIVE", missing_count: int = 0,
         cid: str = "ACT0001001", event: str | None = None,
         previous_status: str = "") -> SkuStatus:
    return SkuStatus(sku=sku, canonical_id=cid, status=status,
                     source_flag="BOTH", sitemap_present=True, listing_present=True,
                     was_yesterday=True, ever_seen=True, first_seen="2026-08-11",
                     previous_status=previous_status,
                     missing_count=missing_count, event=event, light=None)


def _known(sku: str = "1001", *, first_seen="2026-08-11", last_seen="2026-08-11",
           last_status="ACTIVE", missing_count="0", last_missing="", offline="",
           obs="2026-08-11", ever_offline="false") -> dict:
    return {sku: {"canonical_id": "ACT0001001", "official_sku": sku,
                  "first_seen_date": first_seen, "last_seen_date": last_seen,
                  "last_status": last_status, "missing_count": missing_count,
                  "last_missing_date": last_missing, "offline_date": offline,
                  "last_state_observation_date": obs, "ever_offline": ever_offline,
                  "last_run_id": "", "updated_at": ""}}


def _trans(known: dict, statuses: dict, run_date="2026-08-11", run_id="R1", offline_runs=3) -> dict:
    return st.apply_state_transition(known, statuses, run_date, run_id, offline_runs)


def _day(known: dict, *, present: bool, ever_seen: bool,
         run_date: str, run_id: str, offline_runs: int = 3) -> dict:
    """用真实 classify + apply_state_transition 模拟一天。

    上一有效状态与 missing_count 都从 known_skus（生命周期状态源）读取。
    """
    rec = known.get("1001", {})
    mc = int(rec.get("missing_count") or 0)
    prev_status = rec.get("last_status") or ""
    cls = classify(today_present=present, previous_status=prev_status, ever_seen=ever_seen,
                   missing_count=mc, offline_runs=offline_runs)
    s = _sku("1001", status=cls.status, missing_count=cls.missing_count, event=cls.event,
             previous_status=prev_status)
    return st.apply_state_transition(known, {"1001": s}, run_date, run_id, offline_runs)


# ---- 1/2. NEW 首次运行写入 known_skus，first_seen = today ----
def test_t1_2_new_creates_record_with_first_seen_today():
    out = _trans({}, {"1001": _sku(status="NEW")}, run_date="2026-08-11")
    rec = out["known"]["1001"]
    assert rec["first_seen_date"] == "2026-08-11"
    assert rec["last_seen_date"] == "2026-08-11"
    assert rec["last_status"] == "ACTIVE"
    assert rec["missing_count"] == "0"
    assert rec["official_sku"] == "1001"


# ---- 3. 第二天 NEW 商品仍存在：不能重新 NEW（first_seen 保留，状态推进） ----
def test_t3_second_day_act1ive_not_new_again():
    d1 = _trans({}, {"1001": _sku(status="NEW")}, run_date="2026-08-11")
    d2 = _trans(d1["known"], {"1001": _sku(status="ACTIVE")}, run_date="2026-08-12")
    rec = d2["known"]["1001"]
    assert rec["first_seen_date"] == "2026-08-11"   # 不覆盖
    assert rec["last_seen_date"] == "2026-08-12"
    assert rec["last_status"] == "ACTIVE"
    assert rec["missing_count"] == "0"


# ---- 4. 第一次缺失：missing_count = 1，且 last_seen 不更新 ----
def test_t4_first_missing_count_1_last_seen_kept():
    k = _known(last_seen="2026-08-11", obs="2026-08-11")
    out = _trans(k, {"1001": _sku(status="MISSING_FIRST", missing_count=1)},
                 run_date="2026-08-12")
    rec = out["known"]["1001"]
    assert rec["missing_count"] == "1"
    assert rec["last_status"] == "MISSING"
    assert rec["last_missing_date"] == "2026-08-12"
    assert rec["last_seen_date"] == "2026-08-11"   # 缺失当天不更新


# ---- 5. 第二天继续缺失：missing_count = 2 ----
def test_t5_second_missing_count_2():
    k = _known(last_status="MISSING", missing_count="1", last_seen="2026-08-11", obs="2026-08-12")
    out = _trans(k, {"1001": _sku(status="MISSING_CONTINUED", missing_count=2)},
                 run_date="2026-08-13")
    assert out["known"]["1001"]["missing_count"] == "2"
    assert out["known"]["1001"]["last_status"] == "MISSING"


# ---- 6/7. 达到 offline_runs：OFFLINE，并进入 offline_skus ----
def test_t6_7_reach_offline_and_appears_in_offline_skus():
    k = _known(last_status="MISSING", missing_count="2", last_seen="2026-08-12", obs="2026-08-13")
    out = _trans(k, {"1001": _sku(status="OFFLINE", missing_count=3)},
                 run_date="2026-08-14")
    rec = out["known"]["1001"]
    assert rec["last_status"] == "OFFLINE"
    assert rec["offline_date"] == "2026-08-14"
    assert rec["first_seen_date"] == "2026-08-11"   # 历史保留
    assert rec["last_seen_date"] == "2026-08-12"    # 历史保留
    assert out["offline"] == [r for r in st.derive_offline_skus(out["known"]) if r["official_sku"] == "1001"]
    assert rec["official_sku"] in {r["official_sku"] for r in out["offline"]}


# ---- 8. OFFLINE 后重新出现：ACTIVE，missing_count 清零，并产生 REAPPEARED ----
def test_t8_reappeared_resets_to_active_and_emits_event():
    k = _known(last_status="OFFLINE", missing_count="3", last_seen="2026-08-12",
               offline="2026-08-14", obs="2026-08-15", ever_offline="true")
    out = _trans(k, {"1001": _sku(status="REAPPEARED", missing_count=0, event="REAPPEARED")},
                 run_date="2026-08-16")
    rec = out["known"]["1001"]
    assert rec["last_status"] == "ACTIVE"
    assert rec["missing_count"] == "0"
    assert rec["last_missing_date"] == ""
    assert rec["last_seen_date"] == "2026-08-16"
    assert rec["first_seen_date"] == "2026-08-11"
    assert rec["ever_offline"] == "true"   # 历史标记保留
    # REAPPEARED 事件进 04
    events = daily_mod._build_lifecycle_events(
        {"1001": _sku(status="REAPPEARED", event="REAPPEARED")}, "2026-08-16", "R1")
    assert len(events) == 1
    ev = events[0]
    assert ev["事件类型"] == "REAPPEARED" and ev["旧值"] == "OFFLINE" and ev["新值"] == "ACTIVE"
    assert ev["SKU"] == "1001" and ev["来源文件"] == "Action_Master.xlsx" and ev["备注"] == "R1"


# ---- 9. 同一天重复正式运行：missing_count 不重复 +1 ----
def test_t9_same_day_rerun_does_not_advance_missing():
    k = _known(last_status="MISSING", missing_count="1", last_seen="2026-08-12", obs="2026-08-13")
    out = _trans(k, {"1001": _sku(status="MISSING_CONTINUED", missing_count=2)},
                 run_date="2026-08-13")  # 同日第二次运行
    assert out["known"]["1001"]["missing_count"] == "1"


# ---- 10/11. dry-run / QA FAIL：known_skus hash 不变（提交门禁 + 转换纯函数） ----
def test_t10_11_gate_blocks_dry_run_and_qa_fail(tmp_path):
    assert daily_mod._should_commit(dry_run=True, qa_passed=True) is False
    assert daily_mod._should_commit(dry_run=False, qa_passed=False) is False
    assert daily_mod._should_commit(dry_run=False, qa_passed=True) is True
    assert daily_mod._should_commit(dry_run=False, qa_passed=True, access_state="COOLDOWN") is False
    assert daily_mod._should_commit(dry_run=False, qa_passed=True, access_state="PROBE") is False
    assert daily_mod._should_commit(dry_run=False, qa_passed=True, access_state="BLOCKED") is False


def test_t10_11_transition_is_pure_no_disk_write(tmp_path):
    state_dir = tmp_path / "state"
    state_dir.mkdir()
    k = _known()
    st.save_known_skus(state_dir, k)
    before = (state_dir / "known_skus.csv").read_bytes()
    _trans(k, {"1001": _sku(status="MISSING_FIRST", missing_count=1)}, run_date="2026-08-12")
    assert (state_dir / "known_skus.csv").read_bytes() == before  # 纯计算不写盘


# ---- 12a. 历史缺口回填：基线在 known 之外的 ACTIVE 首次纳入时补 first_seen，已有则不覆盖 ----
def test_t12a_active_backfills_first_seen_when_empty():
    out = _trans({}, {"1001": _sku(status="ACTIVE")}, run_date="2026-08-11")
    assert out["known"]["1001"]["first_seen_date"] == "2026-08-11"  # 回填
    # 已存在的 first_seen 绝不因 ACTIVE 覆盖
    k = _known(first_seen="2026-01-09")
    out = _trans(k, {"1001": _sku(status="ACTIVE")}, run_date="2026-08-12")
    assert out["known"]["1001"]["first_seen_date"] == "2026-01-09"


# ---- 12. first_seen_date 永不被后续运行覆盖 ----
def test_t12_first_seen_never_overwritten():
    k = _known(first_seen="2026-01-09")
    for run in ["2026-08-11", "2026-08-12", "2026-08-13"]:
        out = _trans(k, {"1001": _sku(status="ACTIVE")}, run_date=run)
        assert out["known"]["1001"]["first_seen_date"] == "2026-01-09"
    out = _trans(k, {"1001": _sku(status="MISSING_FIRST", missing_count=1)}, run_date="2026-08-14")
    assert out["known"]["1001"]["first_seen_date"] == "2026-01-09"


# ---- 13. last_seen 缺失当天不能更新（test_t4 已覆盖，这里补 OFFLINE 分支） ----
def test_t13_last_seen_not_updated_on_missing_or_offline():
    k = _known(last_seen="2026-08-11", obs="2026-08-13")
    m = _trans(k, {"1001": _sku(status="MISSING_FIRST", missing_count=1)}, run_date="2026-08-14")
    assert m["known"]["1001"]["last_seen_date"] == "2026-08-11"
    k2 = _known(last_status="MISSING", missing_count="2", last_seen="2026-08-11", obs="2026-08-14")
    o = _trans(k2, {"1001": _sku(status="OFFLINE", missing_count=3)}, run_date="2026-08-15")
    assert o["known"]["1001"]["last_seen_date"] == "2026-08-11"


# ---- 14. 状态 CSV 写入失败：原文件不损坏（原子写） ----
def test_t14_atomic_write_keeps_original_on_failure(tmp_path, monkeypatch):
    state_dir = tmp_path / "state"
    k = _known()
    st.save_known_skus(state_dir, k)
    orig = (state_dir / "known_skus.csv").read_bytes()

    def boom(src, dst):
        raise OSError("模拟 os.replace 失败")

    monkeypatch.setattr(os, "replace", boom)
    with pytest.raises(OSError):
        st.save_known_skus(state_dir, k)
    assert (state_dir / "known_skus.csv").read_bytes() == orig


# ---- 15. offline_skus 完全等价于 known_skus[last_status == OFFLINE] ----
def test_t15_offline_skus_derived_only():
    k = _known()
    k["2002"] = {"canonical_id": "ACT0002002", "official_sku": "2002",
                 "first_seen_date": "2026-01-01", "last_seen_date": "2026-08-10",
                 "last_status": "OFFLINE", "missing_count": "3",
                 "last_missing_date": "2026-08-10", "offline_date": "2026-08-10",
                 "last_state_observation_date": "2026-08-10", "ever_offline": "true",
                 "last_run_id": "OLD", "updated_at": "x"}
    derived = st.derive_offline_skus(k)
    assert [r["official_sku"] for r in derived] == ["2002"]
    assert all(r["last_status"] == "OFFLINE" for r in derived)


# ---- 完整跨日生命周期（真实 classify 驱动）----
def test_full_lifecycle_5_days():
    state = {}
    # Day1 首次出现 -> NEW -> known 记 ACTIVE
    state = _day(state, present=True, ever_seen=False, run_date="2026-08-11", run_id="D1")
    assert state["known"]["1001"]["last_status"] == "ACTIVE"
    assert state["known"]["1001"]["first_seen_date"] == "2026-08-11"
    assert state["known"]["1001"]["missing_count"] == "0"
    # Day2 仍在售 -> ACTIVE
    state = _day(state["known"], present=True, ever_seen=True, run_date="2026-08-12", run_id="D2")
    assert state["known"]["1001"]["last_status"] == "ACTIVE"
    assert state["known"]["1001"]["missing_count"] == "0"
    assert state["known"]["1001"]["last_seen_date"] == "2026-08-12"
    # Day3 缺失 -> MISSING_FIRST, missing_count 1
    state = _day(state["known"], present=False, ever_seen=True, run_date="2026-08-13", run_id="D3")
    assert state["known"]["1001"]["last_status"] == "MISSING"
    assert state["known"]["1001"]["missing_count"] == "1"
    assert state["known"]["1001"]["last_seen_date"] == "2026-08-12"  # 缺失当天不更新
    # Day4 继续缺失 -> missing_count 2
    state = _day(state["known"], present=False, ever_seen=True, run_date="2026-08-14", run_id="D4")
    assert state["known"]["1001"]["missing_count"] == "2"
    # Day5 达到阈值 -> OFFLINE, missing_count 3，进 offline_skus
    state = _day(state["known"], present=False, ever_seen=True, run_date="2026-08-15", run_id="D5")
    assert state["known"]["1001"]["last_status"] == "OFFLINE"
    assert state["known"]["1001"]["missing_count"] == "3"
    assert state["known"]["1001"]["offline_date"] == "2026-08-15"
    assert "1001" in {r["official_sku"] for r in state["offline"]}
    # Day6 重新出现 -> REAPPEARED -> ACTIVE, missing_count 清零
    state = _day(state["known"], present=True, ever_seen=True, run_date="2026-08-16", run_id="D6")
    assert state["known"]["1001"]["last_status"] == "ACTIVE"
    assert state["known"]["1001"]["missing_count"] == "0"
    assert "1001" not in {r["official_sku"] for r in state["offline"]}


# ==================== REAPPEARED 新规则（跨日状态推进） ====================

# ---- 场景 1：Day1 ACTIVE → Day2 MISSING_FIRST（仍保留 CURRENT）→ Day3 重现 ----
def test_reappeared_after_missing_first():
    state = {}
    # Day1 首次出现 -> NEW -> ACTIVE
    state = _day(state, present=True, ever_seen=False, run_date="2026-08-11", run_id="D1")
    assert state["known"]["1001"]["last_status"] == "ACTIVE"
    # Day2 缺失 -> MISSING_FIRST（商品仍保留在 CURRENT，known 记为 MISSING）
    state = _day(state["known"], present=False, ever_seen=True, run_date="2026-08-12", run_id="D2")
    assert state["known"]["1001"]["last_status"] == "MISSING"
    assert state["known"]["1001"]["missing_count"] == "1"
    # Day3 重现 -> REAPPEARED -> ACTIVE，missing_count 清零
    state = _day(state["known"], present=True, ever_seen=True, run_date="2026-08-13", run_id="D3")
    assert state["known"]["1001"]["last_status"] == "ACTIVE"
    assert state["known"]["1001"]["missing_count"] == "0"
    assert state["known"]["1001"]["last_seen_date"] == "2026-08-13"
    assert state["known"]["1001"]["first_seen_date"] == "2026-08-11"   # 保持不变


# ---- 场景 2：Day1 ACTIVE → Day2 MISSING_FIRST → Day3 MISSING_CONTINUED → Day4 重现 ----
def test_reappeared_after_missing_continued():
    state = {}
    state = _day(state, present=True, ever_seen=False, run_date="2026-08-11", run_id="D1")
    state = _day(state["known"], present=False, ever_seen=True, run_date="2026-08-12", run_id="D2")
    state = _day(state["known"], present=False, ever_seen=True, run_date="2026-08-13", run_id="D3")
    assert state["known"]["1001"]["last_status"] == "MISSING"
    assert state["known"]["1001"]["missing_count"] == "2"
    state = _day(state["known"], present=True, ever_seen=True, run_date="2026-08-14", run_id="D4")
    assert state["known"]["1001"]["last_status"] == "ACTIVE"
    assert state["known"]["1001"]["missing_count"] == "0"


# ---- 场景 3：OFFLINE 后重现 -> REAPPEARED + ACTIVE ----
def test_reappeared_after_offline():
    k = _known(last_status="OFFLINE", missing_count="3", last_seen="2026-08-12",
               offline="2026-08-14", obs="2026-08-15", ever_offline="true")
    out = _trans(k, {"1001": _sku(status="REAPPEARED", missing_count=0, event="REAPPEARED",
                                  previous_status="OFFLINE")},
                 run_date="2026-08-16")
    assert out["known"]["1001"]["last_status"] == "ACTIVE"
    assert out["known"]["1001"]["missing_count"] == "0"
    assert out["known"]["1001"]["first_seen_date"] == "2026-08-11"   # 保持不变


# ---- 场景 4：连续 ACTIVE -> 不产生 REAPPEARED ----
def test_consecutive_active_no_reappeared():
    c = classify(today_present=True, previous_status="ACTIVE", ever_seen=True)
    assert c.status == "ACTIVE"
    assert c.event is None
    assert c.event != "REAPPEARED"


# ---- 场景 5：首次出现 -> FIRST_SEEN，不得 REAPPEARED ----
def test_first_seen_never_reappeared():
    out = _trans({}, {"1001": _sku(status="NEW", event="FIRST_SEEN")}, run_date="2026-08-11")
    assert out["known"]["1001"]["last_status"] == "ACTIVE"
    assert out["known"]["1001"]["first_seen_date"] == "2026-08-11"
    events = daily_mod._build_lifecycle_events(
        {"1001": _sku(status="NEW", event="FIRST_SEEN")}, "2026-08-11", "R1")
    assert events == []   # FIRST_SEEN 不产生 REAPPEARED 事件


# ---- REAPPEARED 事件旧值反映真实上一状态（MISSING 而非硬编码 OFFLINE）----
def test_reappeared_event_old_value_reflects_previous_state():
    events = daily_mod._build_lifecycle_events(
        {"1001": _sku(status="REAPPEARED", event="REAPPEARED", previous_status="MISSING")},
        "2026-08-16", "R1")
    assert len(events) == 1
    assert events[0]["事件类型"] == "REAPPEARED"
    assert events[0]["旧值"] == "MISSING"
    assert events[0]["新值"] == "ACTIVE"


# ---- 同日重复运行的完整模拟（验收示例：08-11 两次 + 08-12 + 08-13 → OFFLINE）----
def test_same_day_rerun_acceptance_example():
    # 08-11 第一次
    k = _known(last_status="MISSING", missing_count="0", last_seen="2026-08-10", obs="2026-08-10")
    out = _trans(k, {"1001": _sku(status="MISSING_FIRST", missing_count=1)}, run_date="2026-08-11")
    assert out["known"]["1001"]["missing_count"] == "1"
    # 08-11 第二次（同一天）：仍然 = 1
    out = _trans(out["known"], {"1001": _sku(status="MISSING_CONTINUED", missing_count=2)}, run_date="2026-08-11")
    assert out["known"]["1001"]["missing_count"] == "1"
    # 08-12：1 → 2
    out = _trans(out["known"], {"1001": _sku(status="MISSING_CONTINUED", missing_count=2)}, run_date="2026-08-12")
    assert out["known"]["1001"]["missing_count"] == "2"
    # 08-13：2 → 3 → OFFLINE
    out = _trans(out["known"], {"1001": _sku(status="OFFLINE", missing_count=3)}, run_date="2026-08-13")
    assert out["known"]["1001"]["missing_count"] == "3"
    assert out["known"]["1001"]["last_status"] == "OFFLINE"


# ---- 提交接线：_commit_phase 原子写 Master + known_skus + offline_skus ----
def _cfg(tmp_path: Path) -> dict:
    return {
        "paths": {
            "master": tmp_path / "master" / "Action_Master.xlsx",
            "backups": tmp_path / "backups",
            "temp": tmp_path / "temp",
            "state": tmp_path / "state",
        }
    }


def _build_master(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "01_SKU_ZH_CURRENT"
    ws.append(["Canonical_ID", "SKU", "中文品名", "当前售价 (€)", "最后确认存在日期"])
    ws.append(["ACT0001001", "1001", "品", 2.99, "2026-08-10"])
    ws2 = wb.create_sheet("02_SKU_ES_CURRENT")
    ws2.append(["Canonical_ID", "SKU", "西班牙语品名", "当前售价 (€)", "最后确认存在日期"])
    ws2.append(["ACT0001001", "1001", "Nombre", 2.99, "2026-08-10"])
    ws3 = wb.create_sheet("03_PRICE_HISTORY")
    ws3.append(writer.PRICE_HISTORY_HEADERS)
    ws4 = wb.create_sheet("04_EVENT_HISTORY")
    ws4.append(writer.EVENT_HISTORY_HEADERS)
    wb.save(path)


def _run_log_row() -> dict:
    return {h: ("" if h not in ("Run ID", "运行日期", "Git Commit", "运行状态") else "x")
            for h in writer.RUN_LOG_HEADERS}


def test_commit_phase_writes_all_three(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    statuses = {"1001": _sku(status="ACTIVE")}
    known = _known()
    rc = daily_mod._commit_phase(
        cfg, statuses=statuses, known=known, run_date="2026-08-11", run_id="R1",
        offline_runs=3, today_records={"1001": {"sku": "1001", "current_price": 2.99}},
        price_events=[], event_events=[], run_log_row=_run_log_row(), review_rows=[])
    assert rc == "FULL_COMMIT"
    # known_skus 已更新
    assert st.load_known_skus(cfg["paths"]["state"])["1001"]["last_status"] == "ACTIVE"
    assert st.load_known_skus(cfg["paths"]["state"])["1001"]["last_state_observation_date"] == "2026-08-11"
    # offline_skus 已生成（无 OFFLINE 则为空表）
    off_path = cfg["paths"]["state"] / "offline_skus.csv"
    assert off_path.exists()
    # Master 已替换（05_RUN_LOG 存在）
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    assert "05_RUN_LOG" in wb.sheetnames
    wb.close()


def test_commit_phase_offline_skus_contains_offline(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    known = _known(last_status="MISSING", missing_count="2", last_seen="2026-08-12", obs="2026-08-13")
    rc = daily_mod._commit_phase(
        cfg, statuses={"1001": _sku(status="OFFLINE", missing_count=3)}, known=known,
        run_date="2026-08-14", run_id="R1", offline_runs=3,
        today_records={}, price_events=[], event_events=[], run_log_row=_run_log_row(), review_rows=[])
    assert rc == "FULL_COMMIT"
    offline = st.load_offline_skus(cfg["paths"]["state"])
    assert "1001" in offline and offline["1001"]["last_status"] == "OFFLINE"


def test_stage_master_does_not_replace(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    orig = cfg["paths"]["master"].read_bytes()
    tmp = writer.stage_master(cfg, updated_records={}, price_events=[], event_events=[])
    assert cfg["paths"]["master"].read_bytes() == orig  # 未替换
    writer.commit_master(tmp, cfg["paths"]["master"])
    assert cfg["paths"]["master"].read_bytes() != orig  # 替换后
