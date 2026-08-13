import datetime as dt

import openpyxl
import pytest

from action_tracker.excel import writer


def _cfg(tmp_path):
    return {
        "paths": {
            "master": tmp_path / "master" / "Action_Master.xlsx",
            "backups": tmp_path / "backups",
            "temp": tmp_path / "temp",
        }
    }


def _build_master(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    zh = wb.active
    zh.title = "01_SKU_ZH_CURRENT"
    zh.append(["Canonical_ID", "SKU", "当前状态"])
    zh.append(["ACT0001001", "1001", "CURRENT"])
    es = wb.create_sheet("02_SKU_ES_CURRENT")
    es.append(["Canonical_ID", "SKU", "当前状态"])
    es.append(["ACT0001001", "1001", "CURRENT"])
    wb.create_sheet("03_PRICE_HISTORY").append(writer.PRICE_HISTORY_HEADERS)
    wb.create_sheet("04_EVENT_HISTORY").append(writer.EVENT_HISTORY_HEADERS)
    wb.create_sheet("05_RUN_LOG").append(writer.RUN_LOG_HEADERS)
    wb.create_sheet("06_REVIEW_QUEUE")
    wb.save(path)
    wb.close()


def _archive_row(official_sku=None):
    return [
        "vivienda_813161_0", official_sku, dt.date(2026, 4, 5), "家居生活", "Vivienda",
        "Keter Madeira 花盆", "Maceta Keter Madeira", 8.48, "Ø 40x37.5 厘米",
        "Ø 40x37.5 cm", "8,48 欧元/件", "8,48 €/ud.", None, None,
        "品名/一级类目/规格/历史售价/单价/促销", "UNMATCHED_ARCHIVE_ID",
        "es.xlsx", "zh.xlsx", "Sheet1 / 商品全量", "图片暂不纳入",
    ]


def test_writer_preserves_and_validates_april_archive(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    archive = wb.create_sheet("07_APRIL_ARCHIVE")
    archive.append(writer.APRIL_ARCHIVE_HEADERS)
    archive.append(_archive_row())
    wb.save(cfg["paths"]["master"])
    wb.close()

    writer.write_master(
        cfg,
        updated_records={
            "1001": {"canonical_id": "ACT0001001", "sku": "1001", "status": "CURRENT"}
        },
        price_events=[],
        event_events=[],
        dry_run=False,
    )

    wb = openpyxl.load_workbook(cfg["paths"]["master"], read_only=True)
    assert "07_APRIL_ARCHIVE" in wb.sheetnames
    assert wb["07_APRIL_ARCHIVE"].max_row == 2
    assert wb["07_APRIL_ARCHIVE"]["A2"].value == "vivienda_813161_0"
    wb.close()


def test_april_archive_rejects_official_sku_inference(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    archive = wb.create_sheet("07_APRIL_ARCHIVE")
    archive.append(writer.APRIL_ARCHIVE_HEADERS)
    archive.append(_archive_row("813161"))
    wb.save(cfg["paths"]["master"])
    wb.close()

    with pytest.raises(RuntimeError, match="official SKU must remain blank"):
        writer._validate(cfg["paths"]["master"])


def test_april_archive_rejects_duplicate_archive_ids(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    archive = wb.create_sheet("07_APRIL_ARCHIVE")
    archive.append(writer.APRIL_ARCHIVE_HEADERS)
    archive.append(_archive_row())
    archive.append(_archive_row())
    wb.save(cfg["paths"]["master"])
    wb.close()

    with pytest.raises(RuntimeError, match="duplicate archive ID"):
        writer._validate(cfg["paths"]["master"])
