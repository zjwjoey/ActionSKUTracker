import json
from pathlib import Path

import openpyxl
from PIL import Image

from action_tracker.exporting.template1_service import export_template1

from test_exporting import _cfg, _record, _run_log, _write_dictionary, _write_master, _write_snapshot


def _write_seed(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ACTION商品上下架明细"
    sheet.append(["序号", "编号", "中文品名", "图片链接", "商品链接", "26.08.24"])
    sheet.append([1, "9001", "历史商品", "https://images.example/9001.jpg", "https://www.action.com/es-es/p/9001/", 1])
    workbook.save(path)
    workbook.close()


def test_template1_builds_history_union_and_three_sheets(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-26_130145"
    record = _record("1001", last_seen="2026-08-26")
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-26")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-26", [record])
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record)

    seed = tmp_path / "seed.xlsx"
    _write_seed(seed)
    history_config = tmp_path / "history_sources.yaml"
    history_config.write_text(
        "seed:\n"
        f"  path: '{seed.as_posix()}'\n"
        "  sheet: ACTION商品上下架明细\n"
        "  sku_header: 编号\n"
        "  presence_capability: true\n"
        "  absence_capability: true\n"
        "  observation_complete: true\n"
        "  evidence_level: A\n"
        "  fields:\n"
        "    name_zh: 中文品名\n"
        "    image_url: 图片链接\n"
        "    product_url: 商品链接\n"
        "sources: []\n",
        encoding="utf-8",
    )
    cfg["history_sources_path"] = history_config

    result = export_template1(cfg, export_date="2026-08-26", run_id=run_id)
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        assert workbook.sheetnames == ["商品上下架明细", "今日西班牙语清单", "今日中文清单"]
        history = workbook["商品上下架明细"]
        headers = [cell.value for cell in history[1]]
        assert headers[:5] == ["序号", "编号", "中文品名", "品牌", "一级类目（中文）"]
        assert headers[-2:] == ["26.08.24", "26.08.26"]
        values = {str(history.cell(row=row, column=2).value): history.cell(row=row, column=len(headers)).value for row in range(2, history.max_row + 1)}
        assert values == {"1001": 1, "9001": 0}
        assert history.freeze_panes == "A2"
        assert workbook["今日西班牙语清单"].max_row == 2
        assert workbook["今日中文清单"].max_row == 2
    finally:
        workbook.close()
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["history_union_sku_count"] == 2
    assert manifest["current_valid_sku_count"] == 1
    assert manifest["presence_one_count"] == 1

    first_workbook = openpyxl.load_workbook(result["output"], read_only=True, data_only=True)
    try:
        first_values = {
            name: [tuple(row) for row in first_workbook[name].iter_rows(values_only=True)]
            for name in first_workbook.sheetnames
        }
    finally:
        first_workbook.close()
    second = export_template1(cfg, export_date="2026-08-26", run_id=run_id)
    workbook = openpyxl.load_workbook(second["output"], read_only=True, data_only=True)
    try:
        second_values = {
            name: [tuple(row) for row in workbook[name].iter_rows(values_only=True)]
            for name in workbook.sheetnames
        }
        assert second_values == first_values
        headers = [cell.value for cell in next(workbook["商品上下架明细"].iter_rows(max_row=1))]
        assert headers.count("26.08.26") == 1
    finally:
        workbook.close()


def test_template1_with_images_only_embeds_today_chinese_sheet(tmp_path):
    cfg = _cfg(tmp_path)
    cfg["paths"]["images"] = tmp_path / "images"
    run_id = "2026-08-26_130145"
    record = _record("1001", last_seen="2026-08-26")
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-26")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-26", [record])
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record)
    seed = tmp_path / "seed.xlsx"
    _write_seed(seed)
    history_config = tmp_path / "history_sources.yaml"
    history_config.write_text(
        "seed:\n"
        f"  path: '{seed.as_posix()}'\n"
        "  sheet: ACTION商品上下架明细\n"
        "  sku_header: 编号\n"
        "  presence_capability: true\n"
        "  absence_capability: true\n"
        "  observation_complete: true\n"
        "  evidence_level: A\n"
        "  fields:\n"
        "    name_zh: 中文品名\n"
        "    image_url: 图片链接\n"
        "    product_url: 商品链接\n"
        "sources: []\n",
        encoding="utf-8",
    )
    cfg["history_sources_path"] = history_config
    derivative = cfg["paths"]["images"] / "derivatives" / "excel_250"
    derivative.mkdir(parents=True)
    Image.new("RGB", (250, 250), "white").save(derivative / "1001.png")

    result = export_template1(cfg, export_date="2026-08-26", run_id=run_id, with_images=True)
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        assert len(workbook["今日中文清单"]._images) == 1
        assert len(workbook["今日西班牙语清单"]._images) == 0
        assert len(workbook["商品上下架明细"]._images) == 0
    finally:
        workbook.close()
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["zh_image_embedded_count"] == 1
    assert manifest["zh_image_missing_count"] == 0
