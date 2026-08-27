import csv
import hashlib
import json
from pathlib import Path
from types import SimpleNamespace

import pytest

from action_tracker.dictionary import DICTIONARY_BASELINE_FILENAMES
import action_tracker.dictionary_apply as dictionary_apply_module
from action_tracker.dictionary_apply import (
    DictionaryApplyError, _commit_allowlisted, _dictionary_binding_is_valid, _gate_errors, _load_apply_master_records,
    _recover_interrupted_apply,
    _preview_rows, dictionary_apply,
)
from action_tracker.dictionary_resolver import FieldResolution, RecordResolution

from test_exporting import _cfg, _record, _run_log, _write_dictionary, _write_master, _write_snapshot


def test_dictionary_apply_dry_run_writes_preview_without_changing_master(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-26_130145"
    record = _record("1001", last_seen="2026-08-26")
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-26")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-26", [record])
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record)
    before = hashlib.sha256(cfg["paths"]["master"].read_bytes()).hexdigest()

    result = dictionary_apply(cfg, run_id=run_id, dry_run=True)
    output_dir = Path(result["output_dir"])
    assert result["dry_run"] is True
    assert (output_dir / "apply_preview.csv").exists()
    assert (output_dir / "review_required.csv").exists()
    manifest = json.loads((output_dir / "apply_manifest.json").read_text(encoding="utf-8"))
    assert manifest["master_hash_before"] == before
    assert manifest["formal_write"] is False
    assert hashlib.sha256(cfg["paths"]["master"].read_bytes()).hexdigest() == before


def test_dictionary_apply_formal_write_is_blocked(tmp_path):
    cfg = _cfg(tmp_path)
    with pytest.raises(DictionaryApplyError, match="PRODUCTION_DICTIONARY_APPLY_DISABLED"):
        dictionary_apply(cfg, run_id="2026-08-26_130145", dry_run=False)


def test_production_string_false_is_rejected_not_treated_as_truthy(tmp_path):
    cfg = _cfg(tmp_path)
    cfg["dictionary_apply"] = {"production_enabled": "false"}
    with pytest.raises(DictionaryApplyError, match="PRODUCTION_DICTIONARY_APPLY_CONFIG_INVALID"):
        dictionary_apply(cfg, run_id="2026-08-26_130145", dry_run=False)


def test_apply_manifest_has_gate_and_diff_contract(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-26_130145"
    record = _record("1001", last_seen="2026-08-26")
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-26")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-26", [record])
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record)
    result = dictionary_apply(cfg, run_id=run_id, dry_run=True)
    manifest = json.loads((Path(result["output_dir"]) / "apply_manifest.json").read_text(encoding="utf-8"))
    required = {"run_id", "run_date", "master_hash_before", "temporary_master_hash", "master_hash_after_if_committed",
                "dictionary_hash", "dictionary_baseline_hash", "total_current_skus", "auto_ready_count",
                "review_required_count", "source_blocked_count", "preview_field_count", "actual_changed_field_count",
                "unchanged_field_count", "immutable_fact_change_count", "production_enabled", "committed", "generated_at"}
    assert required.issubset(manifest)
    rows = list(csv.DictReader((Path(result["output_dir"]) / "field_diff.csv").open(encoding="utf-8-sig")))
    assert {row["field"] for row in rows} <= {"name_zh", "cat1_zh", "cat2_zh", "spec_zh", "desc_zh", "details_zh"}


def test_diff_does_not_count_equal_value_as_actual_change():
    record = {"sku": "1", "name_zh": "同名", "cat1_zh": "家务清洁", "cat2_zh": "", "spec_zh": "", "desc_zh": "", "details_zh": ""}
    fields = {
        key: FieldResolution(value, "manual_override", "READY")
        for key, value in (("name", "同名"), ("cat1", "家务清洁"), ("cat2", ""), ("spec", ""), ("description", ""), ("details", ""))
    }
    resolution = RecordResolution("1", fields, "MATCH", "OK", "AUTO_READY", ())
    rows, summary = _preview_rows([record], [resolution])
    assert rows == []
    assert summary["actual_changed_field_count"] == 0
    assert summary["unchanged_field_count"] == 6


def test_apply_gate_rejects_non_formal_or_failed_qa_before_write(tmp_path):
    cfg = _cfg(tmp_path)
    cfg["dictionary_apply"] = {"production_enabled": True}
    run_id = "2026-08-26_130145"
    record = _record("1001", last_seen="2026-08-26")
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-26")], [record])
    snapshot = cfg["paths"]["snapshots"] / "2026-08-26" / run_id
    snapshot.mkdir(parents=True)
    (snapshot / "run_report.json").write_text(json.dumps({"run_id": run_id, "run_date": "2026-08-26", "dry_run": False, "commit_status": "FULL_COMMIT"}), encoding="utf-8")
    (snapshot / "qa_report.json").write_text(json.dumps({"state": "FAIL", "passed": False}), encoding="utf-8")
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record)
    before = hashlib.sha256(cfg["paths"]["master"].read_bytes()).hexdigest()
    source = SimpleNamespace(directory=snapshot, records=())
    errors = _gate_errors(cfg, source, [], object(), {}, before)
    assert "QA_NOT_PASS" in errors
    assert hashlib.sha256(cfg["paths"]["master"].read_bytes()).hexdigest() == before


def test_dictionary_binding_requires_baseline_and_selected_file_hashes(tmp_path):
    cfg = _cfg(tmp_path)
    record = _record("1001")
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record)
    entries = {}
    for filename in DICTIONARY_BASELINE_FILENAMES:
        payload = (cfg["paths"]["dictionary_baseline"] / filename).read_bytes()
        entries[filename] = {"sha256": hashlib.sha256(payload).hexdigest()}
    (cfg["paths"]["dictionary_baseline"] / "baseline_manifest.json").write_text(json.dumps({"files": entries}), encoding="utf-8")
    import shutil
    shutil.copytree(cfg["paths"]["dictionary_baseline"], cfg["paths"]["dictionary"])
    context = SimpleNamespace(directory=cfg["paths"]["dictionary"])
    assert _dictionary_binding_is_valid(cfg, context)
    with (cfg["paths"]["dictionary"] / "product_dictionary.csv").open("a", encoding="utf-8") as handle:
        handle.write("tampered\n")
    assert not _dictionary_binding_is_valid(cfg, context)


def test_apply_master_reader_rejects_duplicate_sku(tmp_path):
    cfg = _cfg(tmp_path)
    record = _record("1001")
    _write_master(cfg["paths"]["master"], [], [record])
    import openpyxl
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    for sheet in ("01_SKU_ZH_CURRENT", "02_SKU_ES_CURRENT"):
        ws = wb[sheet]
        ws.append([cell.value for cell in ws[2]])
    wb.save(cfg["paths"]["master"])
    wb.close()
    with pytest.raises(DictionaryApplyError, match="MASTER_DUPLICATE_SKU"):
        _load_apply_master_records(cfg["paths"]["master"])


def test_apply_master_reader_rejects_nonempty_row_with_empty_sku(tmp_path):
    cfg = _cfg(tmp_path)
    record = _record("1001")
    _write_master(cfg["paths"]["master"], [], [record])
    import openpyxl
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    ws = wb["01_SKU_ZH_CURRENT"]
    ws.cell(row=3, column=3).value = "orphan data"
    wb.save(cfg["paths"]["master"])
    wb.close()
    with pytest.raises(DictionaryApplyError, match="MASTER_EMPTY_SKU"):
        _load_apply_master_records(cfg["paths"]["master"])


def test_interrupted_apply_without_backup_pointer_blocks_overwrite(tmp_path):
    cfg = _cfg(tmp_path)
    run_dir = cfg["paths"]["dictionary"] / "apply" / "recovery-test"
    run_dir.mkdir(parents=True)
    (run_dir / "apply_manifest.json").write_text(json.dumps({
        "commit_state": "PENDING", "master_hash_before": "abc", "backup_path": None,
    }), encoding="utf-8")
    with pytest.raises(DictionaryApplyError, match="MASTER_RECOVERY_REQUIRED"):
        _recover_interrupted_apply(cfg, "recovery-test")


def test_post_commit_validation_failure_restores_exact_backup(tmp_path, monkeypatch):
    """一旦替换后验证失败，字典 Apply 必须恢复替换前的原始 Master。"""
    cfg = _cfg(tmp_path)
    cfg["paths"]["backups"] = tmp_path / "backups"
    cfg["paths"]["temp"] = tmp_path / "temp"
    record = _record("1001")
    _write_master(cfg["paths"]["master"], [], [record])
    import openpyxl
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    for sheet in ("03_PRICE_HISTORY", "04_EVENT_HISTORY", "06_REVIEW_QUEUE"):
        wb.create_sheet(sheet)
    wb.save(cfg["paths"]["master"])
    wb.close()
    original = cfg["paths"]["master"].read_bytes()
    master_records = _load_apply_master_records(cfg["paths"]["master"])
    fields = {
        field: FieldResolution("", "none", "READY")
        for field in ("name", "cat1", "cat2", "spec", "description", "details")
    }
    resolution = RecordResolution("1001", fields, "MATCH", "OK", "AUTO_READY", (), "CONFIRMED")
    real_assert = dictionary_apply_module._assert_master_safe
    calls = {"count": 0}

    def fail_only_after_replace(before, staged):
        calls["count"] += 1
        if calls["count"] == 2:
            raise DictionaryApplyError("forced post-commit verification failure")
        return real_assert(before, staged)

    monkeypatch.setattr(dictionary_apply_module, "_assert_master_safe", fail_only_after_replace)
    before_hash = hashlib.sha256(original).hexdigest()
    with pytest.raises(DictionaryApplyError, match="POST_COMMIT_FAILURE_ROLLED_BACK"):
        _commit_allowlisted(cfg, [record], [resolution], master_records, before_hash, "rollback-test")
    assert cfg["paths"]["master"].read_bytes() == original
