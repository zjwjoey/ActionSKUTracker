"""规范 §60 测试 20-23：写入安全 / Excel 类型 / 幂等。"""
import datetime as dt
from pathlib import Path

import openpyxl
import pytest

from action_tracker.excel import writer
from action_tracker.services.change import compute_changes


def _cfg(tmp_path: Path) -> dict:
    return {
        "paths": {
            "master": tmp_path / "master" / "Action_Master.xlsx",
            "backups": tmp_path / "backups",
            "temp": tmp_path / "temp",
        }
    }


def _build_master(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "01_SKU_ZH_CURRENT"
    ws.append(["Canonical_ID", "SKU", "中文品名", "西班牙语品名", "当前售价 (€)", "首次发现日期", "最后确认存在日期", "商品链接"])
    ws.append(["ACT0001001", "1001", "品", "Nombre", 2.99, dt.date(2026, 1, 1), "2026-08-10", "https://x/p/1001/"])

    ws2 = wb.create_sheet("02_SKU_ES_CURRENT")
    ws2.append(["Canonical_ID", "SKU", "西班牙语品名", "当前售价 (€)", "首次发现日期", "最后确认存在日期", "商品链接"])
    ws2.append(["ACT0001001", "1001", "Nombre", 2.99, dt.date(2026, 1, 1), "2026-08-10", "https://x/p/1001/"])

    ws3 = wb.create_sheet("03_PRICE_HISTORY")
    ws3.append(["Canonical_ID", "SKU", "日期", "旧售价 (€)", "新售价 (€)", "变化类型"])
    ws4 = wb.create_sheet("04_EVENT_HISTORY")
    ws4.append(["Canonical_ID", "SKU", "日期", "事件类型", "旧值", "新值"])
    ws5 = wb.create_sheet("05_RUN_LOG")
    ws5.append(writer.RUN_LOG_HEADERS)
    ws6 = wb.create_sheet("06_REVIEW_QUEUE")
    ws6.append(writer.REVIEW_HEADERS)
    wb.save(path)


# ---- 测试 22/23：Excel 日期是真日期、价格是真数值 ----
def test_t22_23_cell_types():
    assert isinstance(writer._cell("2026-08-10"), dt.date)
    assert not isinstance(writer._cell("2026-08-10"), str)
    assert isinstance(writer._cell(3.99), float)
    assert writer._cell(3.99) == 3.99


# ---- 测试 22/23b：写入后回读类型 ----
def test_t22_23_write_readback(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    updated = {
        "1001": {
            "canonical_id": "ACT0001001", "sku": "1001", "current_price": 3.99,
            "last_seen": "2026-08-11", "name_es": "Nuevo nombre",
        }
    }
    writer.write_master(cfg, updated_records=updated, price_events=[], event_events=[], dry_run=False)
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    ws = wb["01_SKU_ZH_CURRENT"]
    price = ws.cell(row=2, column=5).value
    last_seen = ws.cell(row=2, column=7).value
    name = ws.cell(row=2, column=4).value
    wb.close()
    assert price == 3.99 and isinstance(price, float)
    assert isinstance(last_seen, dt.date)
    assert name == "Nuevo nombre"


# ---- 测试 21：正式 Master 写入失败 -> 自动保留原 Master ----
def test_t21_write_failure_keeps_master(tmp_path, monkeypatch):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    orig = cfg["paths"]["master"].read_bytes()
    calls = {"n": 0}
    real_save = openpyxl.Workbook.save

    def fake_save(self, *a, **k):
        calls["n"] += 1
        if calls["n"] == 1:
            raise OSError("模拟写盘失败")
        return real_save(self, *a, **k)

    monkeypatch.setattr(openpyxl.Workbook, "save", fake_save)
    with pytest.raises(OSError):
        writer.write_master(cfg, updated_records={}, price_events=[], event_events=[], dry_run=False)
    # 原 Master 未被替换/损坏
    assert cfg["paths"]["master"].read_bytes() == orig
    # 备份仍产生
    backups = list(cfg["paths"]["backups"].glob("Action_Master_*.xlsx"))
    assert backups, "应自动备份"


def test_backup_names_are_unique_and_restore_is_atomic(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    original = cfg["paths"]["master"].read_bytes()
    first = writer._backup(cfg["paths"]["master"], cfg["paths"]["backups"])
    second = writer._backup(cfg["paths"]["master"], cfg["paths"]["backups"])
    assert first != second and first.read_bytes() == original and second.read_bytes() == original
    cfg["paths"]["master"].write_bytes(b"broken")
    writer.restore_master_from_backup(first, cfg["paths"]["master"])
    assert cfg["paths"]["master"].read_bytes() == original


# ---- 测试 20：重复运行同一 run -> 不重复增加 PRICE_HISTORY/EVENT_HISTORY ----
def test_t20_no_duplicate_events_on_repeat():
    before = {"sku": "1001", "current_price": 4.99, "raw_tags": "Nuevo", "canonical_id": "ACT0001001"}
    after = {"sku": "1001", "current_price": 3.99, "raw_tags": "Nuevo", "canonical_id": "ACT0001001"}
    first = compute_changes("1001", "ACT0001001", before, after, "2026-08-10", 0.01, 1000)
    assert len(first.price_events) == 1
    assert first.price_events[0]["变化类型"] == "DOWN"
    # 第二次运行：before 已是 after（价格不再变化）-> 无事件
    second = compute_changes("1001", "ACT0001001", after, after, "2026-08-11", 0.01, 1000)
    assert second.price_events == []
    assert second.badge_events == []


# ---- 写入行数断言 ----
def test_writer_updates_row_count(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    updated = {"1001": {"canonical_id": "ACT0001001", "sku": "1001", "current_price": 2.50}}
    # 新 SKU 追加
    updated["2002"] = {"canonical_id": "ACT0002002", "sku": "2002", "current_price": 5.0, "name_es": "Nuevo"}
    writer.write_master(cfg, updated_records=updated, price_events=[], event_events=[], dry_run=False)
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    ws = wb["01_SKU_ZH_CURRENT"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert len(rows) == 2  # 原 1001 更新 + 2002 追加


def test_stage_master_revises_existing_run_log_without_duplicate(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    row = {header: None for header in writer.RUN_LOG_HEADERS}
    row.update({"Run ID": "R1", "CONTENT_CHANGE": 99, "PROMO_START": 8})
    wb["05_RUN_LOG"].append([row[header] for header in writer.RUN_LOG_HEADERS])
    wb.save(cfg["paths"]["master"])
    wb.close()

    tmp = writer.stage_master(
        cfg, updated_records={}, price_events=[], event_events=[],
        run_log_revisions={"R1": {"CONTENT_CHANGE": 4, "PROMO_START": 1}},
    )
    writer.commit_master(tmp, cfg["paths"]["master"])
    wb = openpyxl.load_workbook(cfg["paths"]["master"], data_only=True)
    rows = list(wb["05_RUN_LOG"].iter_rows(values_only=True))
    wb.close()
    assert len(rows) == 2
    assert rows[1][writer.RUN_LOG_HEADERS.index("CONTENT_CHANGE")] == 4
    assert rows[1][writer.RUN_LOG_HEADERS.index("PROMO_START")] == 1
