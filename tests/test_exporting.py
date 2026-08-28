import csv
import hashlib
import json
from pathlib import Path

import openpyxl
import pytest

from action_tracker.dictionary import (
    BRAND_DICTIONARY_HEADERS,
    CATEGORY_DICTIONARY_HEADERS,
    MODEL_TRANSLATION_HEADERS,
    OVERRIDE_HEADERS,
    PRODUCT_DICTIONARY_HEADERS,
    SOURCE_DAMAGE_HEADERS,
    TERM_DICTIONARY_HEADERS,
)
from action_tracker.excel.reader import ES_MAP, ZH_MAP
from action_tracker.excel.writer import RUN_LOG_HEADERS
from action_tracker.exporting.service import ExportValidationError, export_catalog


def _cfg(tmp_path: Path) -> dict:
    root = Path(__file__).resolve().parents[1]
    return {
        "project_root": root,
        "paths": {
            "master": tmp_path / "master.xlsx",
            "snapshots": tmp_path / "snapshots",
            "exports": tmp_path / "exports",
            "state": tmp_path / "state",
            "dictionary": tmp_path / "dictionary",
            "dictionary_baseline": tmp_path / "dictionary_baseline",
        },
    }


def _record(sku="1001", *, status="CURRENT", current=2.5, original=3.0, last_seen="2026-08-24"):
    return {
        "canonical_id": f"ACT{sku.zfill(7)}", "sku": sku, "name_es": "Producto español",
        "name_zh": "中文商品", "cat1_es": "Hogar", "cat2_es": "Limpieza",
        "cat1_zh": "家居布置", "cat2_zh": "清洁用品", "spec_es": "2 unidades", "spec_zh": "2件装",
        "current_price": current, "original_price": original, "unit_price": "1,25 €/ud.",
        "desc_es": "Descripción española", "details_es": "Detalles españoles",
        "desc_zh": "中文描述", "details_zh": "中文详情",
        "product_url": f"https://www.action.com/es-es/p/{sku}/", "image_url": f"https://images.example/{sku}.jpg",
        "is_new_badge": True, "promotion": True, "sustainable": False, "discount": 10,
        "raw_tags": "Nuevo", "status": status, "first_seen": "2026-08-10", "last_seen": last_seen,
    }


def _write_master(path: Path, run_rows: list[dict], records: list[dict]):
    wb = openpyxl.Workbook()
    ws_zh = wb.active
    ws_zh.title = "01_SKU_ZH_CURRENT"
    zh_headers = list(ZH_MAP)
    ws_zh.append(zh_headers)
    for record in records:
        ws_zh.append([record.get(ZH_MAP[header]) for header in zh_headers])
    ws_es = wb.create_sheet("02_SKU_ES_CURRENT")
    es_headers = list(ES_MAP)
    ws_es.append(es_headers)
    for record in records:
        ws_es.append([record.get(ES_MAP[header]) for header in es_headers])
    run_log = wb.create_sheet("05_RUN_LOG")
    run_log.append(RUN_LOG_HEADERS)
    for row in run_rows:
        run_log.append([row.get(header) for header in RUN_LOG_HEADERS])
    wb.save(path)
    wb.close()


def _run_log(run_id: str, date: str):
    return {"Run ID": run_id, "运行日期": date, "QA状态": "PASS", "运行状态": "SUCCESS", "备注": "正式写库"}


def _write_snapshot(root: Path, run_id: str, date: str, records: list[dict], *, commit="FULL_COMMIT", passed=True):
    directory = root / date / run_id
    directory.mkdir(parents=True)
    (directory / "run_report.json").write_text(json.dumps({"run_id": run_id, "run_date": date, "commit_status": commit}), encoding="utf-8")
    (directory / "qa_report.json").write_text(json.dumps({"state": "PASS", "passed": passed}), encoding="utf-8")
    with (directory / "products_normalized.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0]))
        writer.writeheader()
        writer.writerows(records)


def _write_csv(path: Path, headers: list[str], rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _write_dictionary(directory: Path, record: dict, *, product=None, manual=None, model=None, categories=None, terms=None):
    payload = "\x1f".join(str(record.get(key) or "").strip() for key in ("name_es", "cat1_es", "cat2_es", "spec_es"))
    source_hash = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    _write_csv(directory / "product_dictionary.csv", PRODUCT_DICTIONARY_HEADERS, [product or {
        "sku": record["sku"], "canonical_id": record["canonical_id"], "name_es_raw": record["name_es"],
        "name_zh_standard": "字典品名", "brand_id": "BrandX", "cat1_es": record["cat1_es"],
        "cat2_es": record["cat2_es"], "cat1_zh": "家务清洁", "cat2_zh": "清洁用品",
        "spec_es_raw": record["spec_es"], "spec_zh_standard": "字典规格", "source_hash": source_hash,
        "translation_status": "MODEL_TRANSLATED", "review_status": "UNREVIEWED", "locked": "0",
    }])
    _write_csv(directory / "brand_dictionary.csv", BRAND_DICTIONARY_HEADERS, [{"brand_id": "BrandX", "canonical_name": "BrandX"}])
    _write_csv(directory / "category_dictionary.csv", CATEGORY_DICTIONARY_HEADERS, categories or [{
        "cat1_es": record["cat1_es"], "cat2_es": record["cat2_es"], "cat1_code": "C08",
        "cat1_zh": "家务清洁", "cat2_zh": "清洁用品", "review_status": "HUMAN_REVIEWED",
    }])
    _write_csv(directory / "term_dictionary.csv", TERM_DICTIONARY_HEADERS, terms or [{
        "term_es": "ud.", "term_zh": "件", "term_type": "unit", "keep_original": "0", "review_status": "HUMAN_REVIEWED",
    }])
    _write_csv(directory / "manual_overrides.csv", OVERRIDE_HEADERS, manual or [])
    _write_csv(directory / "model_translation_overrides.csv", MODEL_TRANSLATION_HEADERS, model or [])
    _write_csv(directory / "source_damage_report.csv", SOURCE_DAMAGE_HEADERS, [])
    return source_hash


def test_es_export_reads_latest_formal_master_and_keeps_sources_read_only(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    records = [_record("1002", original=2.5), _record("1001")]
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], records)
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", records)
    before = hashlib.sha256(cfg["paths"]["master"].read_bytes()).hexdigest()

    result = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)

    assert result["source_kind"] == "MASTER_CURRENT"
    assert result["sku_count"] == 2
    assert hashlib.sha256(cfg["paths"]["master"].read_bytes()).hexdigest() == before
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        ws = workbook["商品全量"]
        assert [cell.value for cell in ws[1]] == ["图片", "编号", "标题", "分类1", "分类2", "规格", "折后价", "原价", "单价", "描述", "产品详情", "图片链接", "商品链接", "备注"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == "A1:N3"
        assert ws.cell(2, 2).value == "1001"  # 稳定按 SKU 排序
        assert ws.cell(2, 7).value == 2.5
        assert ws.cell(2, 8).value == 3.0
        assert ws.cell(3, 8).value is None  # 原价等于当前价，不能显示为促销原价
        assert ws.cell(2, 13).value == "查看商品"
        assert ws.cell(2, 13).hyperlink.target == "https://www.action.com/es-es/p/1001/"
    finally:
        workbook.close()
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["run_id"] == run_id
    assert manifest["source_master_file_hash"] == before
    assert len(manifest["source_master_hash"]) == 64


def test_es_historical_export_reads_formal_snapshot_not_newer_master(tmp_path):
    cfg = _cfg(tmp_path)
    historical_run = "2026-08-24_010000"
    current_run = "2026-08-25_010000"
    historical = [_record("1001", current=2.5, last_seen="2026-08-24")]
    current = [_record("2001", current=4.0, last_seen="2026-08-25")]
    _write_master(cfg["paths"]["master"], [_run_log(historical_run, "2026-08-24"), _run_log(current_run, "2026-08-25")], current)
    _write_snapshot(cfg["paths"]["snapshots"], historical_run, "2026-08-24", historical)
    _write_snapshot(cfg["paths"]["snapshots"], current_run, "2026-08-25", current)

    result = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)

    assert result["source_kind"] == "FORMAL_SNAPSHOT"
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        assert workbook["商品全量"].cell(2, 2).value == "1001"
    finally:
        workbook.close()
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["source_master_file_hash"] is None
    assert manifest["run_id"] == historical_run


def test_export_rejects_duplicate_or_non_current_snapshot_sku(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    current = [_record("2001", last_seen="2026-08-25")]
    _write_master(cfg["paths"]["master"], [_run_log("2026-08-25_010000", "2026-08-25")], current)
    bad = [_record("1001"), _record("1001", status="HISTORICAL")]
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", bad)

    with pytest.raises(ExportValidationError, match="DUPLICATE_SKU"):
        export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    assert not cfg["paths"]["exports"].exists()


def test_es_export_rejects_chinese_pollution_in_spanish_fact_fields(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    bad = [_record("1001")]
    bad[0]["name_es"] = "中文商品"
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], bad)
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", bad)

    with pytest.raises(ExportValidationError, match="NON_SPANISH_FACT: 1001/name_es"):
        export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)


def test_zh_export_uses_field_priority_and_preserves_fact_columns(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    record = _record("1001")
    record.update({"desc_zh": "", "details_zh": "中文详情", "unit_price": "1,25 €/ud."})
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", [record])
    source_hash = _write_dictionary(
        cfg["paths"]["dictionary_baseline"], record,
        product={
            "sku": "1001", "canonical_id": record["canonical_id"], "name_es_raw": record["name_es"],
            "name_zh_standard": "字典品名", "brand_id": "BrandX", "cat1_es": record["cat1_es"],
            "cat2_es": record["cat2_es"], "cat1_zh": "", "cat2_zh": "", "spec_es_raw": record["spec_es"],
            "spec_zh_standard": "", "source_hash": source_hash if False else "", "translation_status": "MODEL_TRANSLATED",
        },
        manual=[{"scope": "product", "key": "1001", "field": "name_zh_standard", "value": "人工品名"}],
    )
    # 写入时 product 的空 source_hash 是为了让类目走正式映射、规格走模型结果。
    _write_dictionary(
        cfg["paths"]["dictionary_baseline"], record,
        product={
            "sku": "1001", "canonical_id": record["canonical_id"], "name_es_raw": record["name_es"],
            "name_zh_standard": "字典品名", "brand_id": "BrandX", "cat1_es": record["cat1_es"],
            "cat2_es": record["cat2_es"], "cat1_zh": "", "cat2_zh": "", "spec_es_raw": record["spec_es"],
            "spec_zh_standard": "", "source_hash": "stale", "translation_status": "MODEL_TRANSLATED",
        },
        manual=[{"scope": "product", "key": "1001", "field": "name_zh_standard", "value": "人工品名"}],
        model=[{"sku": "1001", "source_hash": source_hash, "spec_zh_standard": "模型规格", "quality_status": "OK"}],
    )

    es = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    zh = export_catalog(cfg, language="zh", export_date="2026-08-24", no_images=True)

    es_wb = openpyxl.load_workbook(es["output"], data_only=True)
    zh_wb = openpyxl.load_workbook(zh["output"], data_only=True)
    try:
        es_row = [cell.value for cell in es_wb["商品全量"][2]]
        zh_row = [cell.value for cell in zh_wb["商品全量"][2]]
        assert zh_row[2] == "人工品名"
        assert zh_row[3:5] == ["家务清洁", "清洁用品"]
        assert zh_row[5] == "模型规格"
        assert zh_row[8] == "1,25 €/件"
        assert zh_row[9] == "Descripción española"
        assert "中文描述待审核" in zh_row[13]
        assert "CURRENT" not in zh_row[13]
        assert [zh_row[index] for index in (1, 6, 7, 11, 12)] == [es_row[index] for index in (1, 6, 7, 11, 12)]
    finally:
        es_wb.close()
        zh_wb.close()
    manifest = json.loads(Path(zh["manifest"]).read_text(encoding="utf-8"))
    assert manifest["dictionary_fallback_counts"] == {"中文描述待审核": 1}


def test_zh_export_adds_confirmed_brand_marker_but_keeps_manual_title(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    record = _record("1001")
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", [record])
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record)
    _write_csv(cfg["paths"]["dictionary_baseline"] / "brand_dictionary.csv", BRAND_DICTIONARY_HEADERS, [{
        "brand_id": "BrandX", "canonical_name": "BrandX", "confidence": "REFERENCE",
    }])

    result = export_catalog(cfg, language="zh", export_date="2026-08-24", no_images=True)
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        assert workbook["商品全量"].cell(2, 3).value == "BrandX牌字典品名"
    finally:
        workbook.close()

    _write_dictionary(
        cfg["paths"]["dictionary_baseline"], record,
        manual=[{"scope": "product", "key": "1001", "field": "name_zh_standard", "value": "人工品名"}],
    )
    _write_csv(cfg["paths"]["dictionary_baseline"] / "brand_dictionary.csv", BRAND_DICTIONARY_HEADERS, [{
        "brand_id": "BrandX", "canonical_name": "BrandX", "confidence": "REFERENCE",
    }])
    manual_result = export_catalog(cfg, language="zh", export_date="2026-08-24", no_images=True)
    manual_workbook = openpyxl.load_workbook(manual_result["output"], data_only=True)
    try:
        assert manual_workbook["商品全量"].cell(2, 3).value == "人工品名"
    finally:
        manual_workbook.close()


def test_zh_export_stale_dictionary_value_falls_back_without_dropping_sku(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    record = _record("1001")
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", [record])
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record, product={
        "sku": "1001", "canonical_id": record["canonical_id"], "name_es_raw": record["name_es"],
        "name_zh_standard": "过期中文", "brand_id": "BrandX", "cat1_es": record["cat1_es"],
        "cat2_es": record["cat2_es"], "cat1_zh": "家务清洁", "cat2_zh": "清洁用品",
        "spec_es_raw": record["spec_es"], "spec_zh_standard": "过期规格", "source_hash": "stale",
        "translation_status": "MODEL_TRANSLATED",
    })

    result = export_catalog(cfg, language="zh", export_date="2026-08-24", no_images=True)

    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        row = [cell.value for cell in workbook["商品全量"][2]]
        assert row[1] == "1001"
        assert row[2] == "Producto español"
        assert row[5] == "2 unidades"
        assert "中文品名待审核" in row[13]
        assert "中文规格待审核" in row[13]
    finally:
        workbook.close()


def test_export_rejects_dry_run_and_qa_fail_sources(tmp_path):
    cfg = _cfg(tmp_path)
    records = [_record("1001")]

    _write_snapshot(cfg["paths"]["snapshots"], "2026-08-24_dry", "2026-08-24", records, commit="DRY_RUN")
    with pytest.raises(ExportValidationError, match="NO_FORMAL_EXPORT_SOURCE"):
        export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)

    # Replace the rejected source with a formal-looking but failed QA run.
    failed = cfg["paths"]["snapshots"] / "2026-08-24" / "2026-08-24_failed"
    failed.mkdir(parents=True)
    (failed / "run_report.json").write_text(
        json.dumps({"run_id": "2026-08-24_failed", "run_date": "2026-08-24", "commit_status": "FULL_COMMIT"}),
        encoding="utf-8",
    )
    (failed / "qa_report.json").write_text(json.dumps({"state": "FAIL", "passed": False}), encoding="utf-8")
    with pytest.raises(ExportValidationError, match="NO_FORMAL_EXPORT_SOURCE"):
        export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)


def test_es_and_zh_exports_have_identical_sku_price_and_link_sets(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    records = [_record("1002", current=4.0, original=5.0), _record("1001", current=2.5, original=3.0)]
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], records)
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", records)
    _write_dictionary(cfg["paths"]["dictionary_baseline"], records[0])

    es = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    zh = export_catalog(cfg, language="zh", export_date="2026-08-24", no_images=True)
    es_wb = openpyxl.load_workbook(es["output"], data_only=True)
    zh_wb = openpyxl.load_workbook(zh["output"], data_only=True)
    try:
        es_rows = {row[1].value: (row[6].value, row[7].value, row[11].value, row[12].value)
                   for row in es_wb["商品全量"].iter_rows(min_row=2)}
        zh_rows = {row[1].value: (row[6].value, row[7].value, row[11].value, row[12].value)
                   for row in zh_wb["商品全量"].iter_rows(min_row=2)}
        assert es_rows == zh_rows
    finally:
        es_wb.close()
        zh_wb.close()


def test_repeated_export_is_data_idempotent(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    records = [_record("1001")]
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], records)
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", records)

    first = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    first_wb = openpyxl.load_workbook(first["output"], data_only=True)
    try:
        first_values = [[cell.value for cell in row] for row in first_wb["商品全量"].iter_rows()]
    finally:
        first_wb.close()
    first_manifest = json.loads(Path(first["manifest"]).read_text(encoding="utf-8"))

    second = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    second_wb = openpyxl.load_workbook(second["output"], data_only=True)
    try:
        second_values = [[cell.value for cell in row] for row in second_wb["商品全量"].iter_rows()]
    finally:
        second_wb.close()
    second_manifest = json.loads(Path(second["manifest"]).read_text(encoding="utf-8"))
    assert second_values == first_values
    for key in ("run_id", "export_date", "sku_count", "source_master_hash", "source_kind", "profile_id", "profile_version"):
        assert second_manifest[key] == first_manifest[key]


def test_failed_output_validation_does_not_replace_previous_file(tmp_path, monkeypatch):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    records = [_record("1001")]
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], records)
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", records)
    first = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    before = Path(first["output"]).read_bytes()

    from action_tracker.exporting import service
    original_verify = service._verify_written_workbook
    monkeypatch.setattr(service, "_verify_written_workbook", lambda *args, **kwargs: (_ for _ in ()).throw(ExportValidationError("forced")))
    with pytest.raises(ExportValidationError, match="forced"):
        export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    monkeypatch.setattr(service, "_verify_written_workbook", original_verify)
    assert Path(first["output"]).read_bytes() == before


def test_historical_export_merges_validated_detail_retry(tmp_path):
    cfg = _cfg(tmp_path)
    historical_run = "2026-08-24_010000"
    current_run = "2026-08-25_010000"
    record = _record("1001", last_seen="2026-08-24")
    incomplete = dict(record, desc_es="", details_es="")
    _write_master(cfg["paths"]["master"], [_run_log(historical_run, "2026-08-24"), _run_log(current_run, "2026-08-25")], [_record("2001", last_seen="2026-08-25")])
    _write_snapshot(cfg["paths"]["snapshots"], historical_run, "2026-08-24", [incomplete])
    retry = cfg["paths"]["snapshots"] / "2026-08-24" / historical_run / "detail_retries" / "retry_1"
    retry.mkdir(parents=True)
    (retry / "detail_retry_report.json").write_text(json.dumps({
        "retry_id": "retry_1", "planned": 1, "completed": 1,
        "pending_after": 0, "detail_retry_pass": True,
    }), encoding="utf-8")
    detail = dict(record, desc_es="Descripción recuperada", details_es="Detalles recuperados")
    (retry / "detail_fetch.jsonl").write_text(json.dumps({"sku": "1001", "detail": detail}, ensure_ascii=False) + "\n", encoding="utf-8")

    result = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        row = [cell.value for cell in workbook["商品全量"][2]]
        assert row[9] == "Descripción recuperada"
        assert row[10] == "Detalles recuperados"
    finally:
        workbook.close()
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["detail_retry_ids"] == ["retry_1"]


def test_historical_export_ignores_incomplete_detail_retry(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    record = _record("1001")
    record["desc_es"] = "Descripción original"
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", [record])
    retry = cfg["paths"]["snapshots"] / "2026-08-24" / run_id / "detail_retries" / "retry_incomplete"
    retry.mkdir(parents=True)
    (retry / "detail_retry_report.json").write_text(json.dumps({
        "retry_id": "retry_incomplete", "planned": 2, "completed": 1,
        "pending_after": 1, "detail_retry_pass": False,
    }), encoding="utf-8")
    (retry / "detail_fetch.jsonl").write_text(json.dumps({
        "sku": "1001", "detail": {"desc_es": "不应覆盖"},
    }, ensure_ascii=False) + "\n", encoding="utf-8")

    result = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        row = [cell.value for cell in workbook["商品全量"][2]]
        assert row[9] == "Descripción original"
    finally:
        workbook.close()
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["detail_retry_ids"] == []


def test_export_rejects_dry_run_and_ambiguous_formal_sources(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", [_record("1001")])
    report = cfg["paths"]["snapshots"] / "2026-08-24" / run_id / "run_report.json"
    report.write_text(json.dumps({"run_id": run_id, "run_date": "2026-08-24", "dry_run": True, "commit_status": "FULL_COMMIT"}), encoding="utf-8")
    with pytest.raises(ExportValidationError, match="NO_FORMAL_EXPORT_SOURCE"):
        export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)

    cfg2 = _cfg(tmp_path / "ambiguous")
    cfg2["paths"]["master"].parent.mkdir(parents=True)
    _write_master(cfg2["paths"]["master"], [_run_log("2026-08-25_010000", "2026-08-25")], [_record("2001", last_seen="2026-08-25")])
    _write_snapshot(cfg2["paths"]["snapshots"], "2026-08-24_010000", "2026-08-24", [_record("1001")])
    _write_snapshot(cfg2["paths"]["snapshots"], "2026-08-24_020000", "2026-08-24", [_record("1002")])
    with pytest.raises(ExportValidationError, match="AMBIGUOUS_FORMAL_EXPORT_SOURCE"):
        export_catalog(cfg2, language="es", export_date="2026-08-24", no_images=True)


def test_zh_export_marks_empty_derived_fields_for_review(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    record = _record("1001")
    record.update({"cat2_es": "", "desc_es": "", "details_es": "", "desc_zh": "", "details_zh": ""})
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", [record])
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record, product={
        "sku": "1001", "canonical_id": record["canonical_id"], "name_es_raw": record["name_es"],
        "name_zh_standard": "字典品名", "brand_id": "BrandX", "cat1_es": record["cat1_es"], "cat2_es": "",
        "cat1_zh": "家务清洁", "cat2_zh": "", "spec_es_raw": record["spec_es"],
        "spec_zh_standard": "字典规格", "source_hash": "", "translation_status": "MODEL_TRANSLATED",
    }, categories=[{
        "cat1_es": record["cat1_es"], "cat2_es": "", "cat1_code": "C08", "cat1_zh": "家务清洁",
        "cat2_zh": "", "review_status": "HUMAN_REVIEWED",
    }])

    result = export_catalog(cfg, language="zh", export_date="2026-08-24", no_images=True)
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        row = [cell.value for cell in workbook["商品全量"][2]]
        assert row[4] is None
        assert row[9] is None
        assert row[10] is None
        assert "中文分类2待审核" in row[13]
        assert "中文描述待审核" in row[13]
        assert "中文产品详情待审核" in row[13]
    finally:
        workbook.close()


def test_zh_export_does_not_use_low_quality_model_result(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    record = _record("1001")
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", [record])
    source_hash = _write_dictionary(cfg["paths"]["dictionary_baseline"], record, product={
        "sku": "1001", "canonical_id": record["canonical_id"], "name_es_raw": record["name_es"],
        "name_zh_standard": "", "brand_id": "BrandX", "cat1_es": record["cat1_es"], "cat2_es": record["cat2_es"],
        "cat1_zh": "家务清洁", "cat2_zh": "清洁用品", "spec_es_raw": record["spec_es"],
        "spec_zh_standard": "", "source_hash": "stale", "translation_status": "NEEDS_REVIEW",
    })
    _write_dictionary(cfg["paths"]["dictionary_baseline"], record, product={
        "sku": "1001", "canonical_id": record["canonical_id"], "name_es_raw": record["name_es"],
        "name_zh_standard": "", "brand_id": "BrandX", "cat1_es": record["cat1_es"], "cat2_es": record["cat2_es"],
        "cat1_zh": "家务清洁", "cat2_zh": "清洁用品", "spec_es_raw": record["spec_es"],
        "spec_zh_standard": "", "source_hash": "stale", "translation_status": "NEEDS_REVIEW",
    }, model=[{
        "sku": "1001", "source_hash": source_hash, "name_zh_standard": "低质量模型名",
        "spec_zh_standard": "低质量规格", "quality_status": "NEEDS_REVIEW",
    }])

    result = export_catalog(cfg, language="zh", export_date="2026-08-24", no_images=True)
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        row = [cell.value for cell in workbook["商品全量"][2]]
        assert row[2] == record["name_es"]
        assert row[5] == record["spec_es"]
        assert "中文品名待审核" in row[13]
        assert "中文规格待审核" in row[13]
    finally:
        workbook.close()


def test_manifest_has_sku_set_hash_and_row_height_is_capped(tmp_path):
    cfg = _cfg(tmp_path)
    run_id = "2026-08-24_010000"
    record = _record("1001")
    record["desc_es"] = "Descripción larga " * 200
    _write_master(cfg["paths"]["master"], [_run_log(run_id, "2026-08-24")], [record])
    _write_snapshot(cfg["paths"]["snapshots"], run_id, "2026-08-24", [record])
    result = export_catalog(cfg, language="es", export_date="2026-08-24", no_images=True)
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert len(manifest["sku_set_hash"]) == 64
    assert manifest["validation_results"]["workbook"] == "PASS"
    workbook = openpyxl.load_workbook(result["output"], data_only=True)
    try:
        assert workbook["商品全量"].row_dimensions[2].height <= 405
    finally:
        workbook.close()
