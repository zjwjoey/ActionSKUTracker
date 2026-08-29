from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from action_tracker.database.connection import connect, transaction
from action_tracker.database.mirror import build_mirror
from action_tracker.database.schema import inspect_schema, migrate
from action_tracker.database import mirror as mirror_module
from action_tracker.database.repository import import_baseline
from action_tracker.database.validation import validate_mirror


def _sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[object]], header_row: int = 1):
    ws = wb.create_sheet(name)
    for _ in range(header_row - 1):
        ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return ws


def make_fixture(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _sheet(wb, "01_SKU_ZH_CURRENT", [
        "SKU", "中文品名", "一级类目（中文）", "二级类目（中文）", "规格（中文）",
        "中文描述", "中文产品详情", "翻译状态",
    ], [["100", "测试商品", "家居", "收纳", "1件", "描述", "详情", "CONFIRMED"]])
    _sheet(wb, "02_SKU_ES_CURRENT", [
        "Canonical_ID", "SKU", "西班牙语品名", "当前售价 (€)", "原价 (€)", "上次售价 (€)",
        "最近一次变价方向", "本期价格变化", "变化金额 (€)", "变化幅度 (%)", "最近变价日期",
        "历史最低价 (€)", "历史最高价 (€)", "一级类目（西语）", "二级类目（西语）",
        "规格（西语）", "单价", "新品", "促销", "可持续", "折扣", "原始标签", "当前状态",
        "首次发现日期", "最后确认存在日期", "描述（西语）", "产品详情（西语）", "商品链接", "图片链接", "匹配状态",
    ], [["ENT-1", "100", "Producto de prueba", 2.5, 3.0, 2.0, "UP", "本期上涨", 0.5, 25, "2026-08-29", 1.0, 3.0, "Hogar", "Almacenaje", "1 unidad", "2,50 €/ud", 0, 0, 0, 0, "", "CURRENT", "2026-01-01", "2026-08-29", "Descripción ES", "Detalle ES", "https://example/100", "https://example/image", "OK"]])
    _sheet(wb, "03_PRICE_HISTORY", ["Canonical_ID", "SKU", "日期", "旧售价 (€)", "新售价 (€)", "原价 (€)", "变化类型", "变化金额 (€)", "变化幅度 (%)", "促销状态", "来源文件", "来源Sheet"], [["ACT-1", "100", "2026-08-29", 2.0, 2.5, 3.0, "UP", 0.5, 25, "", "run.xlsx", "Sheet1"], ["ACT-X", "", "2026-04-05", None, 1.59, None, "NEW", None, None, "", "old.xlsx", "Sheet1"]])
    _sheet(wb, "04_EVENT_HISTORY", ["Canonical_ID", "SKU", "日期", "事件类型", "旧值", "新值", "来源文件", "备注"], [["ACT-1", "100", "2026-08-29", "PRICE_UP", "2", "2.5", "run.xlsx", "evidence"], ["ACT-X", "", "2026-04-05", "FIRST_SEEN", "", "", "old.xlsx", "pending"]])
    _sheet(wb, "05_RUN_LOG", ["Run ID", "运行日期", "开始时间", "结束时间", "Git Commit", "Sitemap SKU数", "Listing SKU数", "ACTIVE", "NEW", "REAPPEARED", "MISSING_FIRST", "MISSING_CONTINUED", "OFFLINE", "PRICE_UP", "PRICE_DOWN", "PROMO_START", "PROMO_END", "NEW_BADGE_ON", "NEW_BADGE_OFF", "CONTENT_CHANGE", "异常数量", "QA状态", "运行状态", "备注"], [["run-1", "2026-08-29", "2026-08-29T01:00:00", "2026-08-29T01:10:00", "abc", 2, 1, 1, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, "PASS", "FULL_COMMIT", ""]])
    _sheet(wb, "06_REVIEW_QUEUE", ["日期", "SKU", "问题类型", "证据", "候选值", "置信度", "建议动作", "人工备注"], [["2026-08-29", "100", "NAME", "evidence", "测试商品", 0.9, "REVIEW", ""]])
    _sheet(wb, "07_APRIL_ARCHIVE", ["四月归档ID", "正式SKU"], [["APR-1", "100"]])
    _sheet(wb, "08_LONG_TERM_MASTER", ["实体ID", "正式SKU", "四月归档ID", "身份类型", "匹配状态", "匹配置信度", "当前状态", "中文品名", "西班牙语品名", "一级类目（中文）", "一级类目（西语）", "规格（中文）", "规格（西语）", "当前售价 (€)", "历史最低价 (€)", "历史最高价 (€)", "首次观察日期", "最后观察日期", "四月原始记录数", "四月归档ID集合", "来源数", "来源工作表", "商品链接", "核对备注"], [["ENT-1", "100", "APR-1", "MATCHED", "MATCHED", 1.0, "CURRENT", "测试商品", "Producto de prueba", "家居", "Hogar", "1件", "1 unidad", 2.5, 1.0, 3.0, "2026-01-01", "2026-08-29", 1, "APR-1", 1, "01", "https://example/100", ""], ["ENT-2", "200", "", "NEW", "MATCHED", 1.0, "HISTORICAL", "历史商品", "Historico", "家居", "Hogar", "1件", "1 unidad", 1.0, 1.0, 1.0, "2026-01-01", "2026-01-02", 0, "", 1, "08", "", ""], ["ENT-3", "", "APR-3", "PENDING", "UNMATCHED", 0.0, "ARCHIVE_PENDING_MATCH", "", "", "", "", "", "", "", "", "2026-04-05", "2026-04-05", 1, "APR-3", 1, "07", "", ""]], header_row=7)
    _sheet(wb, "09_APRIL_MATCH_AUDIT", ["四月归档ID", "正式SKU"], [["APR-1", "100"]])
    _sheet(wb, "10_SOURCE_SCHEMA", ["日期", "文件名", "Sheet", "Raw 行数", "Raw 列数", "真实 Raw Schema", "来源作用", "数据状态", "备注"], [["2026-08-29", "run.xlsx", "Sheet1", 1, 1, "SKU", "事实", "OK", ""]])
    wb.save(path)


def test_schema_v1_and_transaction_rollback(tmp_path):
    db = tmp_path / "empty.db"
    migrate(db)
    with connect(db) as conn:
        assert conn.execute("PRAGMA foreign_keys").fetchone()[0] == 1
        names = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type IN ('table','view')")}
        assert {"products", "product_localizations", "observations", "price_history", "events", "runs", "reviews", "v_db_current_skus"} <= names
        with pytest.raises(RuntimeError):
            with transaction(conn) as tx:
                tx.execute("INSERT INTO products(sku, canonical_id, source_sheet, source_row_no, source_raw_json) VALUES ('x','x','T',1,'{}')")
                raise RuntimeError("rollback")
        assert conn.execute("SELECT COUNT(*) FROM products").fetchone()[0] == 0


def test_fixture_mirror_parity_and_master_unchanged(tmp_path):
    master = tmp_path / "Master.xlsx"
    db = tmp_path / "runtime" / "db" / "action_tracker.db"
    make_fixture(master)
    before = master.read_bytes()
    result = build_mirror(master, db, tmp_path / "reports")
    assert result["status"] == "PASS", result
    assert master.read_bytes() == before
    with connect(db, read_only=True) as conn:
        assert conn.execute("SELECT COUNT(*) FROM products").fetchone()[0] == 2
        assert conn.execute("SELECT COUNT(*) FROM v_db_current_skus").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM price_history").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM events").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM migration_source_issues").fetchone()[0] >= 3
        migration = conn.execute("SELECT status, validation_status, report_path FROM migration_runs").fetchone()
        assert migration[0] == "VALIDATED"
        assert migration[1] == "SQLITE MIRROR VALIDATED WITH SOURCE DATA ISSUES"
        assert migration[2]
    validation = validate_mirror(master, db)
    assert validation["status"] == "PASS", validation
    assert validation["checks"]["es_canonical_exact"] is True


def test_validation_rejects_current_set_mismatch(tmp_path):
    master = tmp_path / "Master.xlsx"
    db = tmp_path / "action.db"
    make_fixture(master)
    assert build_mirror(master, db, tmp_path / "reports")["status"] == "PASS"
    wb = load_workbook(master)
    ws = wb["02_SKU_ES_CURRENT"]
    ws["B2"] = "200"
    wb.save(master)
    result = validate_mirror(master, db)
    assert result["status"] == "FAIL"
    assert result["checks"]["es_db_current_equal"] is False


def test_failed_promotion_preserves_existing_mirror(tmp_path, monkeypatch):
    master = tmp_path / "Master.xlsx"
    db = tmp_path / "action.db"
    make_fixture(master)
    old_bytes = b"existing mirror placeholder"
    db.write_bytes(old_bytes)

    def refuse_replace(self, target):
        raise OSError("simulated promotion failure")

    monkeypatch.setattr(Path, "replace", refuse_replace)
    result = build_mirror(master, db, tmp_path / "reports")
    assert result["status"] == "FAIL"
    assert result["validation"]["rollback_preserved_old_mirror"] is True
    assert db.read_bytes() == old_bytes


@pytest.mark.parametrize(
    ("table", "column", "value"),
    [
        ("products", "current_price", 999.0),
        ("product_localizations", "name", "被篡改"),
        ("price_history", "new_price", 999.0),
        ("events", "event_type", "TAMPERED"),
        ("runs", "qa_status", "FAIL"),
        ("reviews", "evidence", "被篡改"),
    ],
)
def test_field_level_parity_rejects_tampering(tmp_path, table, column, value):
    master, db = tmp_path / "Master.xlsx", tmp_path / "action.db"
    make_fixture(master)
    assert build_mirror(master, db, tmp_path / "reports")["status"] == "PASS"
    key = "sku='100'" if table in {"products", "product_localizations", "price_history", "events"} else "run_id='run-1'" if table == "runs" else "review_id='MASTER06:2'"
    with connect(db) as conn:
        conn.execute(f"UPDATE {table} SET {column}=? WHERE {key}", (value,))
    result = validate_mirror(master, db)
    assert result["status"] == "FAIL"
    assert result["checks"]["field_parity"] is False
    parity_key = "localizations" if table == "product_localizations" else table
    assert result["field_parity"][parity_key]["mismatch_count"] > 0


def test_source_evidence_parity_rejects_hash_or_row_loss(tmp_path):
    master, db = tmp_path / "Master.xlsx", tmp_path / "action.db"
    make_fixture(master)
    assert build_mirror(master, db, tmp_path / "reports")["status"] == "PASS"
    conn = connect(db)
    try:
        conn.execute("UPDATE source_records SET raw_json='{}' WHERE source_sheet='08_LONG_TERM_MASTER' AND source_row_no=8")
        conn.commit()
    finally:
        conn.close()
    result = validate_mirror(master, db)
    assert result["status"] == "FAIL"
    assert result["checks"]["source_evidence_parity"] is False

    # Restore a clean mirror and remove one evidence row entirely.
    assert build_mirror(master, db, tmp_path / "reports")["status"] == "PASS"
    conn = connect(db)
    try:
        conn.execute("DELETE FROM source_records WHERE source_sheet='10_SOURCE_SCHEMA' AND source_row_no=2")
        conn.commit()
    finally:
        conn.close()
    result = validate_mirror(master, db)
    assert result["status"] == "FAIL"
    assert result["source_evidence_parity"]["exact_row_identity"] is False


def test_db_init_rejects_legacy_shape(tmp_path):
    legacy = tmp_path / "legacy.db"
    import sqlite3
    with sqlite3.connect(legacy) as conn:
        conn.execute("CREATE TABLE products (official_sku TEXT PRIMARY KEY)")
    assert inspect_schema(legacy) == "LEGACY"


def test_schema_identity_requires_family_version_and_all_v1_tables(tmp_path):
    shaped = tmp_path / "shaped.db"
    import sqlite3
    with sqlite3.connect(shaped) as conn:
        conn.execute("CREATE TABLE products (sku TEXT PRIMARY KEY, canonical_id TEXT, source_sheet TEXT, source_row_no INTEGER, source_raw_json TEXT, current_status_raw TEXT)")
    assert inspect_schema(shaped) == "LEGACY"

    v1 = tmp_path / "v1.db"
    migrate(v1)
    with sqlite3.connect(v1) as conn:
        conn.execute("DELETE FROM schema_metadata WHERE key='schema_family'")
        conn.commit()
    assert inspect_schema(v1) == "LEGACY"

    migrate(v1)
    with sqlite3.connect(v1) as conn:
        conn.execute("UPDATE schema_metadata SET value='0.9.0' WHERE key='schema_version'")
        conn.commit()
    assert inspect_schema(v1) == "LEGACY"


def test_db_init_is_idempotent_for_v1(tmp_path):
    db = tmp_path / "new.db"
    assert inspect_schema(db) == "NEW"
    migrate(db)
    assert inspect_schema(db) == "V1"
    migrate(db)
    assert inspect_schema(db) == "V1"


def test_import_baseline_preserves_multi_sku_identity(tmp_path):
    db = tmp_path / "baseline.db"
    count = import_baseline(
        db,
        {"100": {"name_es": "A"}, "200": {"name_es": "B"}, "300": {"name_es": "C"}},
        "2026-08-29",
    )
    assert count == 3
    with connect(db, read_only=True) as conn:
        assert conn.execute("SELECT COUNT(*) FROM products").fetchone()[0] == 3
        assert conn.execute("SELECT COUNT(DISTINCT source_row_no) FROM products WHERE source_sheet='BASELINE'").fetchone()[0] == 3


def test_import_baseline_keeps_source_rows_when_input_order_changes(tmp_path):
    db = tmp_path / "baseline.db"
    import_baseline(db, {"100": {"name_es": "A"}, "200": {"name_es": "B"}}, "2026-08-29")
    conn = connect(db, read_only=True)
    try:
        before = dict(conn.execute("SELECT sku,source_row_no FROM products WHERE source_sheet='BASELINE'").fetchall())
    finally:
        conn.close()
    import_baseline(db, {"300": {"name_es": "C"}, "100": {"name_es": "A"}, "200": {"name_es": "B"}}, "2026-08-29")
    conn = connect(db, read_only=True)
    try:
        after = dict(conn.execute("SELECT sku,source_row_no FROM products WHERE source_sheet='BASELINE'").fetchall())
    finally:
        conn.close()
    assert after["100"] == before["100"]
    assert after["200"] == before["200"]
    assert after["300"] > max(before.values())


def test_final_master_hash_gate_preserves_old_mirror(tmp_path, monkeypatch):
    master, db = tmp_path / "Master.xlsx", tmp_path / "action.db"
    make_fixture(master)
    old_bytes = b"old mirror"
    db.write_bytes(old_bytes)
    original_hash = mirror_module.sha256_file(master)
    calls = iter([original_hash, original_hash, "changed-before-promotion"])
    monkeypatch.setattr(mirror_module, "sha256_file", lambda _: next(calls))
    result = build_mirror(master, db, tmp_path / "reports")
    assert result["status"] == "FAIL"
    assert result["validation"]["failure_reason"] == "MASTER_CHANGED_BEFORE_MIRROR_PROMOTION"
    assert db.read_bytes() == old_bytes


def test_post_promotion_validation_failure_restores_backup(tmp_path, monkeypatch):
    master, db = tmp_path / "Master.xlsx", tmp_path / "action.db"
    make_fixture(master)
    old_bytes = b"old mirror"
    db.write_bytes(old_bytes)
    monkeypatch.setattr(mirror_module, "_post_promotion_validate", lambda *_: {"status": "FAIL", "checks": {"integrity_check": False}})
    result = build_mirror(master, db, tmp_path / "reports")
    assert result["status"] == "FAIL"
    assert result["validation"]["failure_reason"] == "POST_PROMOTION_VALIDATION_FAILED"
    assert result["validation"]["rollback_restored_old_mirror"] is True
    assert db.read_bytes() == old_bytes


def test_success_report_contains_final_hash_and_post_promotion_validation(tmp_path):
    master, db = tmp_path / "Master.xlsx", tmp_path / "action.db"
    make_fixture(master)
    result = build_mirror(master, db, tmp_path / "reports")
    assert result["status"] == "PASS"
    import json
    report_dir = Path(result["report_dir"])
    migration_report = json.loads((report_dir / "migration_report.json").read_text(encoding="utf-8"))
    validation_report = json.loads((report_dir / "validation_report.json").read_text(encoding="utf-8"))
    assert migration_report["final_master_hash"] == validation_report["final_master_hash"]
    assert validation_report["final_master_hash"] == validation_report["master_hash_before"]
    assert validation_report["post_promotion_validation"]["status"] == "PASS"


def test_provenance_tampering_fails_field_parity(tmp_path):
    master, db = tmp_path / "Master.xlsx", tmp_path / "action.db"
    make_fixture(master)
    assert build_mirror(master, db, tmp_path / "reports")["status"] == "PASS"
    conn = connect(db)
    try:
        conn.execute("UPDATE products SET source_row_no=999 WHERE sku='100'")
        conn.commit()
    finally:
        conn.close()
    result = validate_mirror(master, db)
    assert result["status"] == "FAIL"
    assert result["field_parity"]["products"]["status"] == "FAIL"


def test_repeat_mirror_business_rows_are_deterministic(tmp_path):
    master = tmp_path / "Master.xlsx"
    make_fixture(master)
    db_a, db_b = tmp_path / "a.db", tmp_path / "b.db"
    assert build_mirror(master, db_a, tmp_path / "reports-a")["status"] == "PASS"
    assert build_mirror(master, db_b, tmp_path / "reports-b")["status"] == "PASS"
    with connect(db_a, read_only=True) as a, connect(db_b, read_only=True) as b:
        for query in (
            "SELECT sku,canonical_id,name_es,current_price,current_status_raw FROM products ORDER BY sku",
            "SELECT sku,language,name,cat1,cat2,spec,description,details FROM product_localizations ORDER BY sku,language",
            "SELECT sku,canonical_id,observed_at,previous_price,new_price,original_price,change_type,promotion_raw,raw_json FROM price_history ORDER BY id",
            "SELECT sku,canonical_id,occurred_at,event_type,old_value,new_value,evidence FROM events ORDER BY id",
        ):
            assert [tuple(row) for row in a.execute(query)] == [tuple(row) for row in b.execute(query)]


def test_migration_rejects_missing_sheet_and_required_column(tmp_path):
    master = tmp_path / "Master.xlsx"
    make_fixture(master)
    wb = load_workbook(master)
    del wb["10_SOURCE_SCHEMA"]
    wb.save(master)
    with pytest.raises(ValueError, match="Missing required Master sheets"):
        build_mirror(master, tmp_path / "missing-sheet.db", tmp_path / "reports")

    make_fixture(master)
    wb = load_workbook(master)
    ws = wb["02_SKU_ES_CURRENT"]
    ws.delete_cols(2)  # remove the required SKU column
    wb.save(master)
    with pytest.raises(ValueError, match="Missing required column 'SKU'"):
        build_mirror(master, tmp_path / "missing-column.db", tmp_path / "reports")


@pytest.mark.parametrize("duplicate_field", ["正式SKU", "实体ID"])
def test_migration_rejects_duplicate_product_identity(tmp_path, duplicate_field):
    master = tmp_path / "Master.xlsx"
    make_fixture(master)
    wb = load_workbook(master)
    ws = wb["08_LONG_TERM_MASTER"]
    ws.cell(row=9, column=2 if duplicate_field == "正式SKU" else 1).value = ws.cell(row=8, column=2 if duplicate_field == "正式SKU" else 1).value
    wb.save(master)
    with pytest.raises(Exception):
        build_mirror(master, tmp_path / f"duplicate-{duplicate_field}.db", tmp_path / "reports")


def test_schema_rejects_orphan_foreign_keys(tmp_path):
    db = tmp_path / "fk.db"
    migrate(db)
    with connect(db) as conn:
        with pytest.raises(Exception):
            conn.execute("INSERT INTO product_localizations(sku,language,source,review_status,source_sheet,source_row_no) VALUES ('missing','zh','T','PENDING','T',1)")
        with pytest.raises(Exception):
            conn.execute("INSERT INTO observations(run_id,sku,observation_date,presence,observation_complete,raw_json) VALUES ('missing','missing','2026-08-29',1,1,'{}')")
