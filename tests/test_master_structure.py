"""规范 §十 / §五十 测试：REVIEW_QUEUE 统一入口、事件来源列、旧版审计表迁移。"""
import datetime as dt
from pathlib import Path

import openpyxl
import pytest

from action_tracker.excel import writer
from action_tracker.services.change import compute_changes
from action_tracker.services.review import REVIEW_HEADERS, add_review_item


def _cfg(tmp_path: Path, with_state: bool = True) -> dict:
    cfg = {
        "paths": {
            "master": tmp_path / "master" / "Action_Master.xlsx",
            "backups": tmp_path / "backups",
            "temp": tmp_path / "temp",
        }
    }
    if with_state:
        cfg["paths"]["state"] = tmp_path / "state"
    return cfg


def _build_master(path: Path, legacy: bool = False):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "01_SKU_ZH_CURRENT"
    ws.append(["Canonical_ID", "SKU", "中文品名", "当前售价 (€)", "最后确认存在日期"])
    ws.append(["ACT0001001", "1001", "品", 2.99, "2026-08-10"])
    ws2 = wb.create_sheet("02_SKU_ES_CURRENT")
    ws2.append(["Canonical_ID", "SKU", "西班牙语品名", "当前售价 (€)", "最后确认存在日期"])
    ws2.append(["ACT0001001", "1001", "Nombre", 2.99, "2026-08-10"])
    ws3 = wb.create_sheet("03_PRICE_HISTORY")
    ws3.append(writer.PRICE_HISTORY_HEADERS)
    ws4 = wb.create_sheet("04_EVENT_HISTORY")
    ws4.append(writer.EVENT_HISTORY_HEADERS)
    if legacy:
        ws5 = wb.create_sheet("05_MAPPING_REVIEW")
        ws5.append(["历史日期", "来源文件", "匹配状态", "人工备注"])
        ws5.append(["2026-08-09", "a.xlsx", "已匹配", "n1"])
        ws5.append(["2026-08-09", "b.xlsx", "冲突", "n2"])
        ws6 = wb.create_sheet("06_SOURCE_SUMMARY")
        ws6.append(["日期", "文件名", "行数", "备注"])
        ws6.append(["2026-08-09", "a.xlsx", 100, "ok"])
    wb.save(path)


# ---- add_review_item：统一中文 key + 问题类型校验 ----
def test_review_item_keys_match_headers():
    item = add_review_item("2026-08-10", "1001", "PRICE_ANOMALY", evidence="price=99",
                           candidates=[3.99], confidence=0.9)
    assert list(item.keys()) == REVIEW_HEADERS
    assert item["日期"] == "2026-08-10"
    assert item["SKU"] == "1001"
    assert item["问题类型"] == "PRICE_ANOMALY"
    assert item["候选值"] == [3.99]
    assert item["置信度"] == 0.9
    assert item["建议动作"] == "人工核对"


def test_review_item_invalid_type_raises():
    with pytest.raises(ValueError):
        add_review_item("2026-08-10", "1001", "NOT_A_TYPE")


def test_review_issue_types_cover_source_flags():
    from action_tracker.services.review import REVIEW_ISSUE_TYPES
    for t in ("SITEMAP_ONLY", "LISTING_ONLY", "UNKNOWN", "PRICE_ANOMALY"):
        assert t in REVIEW_ISSUE_TYPES


# ---- change.py：事件 dict 与 03/04 表头对齐 ----
def test_change_price_event_has_source_columns():
    before = {"sku": "1001", "canonical_id": "ACT0001001", "current_price": 4.99, "raw_tags": ""}
    after = {"sku": "1001", "canonical_id": "ACT0001001", "current_price": 3.99, "raw_tags": ""}
    out = compute_changes("1001", "ACT0001001", before, after, "2026-08-10", 0.01, 1000)
    assert len(out.price_events) == 1
    ev = out.price_events[0]
    assert "来源文件" in ev and ev["来源文件"] == "Action_Master.xlsx"
    assert "来源Sheet" in ev and ev["来源Sheet"] == "03_PRICE_HISTORY"
    # 与 03 sheet 表头完全对齐
    assert list(ev.keys()) == writer.PRICE_HISTORY_HEADERS


def test_change_badge_event_has_source_and_note():
    before = {"sku": "1001", "canonical_id": "ACT0001001", "current_price": 3.99, "raw_tags": ""}
    after = {"sku": "1001", "canonical_id": "ACT0001001", "current_price": 3.99, "raw_tags": "Nuevo"}
    out = compute_changes("1001", "ACT0001001", before, after, "2026-08-10", 0.01, 1000, run_id="R1")
    assert len(out.badge_events) == 1
    ev = out.badge_events[0]
    assert ev["事件类型"] == "ACTION_NEW_BADGE_ON"
    assert ev["来源文件"] == "Action_Master.xlsx"
    assert ev["备注"] == "R1"
    assert list(ev.keys()) == writer.EVENT_HISTORY_HEADERS


# ---- writer：旧版审计表迁移（行数验证 + 幂等） ----
def test_writer_migrates_legacy_sheets(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"], legacy=True)
    writer.write_master(cfg, updated_records={}, price_events=[], event_events=[], dry_run=False)

    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    assert set(wb.sheetnames) == {"01_SKU_ZH_CURRENT", "02_SKU_ES_CURRENT", "03_PRICE_HISTORY",
                                  "04_EVENT_HISTORY", "05_RUN_LOG", "06_REVIEW_QUEUE"}
    # 05/06 已是新表（含表头）
    assert [c.value for c in wb["05_RUN_LOG"][1]] == writer.RUN_LOG_HEADERS
    assert [c.value for c in wb["06_REVIEW_QUEUE"][1]] == REVIEW_HEADERS
    wb.close()

    csv_path = cfg["paths"]["state"] / "legacy_mapping_review.csv"
    assert csv_path.exists()
    import csv as _csv
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(_csv.reader(f))
    assert len(rows) - 1 == 2  # 表头 + 2 行审计数据
    assert rows[1][0] == "2026-08-09" and rows[2][2] == "冲突"
    assert (cfg["paths"]["backups"] / "legacy_source_summary.csv").exists()


def test_legacy_master_writer_is_blocked_in_sqlite_primary(tmp_path):
    cfg = _cfg(tmp_path)
    cfg["storage"] = {"mode": "SQLITE_PRIMARY"}
    _build_master(cfg["paths"]["master"])
    with pytest.raises(RuntimeError, match="LEGACY_MASTER_WRITER_BLOCKED"):
        writer.stage_master(cfg, updated_records={}, price_events=[], event_events=[])


def test_writer_migrate_idempotent(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"], legacy=True)
    writer.write_master(cfg, updated_records={}, price_events=[], event_events=[], dry_run=False)
    csv_path = cfg["paths"]["state"] / "legacy_mapping_review.csv"
    before = csv_path.read_bytes()
    # 第二次写：表已删，不应重复归档/重复行
    writer.write_master(cfg, updated_records={}, price_events=[], event_events=[], dry_run=False)
    assert csv_path.read_bytes() == before


# ---- writer：05_RUN_LOG / 06_REVIEW_QUEUE 写入 ----
def test_writer_run_log_and_review(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    run_log_row = {h: (h if h != "Git Commit" else "abc123") for h in writer.RUN_LOG_HEADERS}
    review_rows = [add_review_item("2026-08-10", "1001", "SITEMAP_ONLY", evidence="sitemap有 listing无")]
    writer.write_master(cfg, updated_records={}, price_events=[], event_events=[],
                        run_log_row=run_log_row, review_rows=review_rows, dry_run=False)
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    rl = list(wb["05_RUN_LOG"].iter_rows(values_only=True))
    assert len(rl) == 2  # 表头 + 1 行
    assert rl[1][0] == "Run ID" and rl[1][4] == "abc123"  # Run ID / Git Commit 列
    rq = list(wb["06_REVIEW_QUEUE"].iter_rows(values_only=True))
    assert len(rq) == 2
    d = rq[1][0]
    assert (d.year, d.month, d.day) == (2026, 8, 10)  # _cell 已转真日期
    assert rq[1][2] == "SITEMAP_ONLY"
    wb.close()


# ---- writer：追加不重复表头 ----
def test_append_rows_no_duplicate_header(tmp_path):
    cfg = _cfg(tmp_path)
    _build_master(cfg["paths"]["master"])
    ev = {"Canonical_ID": "ACT0001001", "SKU": "1001", "日期": "2026-08-10",
          "旧售价 (€)": 2.99, "新售价 (€)": 3.99, "原价 (€)": None, "变化类型": "UP",
          "变化金额 (€)": 1.0, "变化幅度 (%)": 33.4, "促销状态": None,
          "来源文件": "Action_Master.xlsx", "来源Sheet": "03_PRICE_HISTORY"}
    writer.write_master(cfg, updated_records={}, price_events=[ev], event_events=[], dry_run=False)
    writer.write_master(cfg, updated_records={}, price_events=[ev], event_events=[], dry_run=False)
    wb = openpyxl.load_workbook(cfg["paths"]["master"])
    rows = list(wb["03_PRICE_HISTORY"].iter_rows(values_only=True))
    wb.close()
    assert rows[0][0] == "Canonical_ID"  # 只有一条表头
    assert len(rows) == 1 + 2  # 表头 + 2 次追加
